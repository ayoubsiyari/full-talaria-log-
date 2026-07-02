#!/usr/bin/env python3
"""
Generate dashboard QA sample sessions as an Excel workbook (one sheet per session).

Each sheet uses the same trade column schema as talaria-test_2-trades-2026-06-20.csv.
80% of trades are realistic; 20% include tail / extreme scenarios.

Usage:
  py chart/scripts/generate_dashboard_session_samples.py
  py chart/scripts/generate_dashboard_session_samples.py /path/to/output.xlsx
"""

from __future__ import annotations

import csv
import json
import random
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

# Column order copied from talaria-test_2-trades-2026-06-20.csv
COLUMNS: list[str] = [
    "journal_trade_id", "trade_id", "client_trade_id", "tradeId", "id", "n",
    "sourceSessionName", "setup", "strategy_id", "strategyName", "symbol", "ticker", "direction", "side", "type",
    "orderType", "quantity", "status", "entryTime", "openTime", "entryDate", "date",
    "exitTime", "closeTime", "exitDate", "entryPrice", "openPrice", "exitPrice",
    "closePrice", "stopLoss", "takeProfit", "pnl", "pnl_dollars_net", "realizedPnL",
    "rMultiple", "actual_rr_net", "actualRR", "rewardToRiskRatio", "riskAmount",
    "riskPerTrade", "plannedRR", "duration", "closeType", "mfe", "mae", "mfe_r",
    "mae_r", "total_mfe_r", "highestPrice", "lowestPrice", "commission_total",
    "commission_at_entry", "spread_pips_at_entry", "postTradeNotes", "preTags",
    "postTags", "tags", "strategy_variables", "post_strategy_variables",
    "partialCloses", "entryScreenshot", "exitScreenshot", "railScreenshots",
    "sourceSessionId", "trading_session_id", "savedAt", "active_sl_at_exit",
    "active_tps_at_exit", "actual_risk_r", "actual_rr_gross", "array_base_price",
    "balance_at_creation", "balance_at_exit", "bar_close_r", "bar_high_r",
    "bar_low_r", "capture_ratio", "chart_trade_id", "dayOfWeek", "entries_locked",
    "entryMarkerTimeMs", "entry_offset_r", "exit_confirmed", "exit_timing_gap",
    "finalClosePnL", "final_exit_bar", "hasMultipleTakeProfits", "hasPartialCloses",
    "holdingTimeDays", "holdingTimeHours", "holdingTimeMs", "hourOfEntry",
    "hourOfExit", "initial_sl", "initial_takeProfit", "isScaledTrade",
    "isSplitEntry", "maeTime", "management_gap", "market", "mfeTime", "month",
    "multiTpSnapshot", "netPnL", "originalRiskAmount", "partialClosePnL",
    "pip_value_at_entry", "plannedEntrySnapshot", "plannedRRAtEntry",
    "plannedTpSnapshot", "planned_risk_pct", "pnl_dollars_gross", "post_checkpoints",
    "post_exit_anchor_time", "post_exit_bar_close_r", "post_exit_bar_high_r",
    "post_exit_bar_low_r", "preTradeNotes", "rulesFollowed", "session",
    "sl_modifications", "sourceFileId", "sourceFilterKey", "sourceKey",
    "sourceLabel", "sourceType", "splitGroupId", "splitIndex", "splitTotal",
    "tag", "total_bars_held", "trail_sl_path", "v9PostTradeTags", "v9TradeNotes",
    "would_have_won", "year",
    # Live-dashboard extras (not in export header but useful for QA)
    "accountType", "planAdherence", "demons", "originSource", "session_mode",
    "category_sheet",
]

INSTRUMENTS: dict[str, dict] = {
    "EURUSD": {"market": "Forex", "pip": 0.0001, "base": 1.085, "pip_value": 10, "spread": 0.8},
    "GBPUSD": {"market": "Forex", "pip": 0.0001, "base": 1.265, "pip_value": 10, "spread": 1.2},
    "USDJPY": {"market": "Forex", "pip": 0.01, "base": 148.5, "pip_value": 9.2, "spread": 0.9},
    "XAUUSD": {"market": "Commodity", "pip": 0.01, "base": 2035.0, "pip_value": 1.0, "spread": 2.5},
    "NAS100": {"market": "Indices", "pip": 1.0, "base": 16850.0, "pip_value": 1.0, "spread": 1.5},
    "BTCUSD": {"market": "Crypto", "pip": 1.0, "base": 52000.0, "pip_value": 1.0, "spread": 12.0},
    "ES": {"market": "Futures", "pip": 0.25, "base": 5280.0, "pip_value": 12.5, "spread": 0.5},
    "NQ": {"market": "Futures", "pip": 0.25, "base": 18450.0, "pip_value": 5.0, "spread": 0.75},
}

SETUPS = [
    "Breakout", "Pullback", "Reversal", "Range Fade", "Trend Continuation",
    "Liquidity Sweep", "Opening Range", "VWAP Bounce", "News Fade", "General",
]

CLOSE_TYPES = ["TP", "SL", "Manual", "Trailing SL", "BE", "Partial", "Time Stop", "Gap Exit"]

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

LIVE_DEMONS = ["revenge", "fomo", "overtrade", "early_exit", "late_entry", "size_up"]

# One workbook tab per dashboard source family.
CATEGORY_SHEETS: dict[str, str] = {
    "standard_backtest": "Standard Backtest",
    "prop_backtest": "Prop",
    "live_prop": "Prop",
    "live_journal": "Journal",
}


def _default_out() -> Path:
    root = Path(__file__).resolve().parents[3]
    return root / "docs" / "dashboard-session-samples-2026-06-20.xlsx"


def _iso_z(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")


def _price_fmt(value: float, pip: float) -> float:
    if pip >= 1:
        return round(value, 2)
    if pip >= 0.01:
        return round(value, 3 if pip == 0.01 else 5)
    return round(value, 5)


def _side(direction: str) -> str:
    return "Long" if direction == "BUY" else "Short"


def _empty_row() -> dict[str, str | int | float | bool]:
    return {c: "" for c in COLUMNS}


def _session_specs() -> list[dict]:
    specs: list[dict] = []
    std_names = [
        "EURUSD Breakout Lab", "GBPUSD Pullback Study", "Multi-Pair Swing",
        "Gold Momentum", "Nasdaq Scalper", "Crypto Range", "Futures ORB",
        "JPY Carry Fade", "Mixed FX Portfolio", "Strategy Validation",
    ]
    prop_bt_names = [
        "FTMO 100K Phase 1", "FTMO 100K Phase 2", "FundedNext 50K",
        "The5ers Bootcamp", "Topstep 50K Combine", "MyForexFunds 10K",
        "Apex 25K Eval", "2-Step Forex Challenge", "Instant Funded Sim",
        "Strict DD Prop Sim",
    ]
    live_jrn_names = [
        "OANDA Personal FX", "MT5 Swing Account", "Binance Spot Journal",
        "Bybit Perps Log", "Multi-Broker Aggregate", "EUR Scalping Live",
        "Commodity Live Log", "Indices Day Journal", "Weekend Crypto Log",
        "Discipline Tracker Live",
    ]
    live_prop_names = [
        "Funded FTMO Live", "FundedNext Live", "Topstep Funded",
        "Prop Challenge Live", "2-Step Funded Account", "Instant Funded Live",
        "Strict Rules Live", "Recovery Phase Live", "Near Breach Live",
        "Consistency Rule Live",
    ]

    for i, name in enumerate(std_names, start=1):
        specs.append({
            "category_sheet": CATEGORY_SHEETS["standard_backtest"],
            "session_id": 1000 + i,
            "name": name,
            "sourceType": "backtest",
            "session_type": "personal",
            "session_mode": "standard_backtest",
            "accountType": "private",
            "start_balance": 10000 + i * 500,
            "trade_count": random.randint(200, 800),
            "instruments": list(INSTRUMENTS.keys())[:4 + (i % 4)],
            "win_rate": 0.48 + (i % 7) * 0.02,
        })

    for i, name in enumerate(prop_bt_names, start=1):
        specs.append({
            "category_sheet": CATEGORY_SHEETS["prop_backtest"],
            "session_id": 2000 + i,
            "name": name,
            "sourceType": "backtest",
            "session_type": "propfirm",
            "session_mode": "prop_backtest",
            "accountType": "prop",
            "start_balance": [10000, 25000, 50000, 100000, 200000][i % 5],
            "trade_count": random.randint(50, 200),
            "instruments": list(INSTRUMENTS.keys())[:2 + (i % 3)],
            "win_rate": 0.45 + (i % 5) * 0.02,
            "max_daily_loss_pct": 5.0,
            "max_dd_pct": 10.0,
        })

    for i, name in enumerate(live_jrn_names, start=1):
        specs.append({
            "category_sheet": CATEGORY_SHEETS["live_journal"],
            "session_id": 3000 + i,
            "name": name,
            "sourceType": "journal",
            "session_type": "personal",
            "session_mode": "live_journal",
            "accountType": "private",
            "start_balance": 5000 + i * 1000,
            "trade_count": random.randint(100, 1000),
            "instruments": list(INSTRUMENTS.keys())[:3 + (i % 5)],
            "win_rate": 0.50 + (i % 4) * 0.015,
        })

    for i, name in enumerate(live_prop_names, start=1):
        specs.append({
            "category_sheet": CATEGORY_SHEETS["live_prop"],
            "session_id": 4000 + i,
            "name": name,
            "sourceType": "journal",
            "session_type": "propfirm",
            "session_mode": "live_prop",
            "accountType": "prop",
            "start_balance": [10000, 50000, 100000][i % 3],
            "trade_count": random.randint(20, 500),
            "instruments": list(INSTRUMENTS.keys())[:2 + (i % 4)],
            "win_rate": 0.46 + (i % 6) * 0.02,
            "max_daily_loss_pct": 4.0 + (i % 3),
            "max_dd_pct": 8.0 + (i % 4),
        })

    return specs


def _pick_scenario(rng: random.Random, is_tail: bool) -> str:
    normal = [
        "win_tp", "loss_sl", "win_manual", "loss_manual", "breakeven",
        "partial_win", "trailing_win", "small_loss", "planned_rr_hit",
    ]
    tails = [
        "mega_win", "mega_loss", "left_on_table", "survived_drawdown",
        "oversized_lot", "micro_lot", "gap_loss", "news_spike_win",
        "multi_day_hold", "split_entry", "scaled_in", "zero_pnl",
        "rule_break", "missed_plan", "out_of_plan",
    ]
    if is_tail:
        return rng.choice(tails + normal[:2])
    return rng.choice(normal)


def _build_trade(
    rng: random.Random,
    spec: dict,
    trade_num: int,
    entry_dt: datetime,
    balance: float,
    is_tail: bool,
) -> tuple[dict, float, datetime]:
    row = _empty_row()
    ticker = rng.choice(spec["instruments"])
    inst = INSTRUMENTS[ticker]
    scenario = _pick_scenario(rng, is_tail)

    direction = rng.choice(["BUY", "SELL"])
    setup = rng.choice(SETUPS)
    risk_pct = rng.uniform(0.5, 1.5)
    risk_amount = balance * (risk_pct / 100.0)
    if scenario == "oversized_lot":
        risk_amount = balance * rng.uniform(0.08, 0.15)
    elif scenario == "micro_lot":
        risk_amount = balance * rng.uniform(0.001, 0.01)

    planned_rr = rng.choice([1.0, 1.5, 2.0, 2.5, 3.0])
    if scenario in ("mega_win", "news_spike_win"):
        planned_rr = rng.uniform(4.0, 8.0)

    pip = inst["pip"]
    entry = inst["base"] * (1 + rng.uniform(-0.03, 0.03))
    entry = _price_fmt(entry, pip)
    sl_dist = rng.uniform(8, 35) * pip
    if scenario == "mega_loss":
        sl_dist = rng.uniform(2, 5) * pip
    if direction == "BUY":
        stop = _price_fmt(entry - sl_dist, pip)
        target = _price_fmt(entry + sl_dist * planned_rr, pip)
    else:
        stop = _price_fmt(entry + sl_dist, pip)
        target = _price_fmt(entry - sl_dist * planned_rr, pip)

    hold_minutes = rng.randint(15, 480)
    if scenario == "multi_day_hold":
        hold_minutes = rng.randint(2880, 10080)
    elif scenario in ("mega_win", "mega_loss", "gap_loss"):
        hold_minutes = rng.randint(5, 45)

    exit_dt = entry_dt + timedelta(minutes=hold_minutes)
    close_type = rng.choice(CLOSE_TYPES)

    win = rng.random() < spec["win_rate"]
    if scenario == "mega_win":
        win = True
        close_type = "TP"
    elif scenario == "mega_loss":
        win = False
        close_type = "SL"
    elif scenario == "breakeven" or scenario == "zero_pnl":
        win = False
        close_type = "BE"
    elif scenario == "left_on_table":
        win = False
        close_type = "Manual"
    elif scenario == "survived_drawdown":
        win = True
        close_type = "Trailing SL"

    if win:
        r_mult = rng.uniform(0.35, min(planned_rr * 1.1, 3.5))
        if scenario == "mega_win":
            r_mult = rng.uniform(5.0, 12.0)
        elif scenario == "news_spike_win":
            r_mult = rng.uniform(3.5, 7.0)
        elif scenario == "partial_win":
            r_mult = rng.uniform(0.5, 1.2)
        if close_type == "TP":
            exit_price = target
        else:
            move = sl_dist * r_mult
            exit_price = _price_fmt(entry + move if direction == "BUY" else entry - move, pip)
    else:
        r_mult = -rng.uniform(0.25, 1.05)
        if scenario in ("mega_loss", "gap_loss"):
            r_mult = -rng.uniform(2.5, 6.0)
        elif scenario == "breakeven" or scenario == "zero_pnl":
            r_mult = 0.0
        elif scenario == "left_on_table":
            r_mult = -rng.uniform(0.15, 0.45)
        if close_type == "SL" and scenario not in ("breakeven", "zero_pnl"):
            exit_price = stop
        elif scenario in ("breakeven", "zero_pnl"):
            exit_price = entry
        else:
            move = sl_dist * abs(r_mult)
            exit_price = _price_fmt(entry - move if direction == "BUY" else entry + move, pip)

    pnl = round(r_mult * risk_amount, 2)
    if scenario == "zero_pnl":
        pnl = 0.0
        r_mult = 0.0

    mae_r = abs(rng.uniform(0.05, 1.2 if not win else 0.8))
    mfe_r = abs(rng.uniform(0.1, 2.8))
    if not win and scenario == "left_on_table":
        mfe_r = rng.uniform(2.5, 6.0)
        mae_r = rng.uniform(0.8, 1.5)
    if win:
        mae_r = min(mae_r, abs(r_mult) * 0.9)
    else:
        mfe_r = min(mfe_r, 1.2)

    qty = round(risk_amount / max(sl_dist / pip * inst["pip_value"], 1), 2)
    if scenario == "oversized_lot":
        qty = round(qty * rng.uniform(2.5, 5.0), 2)
    elif scenario == "micro_lot":
        qty = round(max(0.01, qty * rng.uniform(0.05, 0.2)), 2)

    has_partial = scenario == "partial_win" or (is_tail and rng.random() < 0.3)
    has_multi_tp = is_tail and rng.random() < 0.15
    is_split = scenario == "split_entry" or (is_tail and rng.random() < 0.1)
    is_scaled = scenario == "scaled_in" or (is_tail and rng.random() < 0.1)
    rules_followed = scenario not in ("rule_break", "missed_plan", "out_of_plan")
    plan_adherence = "according-to-plan"
    if scenario == "out_of_plan":
        plan_adherence = "out-of-plan"
    elif scenario == "missed_plan":
        plan_adherence = "missed-trade"
    elif not rules_followed:
        plan_adherence = "out-of-plan"

    post_tag = "Win" if pnl > 0 else "Loss" if pnl < 0 else "BE"
    entry_ms = int(entry_dt.timestamp() * 1000)
    exit_ms = int(exit_dt.timestamp() * 1000)
    holding_ms = exit_ms - entry_ms
    holding_hours = round(holding_ms / 3600000, 4)
    holding_days = round(holding_hours / 24, 4)

    if direction == "BUY":
        highest = _price_fmt(max(entry, exit_price) + rng.uniform(0, 3) * pip, pip)
        lowest = _price_fmt(min(entry, exit_price) - rng.uniform(0, 3) * pip, pip)
        mfe_price = _price_fmt(entry + mfe_r * sl_dist, pip)
        mae_price = _price_fmt(entry - mae_r * sl_dist, pip)
    else:
        highest = _price_fmt(max(entry, exit_price) + rng.uniform(0, 3) * pip, pip)
        lowest = _price_fmt(min(entry, exit_price) - rng.uniform(0, 3) * pip, pip)
        mfe_price = _price_fmt(entry - mfe_r * sl_dist, pip)
        mae_price = _price_fmt(entry + mae_r * sl_dist, pip)

    balance_after = round(balance + pnl, 2)
    trade_id = spec["session_id"] * 10000 + trade_num
    client_id = trade_num
    capture = abs(r_mult) / mfe_r if mfe_r else 0

    demons = []
    if plan_adherence == "out-of-plan" and rng.random() < 0.7:
        demons = rng.sample(LIVE_DEMONS, k=rng.randint(1, 2))

    row.update({
        "journal_trade_id": trade_id,
        "trade_id": trade_id,
        "client_trade_id": client_id,
        "tradeId": trade_id,
        "id": trade_id,
        "n": trade_num,
        "sourceSessionName": spec["name"],
        "setup": setup,
        "symbol": ticker,
        "ticker": ticker,
        "direction": direction,
        "side": _side(direction),
        "type": direction,
        "orderType": "market",
        "quantity": qty,
        "status": "closed",
        "entryTime": entry_ms,
        "openTime": entry_ms,
        "entryDate": _iso_z(entry_dt),
        "date": entry_dt.strftime("%Y-%m-%d"),
        "exitTime": exit_ms,
        "closeTime": _iso_z(exit_dt),
        "exitDate": _iso_z(exit_dt),
        "entryPrice": entry,
        "openPrice": entry,
        "exitPrice": exit_price,
        "closePrice": exit_price,
        "stopLoss": stop,
        "takeProfit": target,
        "pnl": pnl,
        "pnl_dollars_net": pnl,
        "realizedPnL": pnl,
        "rMultiple": round(r_mult, 4),
        "actual_rr_net": round(abs(r_mult), 4),
        "actualRR": round(abs(r_mult), 4),
        "rewardToRiskRatio": round(planned_rr, 2),
        "riskAmount": round(risk_amount, 2),
        "riskPerTrade": round(risk_amount, 2),
        "plannedRR": planned_rr,
        "duration": max(1, hold_minutes),
        "closeType": close_type,
        "mfe": mfe_price,
        "mae": mae_price,
        "mfe_r": round(mfe_r, 4),
        "mae_r": round(mae_r, 4),
        "total_mfe_r": round(mfe_r, 4),
        "highestPrice": highest,
        "lowestPrice": lowest,
        "commission_total": round(rng.uniform(0, 4), 2),
        "commission_at_entry": round(rng.uniform(0, 3), 2),
        "spread_pips_at_entry": inst["spread"],
        "postTradeNotes": json.dumps({"tags": post_tag, "scenario": scenario}),
        "preTags": json.dumps([setup]),
        "postTags": json.dumps([post_tag]),
        "tags": json.dumps([]),
        "strategy_variables": "",
        "post_strategy_variables": "",
        "partialCloses": json.dumps([{"pct": 50, "pnl": round(pnl * 0.5, 2)}]) if has_partial else "[]",
        "entryScreenshot": "",
        "exitScreenshot": "",
        "railScreenshots": "[]",
        "sourceSessionId": spec["session_id"],
        "trading_session_id": spec["session_id"],
        "savedAt": int(datetime.now(tz=timezone.utc).timestamp() * 1000),
        "active_sl_at_exit": stop,
        "active_tps_at_exit": json.dumps([{"price": target, "percentage": 100, "hit": close_type == "TP"}]),
        "actual_risk_r": 1,
        "actual_rr_gross": round(abs(r_mult), 4),
        "array_base_price": entry,
        "balance_at_creation": round(balance, 2),
        "balance_at_exit": balance_after,
        "bar_close_r": "[]",
        "bar_high_r": "[]",
        "bar_low_r": "[]",
        "capture_ratio": round(capture, 4),
        "chart_trade_id": trade_num,
        "dayOfWeek": DAYS[entry_dt.weekday()],
        "entries_locked": False,
        "entryMarkerTimeMs": entry_ms,
        "entry_offset_r": 0,
        "exit_confirmed": True,
        "exit_timing_gap": round(mfe_r - abs(r_mult), 4),
        "finalClosePnL": pnl,
        "final_exit_bar": rng.randint(20, 400),
        "hasMultipleTakeProfits": has_multi_tp,
        "hasPartialCloses": has_partial,
        "holdingTimeDays": holding_days,
        "holdingTimeHours": holding_hours,
        "holdingTimeMs": holding_ms,
        "hourOfEntry": entry_dt.hour,
        "hourOfExit": exit_dt.hour,
        "initial_sl": stop,
        "initial_takeProfit": target,
        "isScaledTrade": is_scaled,
        "isSplitEntry": is_split,
        "maeTime": exit_ms - rng.randint(1000, 600000),
        "management_gap": round(rng.uniform(0, 1.5), 4),
        "market": inst["market"],
        "mfeTime": entry_ms + rng.randint(1000, max(2000, holding_ms // 2)),
        "month": MONTHS[entry_dt.month - 1],
        "multiTpSnapshot": json.dumps([target, _price_fmt(target + pip * 20, pip)]) if has_multi_tp else "",
        "netPnL": pnl,
        "originalRiskAmount": round(risk_amount, 2),
        "partialClosePnL": round(pnl * 0.5, 2) if has_partial else 0,
        "pip_value_at_entry": inst["pip_value"],
        "plannedEntrySnapshot": "",
        "plannedRRAtEntry": round(planned_rr, 4),
        "plannedTpSnapshot": "",
        "planned_risk_pct": round(risk_pct, 4),
        "pnl_dollars_gross": pnl,
        "post_checkpoints": "[]",
        "post_exit_anchor_time": exit_ms,
        "post_exit_bar_close_r": "[]",
        "post_exit_bar_high_r": "[]",
        "post_exit_bar_low_r": "[]",
        "preTradeNotes": json.dumps({}),
        "rulesFollowed": rules_followed,
        "session": "Session",
        "sl_modifications": "[]",
        "sourceFileId": spec["session_id"] % 100,
        "sourceFilterKey": f"session:{spec['session_id']}",
        "sourceKey": f"session:{spec['session_id']}",
        "sourceLabel": spec["name"],
        "sourceType": spec["sourceType"],
        "splitGroupId": f"split-{trade_num}" if is_split else "",
        "splitIndex": 1 if is_split else "",
        "splitTotal": 2 if is_split else "",
        "tag": setup,
        "total_bars_held": rng.randint(5, 300),
        "trail_sl_path": "[]",
        "v9PostTradeTags": json.dumps([post_tag]),
        "v9TradeNotes": "",
        "would_have_won": scenario == "left_on_table",
        "year": entry_dt.year,
        "accountType": spec["accountType"],
        "planAdherence": plan_adherence if spec["session_mode"].startswith("live") else "",
        "demons": json.dumps(demons),
        "originSource": spec["session_mode"],
        "session_mode": spec["session_mode"],
        "category_sheet": spec["category_sheet"],
    })

    next_dt = exit_dt + timedelta(minutes=rng.randint(10, 360))
    return row, balance_after, next_dt


def generate_session_trades(spec: dict, seed: int) -> list[dict]:
    rng = random.Random(seed)
    count = spec["trade_count"]
    tail_count = max(1, int(round(count * 0.2)))
    tail_indices = set(rng.sample(range(count), tail_count))

    start = datetime(2022, 1, 3, 8, 0, tzinfo=timezone.utc) + timedelta(days=spec["session_id"] % 90)
    balance = float(spec["start_balance"])
    entry_dt = start
    rows: list[dict] = []

    for i in range(count):
        row, balance, entry_dt = _build_trade(
            rng, spec, i + 1, entry_dt, balance, is_tail=(i in tail_indices)
        )
        rows.append(row)

    rows.sort(key=lambda r: (int(r["entryTime"]), int(r["tradeId"])))
    for idx, row in enumerate(rows, start=1):
        row["n"] = idx
        row["chart_trade_id"] = idx
    return rows


def _write_index_sheet(ws, specs: list[dict], trade_counts: dict[int, int]) -> None:
    headers = [
        "Category Sheet", "Session ID", "Session Name", "Mode", "Source Type",
        "Session Type", "Account Type", "Start Balance", "Trade Count", "Instruments",
    ]
    ws.append(headers)
    for spec in specs:
        ws.append([
            spec["category_sheet"],
            spec["session_id"],
            spec["name"],
            spec["session_mode"],
            spec["sourceType"],
            spec["session_type"],
            spec["accountType"],
            spec["start_balance"],
            trade_counts[spec["session_id"]],
            ", ".join(spec["instruments"]),
        ])


def _sort_trades(rows: list[dict]) -> list[dict]:
    rows.sort(
        key=lambda r: (
            str(r["sourceSessionName"]),
            int(r["sourceSessionId"]),
            int(r["entryTime"]),
            int(r["tradeId"]),
        )
    )
    return rows


def _autosize_columns(ws, max_width: int = 28) -> None:
    for col_idx, col in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max_width, max(10, len(col) + 2))


def build_workbook(out_path: Path, seed: int = 20260620) -> None:
    random.seed(seed)
    specs = _session_specs()
    wb = Workbook()
    index_ws = wb.active
    index_ws.title = "Index"
    trade_counts: dict[int, int] = {}
    by_category: dict[str, list[dict]] = {
        CATEGORY_SHEETS["standard_backtest"]: [],
        CATEGORY_SHEETS["prop_backtest"]: [],
        CATEGORY_SHEETS["live_journal"]: [],
    }

    for spec in specs:
        trades = generate_session_trades(spec, seed=seed + spec["session_id"])
        trade_counts[spec["session_id"]] = len(trades)
        by_category[spec["category_sheet"]].extend(trades)

    sheet_order = [
        CATEGORY_SHEETS["standard_backtest"],
        CATEGORY_SHEETS["prop_backtest"],
        CATEGORY_SHEETS["live_journal"],
    ]
    for sheet_name in sheet_order:
        trades = _sort_trades(by_category[sheet_name])
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(COLUMNS)
        for trade in trades:
            ws.append([trade.get(c, "") for c in COLUMNS])
        _autosize_columns(ws)

    _write_index_sheet(index_ws, specs, trade_counts)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def export_reference_csv(out_dir: Path, seed: int = 20260620) -> None:
    """Write one CSV per category sheet plus a merged file."""
    random.seed(seed)
    specs = _session_specs()
    by_category: dict[str, list[dict]] = {
        CATEGORY_SHEETS["standard_backtest"]: [],
        CATEGORY_SHEETS["prop_backtest"]: [],
        CATEGORY_SHEETS["live_journal"]: [],
    }
    for spec in specs:
        by_category[spec["category_sheet"]].extend(
            generate_session_trades(spec, seed=seed + spec["session_id"])
        )

    out_dir.mkdir(parents=True, exist_ok=True)
    merged: list[dict] = []
    slug_map = {
        CATEGORY_SHEETS["standard_backtest"]: "standard-backtest",
        CATEGORY_SHEETS["prop_backtest"]: "prop",
        CATEGORY_SHEETS["live_journal"]: "journal",
    }
    for sheet_name, rows in by_category.items():
        sorted_rows = _sort_trades(rows)
        merged.extend(sorted_rows)
        slug = slug_map[sheet_name]
        out_path = out_dir / f"dashboard-session-samples-{slug}-2026-06-20.csv"
        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(sorted_rows)

    merged_path = out_dir / "dashboard-session-samples-2026-06-20.csv"
    merged.sort(
        key=lambda r: (
            str(r.get("category_sheet") or r["session_mode"]),
            str(r["sourceSessionName"]),
            int(r["entryTime"]),
        )
    )
    with merged_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(merged)


def main(argv: list[str]) -> None:
    out = Path(argv[1]).expanduser() if len(argv) > 1 else _default_out()
    build_workbook(out)
    export_reference_csv(out.parent)
    print(f"Wrote workbook: {out}")
    print("Sheets: Index + Standard Backtest + Prop + Journal")
    print("CSVs: standard-backtest, prop, journal, and merged (docs/)")


if __name__ == "__main__":
    main(sys.argv)
