#!/usr/bin/env python3
"""
Adapt mentor-platform backtest Excel (e.g. mentor data/alae2.xlsx) to Talaria journal trades.

Maps mentor columns (variables_json, high/low price, etc.) onto the full dashboard trade
schema used by QA sessions — including MAE/MFE, strategy variables/tags, and bar paths.

Default target session: QA T1 · EURUSD Scalper BT (strategy id 57).

Usage:
  py scripts/adapt_mentor_xlsx_to_talaria.py
  py scripts/adapt_mentor_xlsx_to_talaria.py "mentor data/alae2.xlsx" -o "mentor data/alae2-talaria-adapted.xlsx"
  py scripts/adapt_mentor_xlsx_to_talaria.py --upload --session-name "QA T1 · EURUSD Scalper BT"
"""
from __future__ import annotations

import argparse
import json
import math
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
CHART_SCRIPTS = ROOT / "chart v 1.4" / "chart" / "scripts"
ANALYTICS_CORE = ROOT / "homepage" / "src" / "app" / "dashboard" / "analytics" / "backend"
for p in (CHART_SCRIPTS, ANALYTICS_CORE):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from generate_dashboard_session_samples import COLUMNS, DAYS, MONTHS  # noqa: E402
from analytics_core.session_seed_trades import (  # noqa: E402
    _FALLBACK_INSTRUMENTS,
    _build_trade_path_arrays,
    _norm_sym,
    _price_fmt,
    _tags_from_strategy_variables,
)

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

DEFAULT_INPUT = ROOT / "mentor data" / "alae2.xlsx"
DEFAULT_OUTPUT = ROOT / "mentor data" / "alae2-talaria-adapted.xlsx"
DEFAULT_SESSION_NAME = "QA T1 · EURUSD Scalper BT"
DEFAULT_STRATEGY_ID = 57
DEFAULT_STRATEGY_LABEL = "1-Min Momentum Scalper"
DEFAULT_SESSION_ID = 0  # resolved on --upload
DEFAULT_TIMEFRAME_MINUTES = 5
POST_EXIT_CANDLES = 20

SYM_ALIASES = {
    "XAU/USD": "XAUUSD",
    "GOLD": "XAUUSD",
    "GC": "XAUUSD",
    "NDX": "NQ",
    "NAS100": "NQ",
    "US100": "NQ",
    "USTEC": "NQ",
}

STRATEGY_BY_TICKER: dict[str, tuple[int, str]] = {
    "EURUSD": (57, "1-Min Momentum Scalper"),
    "GBPUSD": (58, "London Open Liquidity Scalp"),
    "USDJPY": (57, "1-Min Momentum Scalper"),
    "AUDUSD": (61, "4H Trend Pullback Swing"),
    "NZDUSD": (61, "4H Trend Pullback Swing"),
    "USDCAD": (61, "4H Trend Pullback Swing"),
    "USDCHF": (61, "4H Trend Pullback Swing"),
    "XAUUSD": (64, "Fibonacci Confluence Swing"),
    "ES": (60, "Opening Range Breakout"),
    "NQ": (59, "VWAP Reclaim Intraday"),
    "MES": (60, "Opening Range Breakout"),
    "MNQ": (59, "VWAP Reclaim Intraday"),
    "BTC": (62, "Liquidity Sweep + FVG"),
    "ETH": (62, "Liquidity Sweep + FVG"),
}

LIVE_DEMONS = ["revenge", "fomo", "overtrade", "early_exit", "late_entry", "size_up"]

DEMON_LABELS: dict[str, str] = {
    "revenge": "Revenge Trading",
    "fomo": "FOMO",
    "overtrade": "Overtrading",
    "early_exit": "Early Exit",
    "late_entry": "Late Entry",
    "size_up": "Size Up",
}

DEMON_EVIDENCE: dict[str, str] = {
    "revenge": "Increased aggression after a prior loss in the session.",
    "fomo": "Entered after the move was already extended without a fresh setup.",
    "overtrade": "Took an extra trade outside the planned session window.",
    "early_exit": "Closed before the planned target without a rule-based trigger.",
    "late_entry": "Entered after the optimal entry window had passed.",
    "size_up": "Position size exceeded the planned risk envelope.",
}

PLAN_ADHERENCE_TO_REVIEW: dict[str, str] = {
    "according-to-plan": "according_to_plan",
    "out-of-plan": "out_of_plan",
    "missed-trade": "missed_trade",
}

PLAN_ADHERENCE_TO_OUTCOME: dict[str, str] = {
    "according-to-plan": "followed",
    "out-of-plan": "violated",
    "missed-trade": "missed",
}

PLAN_ADHERENCE_TO_REASON: dict[str, str] = {
    "according-to-plan": "According to plan",
    "out-of-plan": "Out of plan",
    "missed-trade": "Missed trade",
}

SOURCE_KINDS = frozenset(
    {"backtest", "prop_backtest", "live_personal", "live_prop"}
)

# Strategy variable contract aligned to mentor "dol" field (maps to our pre-trade variables / tags).
PRE_VAR_DEFS: list[dict[str, Any]] = [
    {
        "id": "dol",
        "name": "dol",
        "vtype": "multi",
        "options": ["taken", "not taken"],
        "timing": "pre",
    }
]

POST_VAR_DEFS: list[dict[str, Any]] = [
    {
        "id": "outcome_review",
        "name": "Outcome review",
        "vtype": "yesno",
        "timing": "post",
    }
]


def _iso_z(ms: int) -> str:
    return datetime.fromtimestamp(ms / 1000.0, tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")


def _parse_dt(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value.astimezone(timezone.utc)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%SZ"):
        try:
            dt = datetime.strptime(text.replace("Z", ""), fmt.replace("Z", ""))
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return dt.astimezone(timezone.utc) if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _to_float(value: Any, default: float | None = None) -> float | None:
    if value is None or value == "":
        return default
    try:
        n = float(value)
        return n if math.isfinite(n) else default
    except (TypeError, ValueError):
        return default


def _infer_symbol(raw_symbol: Any, entry: float) -> str:
    if raw_symbol is not None and str(raw_symbol).strip():
        sym = str(raw_symbol).strip()
        if sym in SYM_ALIASES:
            return SYM_ALIASES[sym]
        return _norm_sym(sym)
    # Missing symbol — infer from price magnitude (mentor export quirk).
    if entry > 50 and entry < 250:
        return "USDJPY"
    if entry > 10000:
        return "NQ"
    if entry > 1500:
        return "XAUUSD"
    if 1.15 <= entry <= 1.35:
        return "GBPUSD"
    if 0.92 <= entry <= 1.02:
        return "AUDUSD"
    if 1.0 <= entry <= 1.15:
        return "EURUSD"
    if 0.95 <= entry < 1.0:
        return "USDCHF"
    return "EURUSD"


def _instrument(ticker: str) -> dict[str, Any]:
    fb = _FALLBACK_INSTRUMENTS.get(ticker) or _FALLBACK_INSTRUMENTS["EURUSD"]
    return {
        "ticker": ticker,
        "pip": float(fb["pip"]),
        "spread": float(fb["spread"]),
        "pip_value": float(fb["pip_value"]),
        "commission": 0.0,
        "market": str(fb["market"]),
    }


def _direction(raw: Any) -> str:
    s = str(raw or "long").strip().lower()
    return "SELL" if s in {"short", "sell", "s"} else "BUY"


def _side(direction: str) -> str:
    return "Long" if direction == "BUY" else "Short"


def _close_type(exit_px: float, stop: float, target: float, entry: float, pip: float) -> str:
    tol = max(pip * 2, abs(entry) * 1e-6)
    if abs(exit_px - target) <= tol:
        return "TP"
    if abs(exit_px - stop) <= tol:
        return "SL"
    if abs(exit_px - entry) <= tol:
        return "BE"
    return "Manual"


def _row_price(row: dict[str, Any], *keys: str) -> float | None:
    for key in keys:
        val = _to_float(row.get(key))
        if val is not None:
            return val
    return None


def _derive_excursion_prices(
    direction: str,
    entry: float,
    exit_px: float,
    stop: float,
    target: float,
) -> tuple[float, float]:
    """Bound high/low from prices we trust (entry, exit, stop, target). Does not change R math."""
    if direction == "BUY":
        return max(entry, exit_px, target), min(entry, exit_px, stop)
    return max(entry, exit_px, stop), min(entry, exit_px, target)


def _excursion_sanitize_reason(
    direction: str,
    entry: float,
    exit_px: float,
    stop: float,
    target: float,
    high: float | None,
    low: float | None,
) -> str | None:
    if high is None or low is None:
        return "missing_high_low"
    pip = _pip_for(entry)
    tol = max(pip, abs(entry) * 1e-6)
    if round(high, 2) == 30000.0 and round(low, 2) == 10000.0:
        return "placeholder_30k_10k"
    if high + tol < low:
        return "inverted_high_low"
    risk = abs(entry - stop) if stop is not None else max(abs(entry) * 0.001, pip)
    cap = max(risk * 50.0, abs(entry) * 0.08, risk + pip * 10.0)
    if abs(high - entry) > cap or abs(entry - low) > cap:
        return "absurd_deviation"
    band_lo = min(entry, exit_px, stop, target) - cap
    band_hi = max(entry, exit_px, stop, target) + cap
    if high < band_lo or high > band_hi or low < band_lo or low > band_hi:
        return "outside_trade_band"
    return None


def _sanitize_excursion_prices(
    direction: str,
    entry: float,
    exit_px: float,
    stop: float,
    target: float,
    high: float | None,
    low: float | None,
) -> tuple[float, float, bool, str | None]:
    reason = _excursion_sanitize_reason(direction, entry, exit_px, stop, target, high, low)
    if reason is None:
        return float(high), float(low), False, None
    derived_high, derived_low = _derive_excursion_prices(direction, entry, exit_px, stop, target)
    return derived_high, derived_low, True, reason


def audit_excursion_inputs(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Scan raw mentor rows for suspicious high/low inputs (and missing excursion prices)."""
    counts: dict[str, int] = {}
    samples: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        entry = _row_price(row, "entry_price", "entryPrice")
        exit_px = _row_price(row, "exit_price", "exitPrice")
        stop = _row_price(row, "stop_loss", "stopLoss")
        target = _row_price(row, "take_profit", "takeProfit")
        high = _row_price(row, "high_price", "highestPrice", "high")
        low = _row_price(row, "low_price", "lowestPrice", "low")
        if None in (entry, exit_px, stop, target):
            continue
        direction = _direction(row.get("direction"))
        reason = _excursion_sanitize_reason(direction, float(entry), float(exit_px), float(stop), float(target), high, low)
        if not reason:
            continue
        counts[reason] = counts.get(reason, 0) + 1
        if reason not in samples:
            samples[reason] = []
        if len(samples[reason]) < 3:
            samples[reason].append(
                {
                    "id": row.get("id"),
                    "entry": entry,
                    "high": high,
                    "low": low,
                    "exit": exit_px,
                }
            )
    return {"rows": len(rows), "bad": sum(counts.values()), "reasons": counts, "samples": samples}


def _excursions(
    direction: str,
    entry: float,
    stop: float,
    high: float,
    low: float,
    exit_px: float,
    r_mult: float,
) -> tuple[float, float, float, float, float, float]:
    risk = abs(entry - stop)
    if risk <= 0:
        risk = max(abs(entry) * 1e-5, 1e-8)
    if direction == "BUY":
        mfe_r = max(0.0, (high - entry) / risk)
        mae_r = max(0.0, (entry - low) / risk)
        mfe_price = _price_fmt(high, _pip_for(entry))
        mae_price = _price_fmt(low, _pip_for(entry))
        highest = _price_fmt(max(high, entry, exit_px), _pip_for(entry))
        lowest = _price_fmt(min(low, entry, exit_px), _pip_for(entry))
    else:
        mfe_r = max(0.0, (entry - low) / risk)
        mae_r = max(0.0, (high - entry) / risk)
        mfe_price = _price_fmt(low, _pip_for(entry))
        mae_price = _price_fmt(high, _pip_for(entry))
        highest = _price_fmt(max(high, entry, exit_px), _pip_for(entry))
        lowest = _price_fmt(min(low, entry, exit_px), _pip_for(entry))
    mfe_r = max(mfe_r, 0.05)
    mae_r = max(mae_r, 0.05)
    if r_mult >= 0:
        mae_r = min(mae_r, max(abs(r_mult), 0.1))
    else:
        mfe_r = min(mfe_r, 1.2)
    return mfe_r, mae_r, mfe_price, mae_price, highest, lowest


def _pip_for(entry: float) -> float:
    if entry > 1000:
        return 0.01 if entry < 5000 else 1.0
    if entry > 50:
        return 0.01
    return 0.0001


def _parse_mentor_variables(row: dict[str, Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    raw = row.get("variables_json")
    if raw:
        try:
            obj = json.loads(raw) if isinstance(raw, str) else raw
            if isinstance(obj, dict):
                for name, val in obj.items():
                    if isinstance(val, list) and val:
                        value = str(val[0])
                    else:
                        value = str(val)
                    out.append(
                        {
                            "id": str(name),
                            "name": str(name),
                            "vtype": "multi",
                            "value": value,
                        }
                    )
        except (json.JSONDecodeError, TypeError):
            pass
    for i in range(1, 11):
        key = f"var{i}"
        val = row.get(key)
        if val is not None and str(val).strip():
            out.append({"id": key, "name": key, "vtype": "multi", "value": str(val).strip()})
    if not out:
        for d in PRE_VAR_DEFS:
            out.append(
                {
                    "id": d["id"],
                    "name": d["name"],
                    "vtype": d.get("vtype", "multi"),
                    "value": "not taken",
                }
            )
    return out


def _post_variables(pnl: float) -> list[dict[str, Any]]:
    return [
        {
            "id": "outcome_review",
            "name": "Outcome review",
            "vtype": "yesno",
            "value": "Yes" if pnl >= 0 else "No",
        }
    ]


def _synthetic_bars(
    entry_ms: int,
    exit_ms: int,
    *,
    entry: float,
    exit_px: float,
    high: float,
    low: float,
    bar_ms: int,
) -> list[dict[str, Any]]:
    span = max(exit_ms - entry_ms, bar_ms)
    n = max(3, min(int(span / bar_ms) + 1, 120))
    bars: list[dict[str, Any]] = []
    for i in range(n):
        t = i / max(n - 1, 1)
        ts = entry_ms + int(t * (exit_ms - entry_ms))
        close = entry + (exit_px - entry) * t
        # Widen range toward observed high/low mid-trade.
        mid_w = math.sin(math.pi * t)
        h = close + (high - close) * mid_w * 0.85
        l = close - (close - low) * mid_w * 0.85
        h = max(h, close, low)
        l = min(l, close, high)
        bars.append({"t": ts, "o": close, "h": h, "l": l, "c": close})
    return bars


def _post_exit_bars(exit_ms: int, exit_px: float, entry: float, bar_ms: int, count: int) -> list[dict[str, Any]]:
    bars: list[dict[str, Any]] = []
    drift = (entry - exit_px) * 0.15
    for i in range(count):
        ts = exit_ms + (i + 1) * bar_ms
        close = exit_px + drift * (i + 1) / count
        bars.append({"t": ts, "o": close, "h": close, "l": close, "c": close})
    return bars


def _cell_value(key: str, val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)) and key not in {
        "bar_close_r", "bar_high_r", "bar_low_r",
        "post_exit_bar_close_r", "post_exit_bar_high_r", "post_exit_bar_low_r",
    }:
        return val
    if isinstance(val, (list, dict)):
        return json.dumps(val, ensure_ascii=False)
    return val


def _empty_row() -> dict[str, Any]:
    return {c: "" for c in COLUMNS}


def dominant_ticker_from_rows(rows: list[dict[str, Any]]) -> str:
    from collections import Counter

    counts: Counter[str] = Counter()
    for row in rows:
        entry = _to_float(row.get("entry_price"), 1.0) or 1.0
        sym = _infer_symbol(row.get("symbol"), float(entry))
        if sym:
            counts[sym] += 1
    if not counts:
        return "EURUSD"
    return counts.most_common(1)[0][0]


def pick_strategy_for_ticker(ticker: str) -> tuple[int, str]:
    key = _norm_sym(ticker)
    if key in STRATEGY_BY_TICKER:
        return STRATEGY_BY_TICKER[key]
    inst = _instrument(key)
    market = str(inst.get("market") or "Forex")
    if market == "Futures":
        return 59, "VWAP Reclaim Intraday"
    if market == "Crypto":
        return 62, "Liquidity Sweep + FVG"
    if key == "XAUUSD":
        return 64, "Fibonacci Confluence Swing"
    return 57, "1-Min Momentum Scalper"


def market_for_ticker(ticker: str) -> str:
    inst = _instrument(_norm_sym(ticker))
    market = str(inst.get("market") or "Forex")
    if market == "Commodity":
        return "Forex"
    return market


def _infer_plan_adherence(row: dict[str, Any]) -> str:
    raw = row.get("variables_json")
    if not raw:
        return "according-to-plan"
    try:
        obj = json.loads(raw) if isinstance(raw, str) else raw
    except (json.JSONDecodeError, TypeError):
        return "according-to-plan"
    if not isinstance(obj, dict):
        return "according-to-plan"
    for val in obj.values():
        parts = val if isinstance(val, list) else [val]
        text = " ".join(str(p) for p in parts).lower()
        if "missed" in text:
            return "missed-trade"
        if "out of plan" in text or "out-of-plan" in text:
            return "out-of-plan"
    return "according-to-plan"


def _variables_dict_from_row(row: dict[str, Any]) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    raw = row.get("variables_json")
    if raw:
        try:
            obj = json.loads(raw) if isinstance(raw, str) else raw
            if isinstance(obj, dict):
                for name, val in obj.items():
                    if isinstance(val, list):
                        out[str(name)] = [str(v) for v in val if str(v).strip()]
                    elif val is not None and str(val).strip():
                        out[str(name)] = [str(val)]
        except (json.JSONDecodeError, TypeError):
            pass
    for i in range(1, 11):
        val = row.get(f"var{i}")
        if val is not None and str(val).strip():
            out[f"var{i}"] = [str(val).strip()]
    return out


def _demon_cost_r(trade: dict[str, Any]) -> float:
    r_mult = abs(float(trade.get("rMultiple") or trade.get("actual_rr_net") or 0))
    return round(max(0.35, min(2.5, r_mult if r_mult > 0 else 0.75)), 2)


def _make_synthetic_demon(
    trade: dict[str, Any],
    demon_type: str,
    *,
    source: str = "auto",
    state: str = "detected",
) -> dict[str, Any]:
    cost_r = _demon_cost_r(trade)
    risk_amount = abs(float(trade.get("riskAmount") or trade.get("riskPerTrade") or 1000))
    return {
        "type": demon_type,
        "label": DEMON_LABELS.get(demon_type, demon_type.replace("_", " ").title()),
        "source": source,
        "state": state,
        "cost_R": cost_r,
        "cost_$": int(round(cost_r * risk_amount)),
        "evidence": DEMON_EVIDENCE.get(demon_type, "Behavioral pattern detected on review."),
        "detected_at": trade.get("closeTime") or trade.get("exitTime") or trade.get("entryTime"),
    }


def _synthesize_live_journal_discipline(
    trade: dict[str, Any],
    rng: Any,
    mentor_plan_adherence: str,
) -> str:
    """Assign a demo-rich discipline/demon mix for live journal imports only."""
    roll = rng.random() * 100.0
    if mentor_plan_adherence in {"out-of-plan", "missed-trade"}:
        bucket = mentor_plan_adherence
        manual_demon = False
    elif roll < 76.0:
        bucket = "according-to-plan"
        manual_demon = False
    elif roll < 84.0:
        bucket = "discipline_breach_sl"
        manual_demon = False
    elif roll < 90.0:
        bucket = "out-of-plan"
        manual_demon = False
    elif roll < 94.0:
        bucket = "out-of-plan"
        manual_demon = True
    elif roll < 97.0:
        bucket = "missed-trade"
        manual_demon = False
    else:
        bucket = "discipline_breach_risk"
        manual_demon = False

    if bucket == "discipline_breach_sl":
        entry = float(trade.get("entryPrice") or 0)
        stop = float(trade.get("stopLoss") or trade.get("initial_sl") or entry)
        risk_pts = abs(entry - stop) or max(abs(entry) * 0.001, 0.0001)
        direction = str(trade.get("direction") or "BUY").upper()
        widen = risk_pts * rng.uniform(0.12, 0.28)
        if direction == "BUY":
            widened_stop = _price_fmt(stop - widen, _pip_for(entry))
        else:
            widened_stop = _price_fmt(stop + widen, _pip_for(entry))
        trade["sl_modifications"] = [
            {
                "field": "SL",
                "trigger": "MANUAL",
                "old": stop,
                "new": widened_stop,
                "at": trade.get("entryTime"),
            }
        ]
        bucket = "according-to-plan"
    elif bucket == "discipline_breach_risk":
        trade["planned_risk_r"] = 1.0
        trade["actual_risk_r"] = round(rng.uniform(1.25, 1.65), 2)
        bucket = "according-to-plan"

    plan_adherence = bucket
    plan_review = PLAN_ADHERENCE_TO_REVIEW[plan_adherence]
    plan_outcome = PLAN_ADHERENCE_TO_OUTCOME[plan_adherence]
    reason = PLAN_ADHERENCE_TO_REASON[plan_adherence]

    trade["planAdherence"] = plan_adherence
    trade["planReviewKey"] = plan_review
    trade["planReview"] = plan_review
    trade["planOutcome"] = plan_outcome
    trade["plan_behavior"] = reason
    trade["planBehavior"] = reason
    trade["rulesFollowed"] = plan_adherence == "according-to-plan"
    trade["missedTrade"] = plan_adherence == "missed-trade"
    trade["would_have_won"] = plan_adherence == "missed-trade" and float(trade.get("mfe_r") or 0) > 1.0

    demons: list[dict[str, Any]] = []
    demon_catcher: dict[str, Any] | None = None
    if plan_adherence == "out-of-plan":
        if manual_demon or rng.random() < 0.35:
            demon_type = rng.choice(LIVE_DEMONS)
            demon_catcher = {
                "planReview": plan_review,
                "category": DEMON_LABELS.get(demon_type, demon_type),
                "trigger": DEMON_EVIDENCE.get(demon_type, "Manual review flagged a behavioral slip."),
                "correction": "Pause one bar, re-check the setup checklist, then re-enter only on plan.",
            }
            trade["demon_category"] = demon_catcher["category"]
            trade["demon_trigger"] = demon_catcher["trigger"]
            trade["demon_correction"] = demon_catcher["correction"]
            demons.append(
                _make_synthetic_demon(trade, demon_type, source="manual", state="confirmed")
            )
        elif rng.random() < 0.72:
            for demon_type in rng.sample(LIVE_DEMONS, k=rng.randint(1, 2)):
                demons.append(_make_synthetic_demon(trade, demon_type))

    trade["demons"] = demons
    trade["demonCatcher"] = demon_catcher

    notes = trade.get("postTradeNotes")
    if not isinstance(notes, dict):
        notes = {}
    notes.update(
        {
            "rule_outcome": plan_outcome,
            "reason": reason,
            "planAdherence": plan_adherence,
            "syntheticDiscipline": True,
        }
    )
    trade["postTradeNotes"] = notes

    pre_notes = trade.get("preTradeNotes")
    if isinstance(pre_notes, dict):
        pre_notes["planAdherence"] = plan_adherence
    else:
        trade["preTradeNotes"] = {"planAdherence": plan_adherence}

    return plan_adherence


def _apply_source_metadata(
    trade: dict[str, Any],
    *,
    source_kind: str,
    source_id: int,
    source_name: str,
    profile_id: int | None,
    mentor_filename: str,
    plan_adherence: str,
    rng_seed: int,
) -> None:
    import random

    rng = random.Random(rng_seed)
    is_live = source_kind in {"live_personal", "live_prop"}
    is_prop = source_kind in {"live_prop", "prop_backtest"}

    if is_live:
        account_key = profile_id if profile_id is not None else source_id
        trade["sourceSessionId"] = source_id
        trade["trading_session_id"] = source_id
        trade["sourceKey"] = f"journalAccount:{account_key}"
        trade["sourceFilterKey"] = f"journalAccount:{account_key}"
        trade["sourceLabel"] = source_name
        trade["sourceType"] = "journal"
        trade["sourceDashboardKind"] = "journal"
        trade["journalAccountKey"] = account_key
        trade["liveJournal"] = True
        trade["session"] = "Journal"
        trade["category_sheet"] = "Prop" if is_prop else "Journal"
        trade["session_mode"] = "prop_live_journal" if is_prop else "live_journal"
        trade["accountType"] = "prop" if is_prop else "private"
        trade["originSource"] = "mentor_import_live"
    else:
        trade["sourceSessionId"] = source_id
        trade["trading_session_id"] = source_id
        trade["sourceKey"] = f"session:{source_id}"
        trade["sourceFilterKey"] = f"session:{source_id}"
        trade["sourceLabel"] = source_name
        trade["sourceType"] = "backtest"
        trade["session"] = source_name
        trade["category_sheet"] = "Prop" if is_prop else "Standard Backtest"
        trade["session_mode"] = "prop_backtest" if is_prop else "standard_backtest"
        trade["accountType"] = "prop" if is_prop else "private"
        trade["originSource"] = "mentor_import_prop" if is_prop else "mentor_import"

    trade["planAdherence"] = plan_adherence
    trade["rulesFollowed"] = plan_adherence == "according-to-plan"
    source_type_num = {
        "backtest": 1,
        "prop_backtest": 2,
        "live_personal": 3,
        "live_prop": 4,
    }.get(source_kind, 1)
    trade["source_type"] = source_type_num
    if is_live:
        plan_adherence = _synthesize_live_journal_discipline(trade, rng, plan_adherence)
    else:
        trade["demons"] = (
            rng.sample(LIVE_DEMONS, k=rng.randint(1, 2))
            if plan_adherence == "out-of-plan"
            else []
        )
        trade["would_have_won"] = plan_adherence == "missed-trade" and float(trade.get("mfe_r") or 0) > 1.0
    if not is_live:
        trade["preTradeNotes"] = {
            "mentorSource": mentor_filename,
            "planAdherence": plan_adherence,
        }
    elif isinstance(trade.get("preTradeNotes"), dict):
        trade["preTradeNotes"]["mentorSource"] = mentor_filename
    else:
        trade["preTradeNotes"] = {
            "mentorSource": mentor_filename,
            "planAdherence": plan_adherence,
        }
    if is_prop:
        trade["propFirm"] = trade.get("propFirm") or "FTMO"


def convert_mentor_row(
    row: dict[str, Any],
    *,
    index: int,
    source_id: int,
    source_name: str,
    strategy_label: str,
    strategy_id: int,
    balance_before: float,
    bar_ms: int,
    post_exit_candles: int,
    mentor_stem: str,
    mentor_filename: str,
    source_kind: str = "backtest",
    profile_id: int | None = None,
) -> tuple[dict[str, Any], float]:
    entry = _row_price(row, "entry_price", "entryPrice")
    exit_px = _row_price(row, "exit_price", "exitPrice")
    stop = _row_price(row, "stop_loss", "stopLoss")
    target = _row_price(row, "take_profit", "takeProfit")
    high = _row_price(row, "high_price", "highestPrice", "high")
    low = _row_price(row, "low_price", "lowestPrice", "low")
    if None in (entry, exit_px, stop, target):
        raise ValueError("missing required prices")

    ticker = _infer_symbol(row.get("symbol"), float(entry))
    inst = _instrument(ticker)
    pip = float(inst["pip"])

    entry = _price_fmt(float(entry), pip)
    exit_px = _price_fmt(float(exit_px), pip)
    stop = _price_fmt(float(stop), pip)
    target = _price_fmt(float(target), pip)
    direction = _direction(row.get("direction"))
    high, low, excursion_sanitized, excursion_sanitize_reason = _sanitize_excursion_prices(
        direction, entry, exit_px, stop, target, high, low
    )
    high = _price_fmt(float(high), pip)
    low = _price_fmt(float(low), pip)
    entry_dt = _parse_dt(row.get("entry_datetime") or row.get("trade_date"))
    exit_dt = _parse_dt(row.get("exit_datetime") or row.get("exit_datetime"))
    if not entry_dt:
        raise ValueError("missing entry datetime")
    if not exit_dt:
        exit_dt = entry_dt
    entry_ms = int(entry_dt.timestamp() * 1000)
    exit_ms = int(exit_dt.timestamp() * 1000)
    if exit_ms <= entry_ms:
        exit_ms = entry_ms + bar_ms

    pnl = _to_float(row.get("pnl"), 0.0) or 0.0
    risk_amount = _to_float(row.get("risk_amount"), 1000.0) or 1000.0
    r_mult = _to_float(row.get("rr"), 0.0)
    if r_mult is None:
        r_mult = (pnl / risk_amount) if risk_amount else 0.0

    sl_dist = abs(entry - stop) or pip
    tp_dist = abs(target - entry)
    planned_rr = round(tp_dist / sl_dist, 4) if sl_dist else 1.0

    mfe_r, mae_r, mfe_price, mae_price, highest, lowest = _excursions(
        direction, entry, stop, high, low, exit_px, float(r_mult)
    )

    close_type = _close_type(exit_px, stop, target, entry, pip)
    hold_ms = exit_ms - entry_ms
    hold_minutes = max(1, int(round(hold_ms / 60000)))

    plan_adherence = _infer_plan_adherence(row)

    pre_vars = _parse_mentor_variables(row)
    post_vars = _post_variables(pnl)
    pre_tags = _tags_from_strategy_variables(pre_vars)
    post_tag = "Win" if pnl > 0 else "Loss" if pnl < 0 else "BE"
    post_tags = _tags_from_strategy_variables(post_vars) + [post_tag]
    post_tags = list(dict.fromkeys(post_tags))

    in_bars = _synthetic_bars(
        entry_ms, exit_ms, entry=entry, exit_px=exit_px, high=high, low=low, bar_ms=bar_ms
    )
    all_bars = in_bars + _post_exit_bars(exit_ms, exit_px, entry, bar_ms, post_exit_candles)
    path = _build_trade_path_arrays(
        all_bars,
        entry_ms=entry_ms,
        exit_ms=exit_ms,
        direction=direction,
        array_base=entry,
        initial_sl=stop,
        post_exit_candles=post_exit_candles,
    )

    qty = _to_float(row.get("quantity"), 1.0) or 1.0
    balance_after = round(balance_before + pnl, 2)
    mentor_id = row.get("id") or index
    trade_id = f"mentor-{mentor_stem}-{mentor_id}"
    try:
        numeric_id = int(mentor_id)
    except (TypeError, ValueError):
        numeric_id = int(source_id) * 10000 + index

    entry_screenshot = str(row.get("entry_screenshot") or "").strip()
    exit_screenshot = str(row.get("exit_screenshot") or "").strip()
    notes = str(row.get("notes") or "").strip()

    capture_ratio = round(abs(float(r_mult)) / mfe_r, 4) if mfe_r else 0.0
    mfe_time = entry_ms + max(1000, hold_ms // 2)
    mae_time = exit_ms - max(1000, min(hold_ms // 3, 600000))

    out = _empty_row()
    out.update({
        "journal_trade_id": numeric_id,
        "trade_id": numeric_id,
        "client_trade_id": index,
        "tradeId": trade_id,
        "id": numeric_id,
        "n": index,
        "sourceSessionName": source_name,
        "setup": strategy_label,
        "symbol": ticker,
        "ticker": ticker,
        "direction": direction,
        "side": _side(direction),
        "type": direction,
        "orderType": "market",
        "quantity": qty,
        "status": "closed",
        "entryTime": entry_ms,
        "openTime": entry_ms,
        "entryDate": _iso_z(entry_ms),
        "date": entry_dt.strftime("%Y-%m-%d"),
        "exitTime": exit_ms,
        "closeTime": _iso_z(exit_ms),
        "exitDate": _iso_z(exit_ms),
        "entryPrice": entry,
        "openPrice": entry,
        "exitPrice": exit_px,
        "closePrice": exit_px,
        "stopLoss": stop,
        "takeProfit": target,
        "pnl": pnl,
        "pnl_dollars_net": pnl,
        "realizedPnL": pnl,
        "rMultiple": round(float(r_mult), 4),
        "actual_rr_net": round(abs(float(r_mult)), 4),
        "actualRR": round(abs(float(r_mult)), 4),
        "rewardToRiskRatio": planned_rr,
        "riskAmount": risk_amount,
        "riskPerTrade": risk_amount,
        "plannedRR": planned_rr,
        "duration": hold_minutes,
        "closeType": close_type,
        "mfe": mfe_price,
        "mae": mae_price,
        "mfe_r": round(mfe_r, 4),
        "mae_r": round(mae_r, 4),
        "total_mfe_r": round(mfe_r, 4),
        "highestPrice": highest,
        "lowestPrice": lowest,
        "commission_total": _to_float(row.get("commission"), 0.0) or 0.0,
        "commission_at_entry": _to_float(row.get("commission"), 0.0) or 0.0,
        "spread_pips_at_entry": inst["spread"],
        "postTradeNotes": {
            "mentorImport": True,
            "mentorTradeId": row.get("id"),
            "mentorFile": mentor_filename,
            "entryScreenshotUrl": entry_screenshot or None,
            "exitScreenshotUrl": exit_screenshot or None,
            "notes": notes or None,
            "postStrategyVariables": post_vars,
            "planAdherence": plan_adherence,
            "excursionSanitized": excursion_sanitized,
            "excursionSanitizeReason": excursion_sanitize_reason,
            "sourceHighPrice": _row_price(row, "high_price", "highestPrice", "high"),
            "sourceLowPrice": _row_price(row, "low_price", "lowestPrice", "low"),
        },
        "preTags": pre_tags,
        "postTags": post_tags,
        "tags": list(dict.fromkeys(pre_tags + post_tags)),
        "strategy_variables": pre_vars,
        "post_strategy_variables": post_vars,
        "partialCloses": [],
        "entryScreenshot": entry_screenshot,
        "exitScreenshot": exit_screenshot,
        "railScreenshots": [],
        "sourceSessionId": source_id,
        "trading_session_id": source_id,
        "savedAt": int(datetime.now(tz=timezone.utc).timestamp() * 1000),
        "active_sl_at_exit": stop,
        "active_tps_at_exit": [{"price": target, "percentage": 100, "hit": close_type == "TP"}],
        "actual_risk_r": 1.0,
        "actual_rr_gross": round(abs(float(r_mult)), 4),
        "array_base_price": entry,
        "balance_at_creation": round(balance_before, 2),
        "balance_at_exit": balance_after,
        "bar_close_r": path["bar_close_r"],
        "bar_high_r": path["bar_high_r"],
        "bar_low_r": path["bar_low_r"],
        "capture_ratio": capture_ratio,
        "chart_trade_id": index,
        "dayOfWeek": DAYS[entry_dt.weekday()],
        "entries_locked": False,
        "entryMarkerTimeMs": float(entry_ms),
        "entry_offset_r": 0.0,
        "exit_confirmed": True,
        "exit_timing_gap": round(mfe_r - abs(float(r_mult)), 4),
        "finalClosePnL": pnl,
        "final_exit_bar": path["final_exit_bar"],
        "hasMultipleTakeProfits": False,
        "hasPartialCloses": False,
        "holdingTimeDays": round(hold_ms / 86400000, 4),
        "holdingTimeHours": round(hold_ms / 3600000, 4),
        "holdingTimeMs": hold_ms,
        "hourOfEntry": entry_dt.hour,
        "hourOfExit": exit_dt.hour,
        "initial_sl": stop,
        "initial_takeProfit": target,
        "isScaledTrade": False,
        "isSplitEntry": False,
        "maeTime": float(mae_time),
        "management_gap": 0.0,
        "market": inst["market"],
        "mfeTime": float(mfe_time),
        "month": MONTHS[entry_dt.month - 1],
        "multiTpSnapshot": [target],
        "netPnL": pnl,
        "originalRiskAmount": risk_amount,
        "partialClosePnL": 0,
        "pip_value_at_entry": inst["pip_value"],
        "plannedEntrySnapshot": entry,
        "plannedRRAtEntry": planned_rr,
        "plannedTpSnapshot": target,
        "planned_risk_pct": round((risk_amount / max(balance_before, 1)) * 100, 4),
        "pnl_dollars_gross": pnl,
        "post_checkpoints": [],
        "post_exit_anchor_time": path["post_exit_anchor_time"],
        "post_exit_bar_close_r": path["post_exit_bar_close_r"],
        "post_exit_bar_high_r": path["post_exit_bar_high_r"],
        "post_exit_bar_low_r": path["post_exit_bar_low_r"],
        "preTradeNotes": {"mentorSource": mentor_filename, "planAdherence": plan_adherence},
        "rulesFollowed": plan_adherence == "according-to-plan",
        "session": source_name,
        "sl_modifications": [],
        "sourceFileId": 0,
        "sourceFilterKey": f"session:{source_id}",
        "sourceKey": f"session:{source_id}",
        "sourceLabel": source_name,
        "sourceType": "backtest",
        "splitGroupId": None,
        "splitIndex": None,
        "splitTotal": None,
        "tag": strategy_label,
        "total_bars_held": len(path["bar_close_r"]),
        "trail_sl_path": [],
        "v9PostTradeTags": post_tags,
        "v9TradeNotes": notes or f"Imported from mentor ({mentor_filename}, id={row.get('id')})",
        "would_have_won": False,
        "year": entry_dt.year,
        "accountType": "private",
        "planAdherence": plan_adherence,
        "demons": [],
        "originSource": "mentor_import",
        "session_mode": "standard_backtest",
        "category_sheet": "Standard Backtest",
        "strategy_id": strategy_id,
        "entries": [{"price": entry, "qty": qty, "quantity": qty}],
        "targets": [{"price": target, "qty": qty, "quantity": qty}],
        "exits": [{"price": exit_px, "qty": qty, "quantity": qty}],
    })
    _apply_source_metadata(
        out,
        source_kind=source_kind,
        source_id=source_id,
        source_name=source_name,
        profile_id=profile_id,
        mentor_filename=mentor_filename,
        plan_adherence=plan_adherence,
        rng_seed=numeric_id,
    )
    return out, balance_after


def read_mentor_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb["Journal"] if "Journal" in wb.sheetnames else wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    header = next(rows_iter)
    headers = [str(h) if h is not None else "" for h in header]
    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        if not any(v is not None and str(v).strip() for v in raw):
            continue
        out.append({headers[i]: raw[i] for i in range(len(headers))})
    wb.close()
    return out


def write_workbook(path: Path, trades: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"
    ws.append(COLUMNS)
    for trade in trades:
        ws.append([_cell_value(k, trade.get(k)) for k in COLUMNS])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def convert_file(
    input_path: Path,
    *,
    source_id: int,
    source_name: str,
    strategy_label: str,
    strategy_id: int,
    start_balance: float = 10000.0,
    source_kind: str = "backtest",
    profile_id: int | None = None,
    mentor_stem: str | None = None,
) -> list[dict[str, Any]]:
    mentor_rows = read_mentor_rows(input_path)
    stem = mentor_stem or input_path.stem
    filename = input_path.name
    bar_ms = DEFAULT_TIMEFRAME_MINUTES * 60 * 1000
    trades: list[dict[str, Any]] = []
    errors: list[str] = []
    balance = start_balance
    for i, row in enumerate(mentor_rows, start=1):
        try:
            trade, balance = convert_mentor_row(
                row,
                index=i,
                source_id=source_id,
                source_name=source_name,
                strategy_label=strategy_label,
                strategy_id=strategy_id,
                balance_before=balance,
                bar_ms=bar_ms,
                post_exit_candles=POST_EXIT_CANDLES,
                mentor_stem=stem,
                mentor_filename=filename,
                source_kind=source_kind,
                profile_id=profile_id,
            )
            trades.append(trade)
        except Exception as exc:
            errors.append(f"row {i} (mentor id={row.get('id')}): {exc}")
    if errors:
        print("Conversion warnings:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
    trades.sort(key=lambda t: (float(t["entryTime"]), str(t["tradeId"])))
    for idx, trade in enumerate(trades, start=1):
        trade["n"] = idx
        trade["client_trade_id"] = idx
        trade["chart_trade_id"] = idx
    return trades


def upload_journal(trades: list[dict[str, Any]], *, session_id: int, origin: str, email: str, password: str) -> None:
    seed_script = ROOT / "scripts" / "seed_session_demo_trades.py"
    if str(ROOT / "scripts") not in sys.path:
        sys.path.insert(0, str(ROOT / "scripts"))
    from seed_session_demo_trades import Client, patch_journal  # noqa: E402

    client = Client(origin)
    client.login(email, password)
    print(f"Uploading {len(trades)} trades to session {session_id} ...")
    patch_journal(client, session_id, trades)
    print("Upload complete.")


def resolve_session_id(origin: str, email: str, password: str, session_name: str) -> int:
    from seed_session_demo_trades import Client, find_session  # noqa: E402

    client = Client(origin)
    client.login(email, password)
    session = find_session(client, session_id=None, session_name=session_name)
    return int(session["id"])


def main() -> int:
    parser = argparse.ArgumentParser(description="Adapt mentor backtest Excel to Talaria journal trades")
    parser.add_argument("input", nargs="?", default=str(DEFAULT_INPUT))
    parser.add_argument("-o", "--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--json-output", default="")
    parser.add_argument("--session-id", type=int, default=DEFAULT_SESSION_ID)
    parser.add_argument("--session-name", default=DEFAULT_SESSION_NAME)
    parser.add_argument("--strategy-id", type=int, default=DEFAULT_STRATEGY_ID)
    parser.add_argument("--strategy-label", default=DEFAULT_STRATEGY_LABEL)
    parser.add_argument("--start-balance", type=float, default=10000.0)
    parser.add_argument("--origin", default="http://31.97.192.82:3000")
    parser.add_argument("--email", default="data@talaria-log.com")
    parser.add_argument("--password", default="data@talaria-log.com")
    parser.add_argument("--upload", action="store_true")
    parser.add_argument("--resolve-session", action="store_true", help="Look up session id by name on origin")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.is_file():
        print(f"Input not found: {input_path}", file=sys.stderr)
        return 1

    session_id = args.session_id
    if args.upload or args.resolve_session:
        session_id = resolve_session_id(args.origin, args.email, args.password, args.session_name)
        print(f"Resolved session id={session_id} ({args.session_name})")

    trades = convert_file(
        input_path,
        source_id=session_id,
        source_name=args.session_name,
        strategy_label=args.strategy_label,
        strategy_id=args.strategy_id,
        start_balance=args.start_balance,
        source_kind="backtest",
    )
    if not trades:
        print("No trades converted.", file=sys.stderr)
        return 1

    out_path = Path(args.output)
    write_workbook(out_path, trades)
    print(f"Wrote {len(trades)} trades -> {out_path}")

    json_path = Path(args.json_output) if args.json_output else out_path.with_suffix(".json")
    json_path.write_text(json.dumps(trades, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote JSON -> {json_path}")

    with_paths = sum(1 for t in trades if t.get("bar_close_r"))
    with_vars = sum(1 for t in trades if t.get("strategy_variables"))
    print(
        f"Summary: wins={sum(1 for t in trades if float(t['pnl']) > 0)} "
        f"losses={sum(1 for t in trades if float(t['pnl']) < 0)} "
        f"bar_paths={with_paths} strategy_variables={with_vars}"
    )
    sample = trades[0]
    print(
        f"Sample: {sample['ticker']} {sample['direction']} pnl={sample['pnl']} "
        f"mae_r={sample['mae_r']} mfe_r={sample['mfe_r']} "
        f"bar_close_r len={len(sample.get('bar_close_r') or [])} "
        f"preTags={sample.get('preTags')}"
    )

    if args.upload:
        upload_journal(
            trades,
            session_id=session_id,
            origin=args.origin,
            email=args.email,
            password=args.password,
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
