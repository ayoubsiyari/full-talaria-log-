#!/usr/bin/env py -3
"""Generate comprehensive Talaria Chart QA tracker for non-technical testers."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

OUTPUT = Path(__file__).resolve().parent / "Talaria-Chart-QA-Tracker.xlsx"
STATUS_LIST = '"☑ Pass,☐ Fail,— Not tested,⚠ Partial"'
ISSUE_LIST = '"☑ Yes,☐ No,— Not tested"'

# ─── Drawing tools from V9 rail (TalariaV8bLive.jsx V9_RAIL_ICONS_BY_GROUP) ───

LINE_TOOLS = [
    ("Trend Line", "trendline", "Left rail → Lines group → Trend Line (shortcut T). Click two points on chart."),
    ("Horizontal Ray", "hray", "Lines group → Horizontal Ray. Click anchor; ray extends right."),
    ("Horizontal Line", "hline", "Lines group → Horizontal Line (shortcut H). Full-width price level."),
    ("Vertical Line", "vline", "Lines group → Vertical Line (shortcut V). Full-height time line."),
    ("Ray", "ray", "Lines group → Ray. One anchor + direction."),
    ("Extended Line", "extendedLine", "Lines group → Extended Line. Line extends both directions."),
    ("Cross Line", "crossLine", "Lines group → Cross Line. Horizontal + vertical through one point."),
    ("Polyline", "polyline", "Lines group → Polyline. Click multiple points; finish with double-click or Enter."),
    ("Path", "pathTool", "Lines group → Path. Multi-segment path without fill."),
    ("Curve", "curve", "Lines group → Curve. Bezier-style curve between points."),
    ("Double Curve", "doubleCurve", "Lines group → Double Curve."),
]

SHAPE_TOOLS = [
    ("Rectangle", "rect", "Shapes group → Rectangle (shortcut R). Drag diagonal corners."),
    ("Triangle", "triangle", "Shapes group → Triangle."),
    ("Arc", "arcShape", "Shapes group → Arc."),
    ("Ellipse", "ellipse", "Shapes group → Ellipse."),
    ("Circle", "circle", "Shapes group → Circle."),
    ("Arrow Marker", "arrowMarker", "Shapes group → Arrow Marker. Place on bar/point."),
    ("Arrow Line", "arrowLine", "Shapes group → Arrow (line with arrowhead)."),
    ("Arrow Mark Up", "arrowUp", "Shapes group → Arrow Mark Up (green triangle)."),
    ("Arrow Mark Down", "arrowDn", "Shapes group → Arrow Mark Down (red triangle)."),
]

CHANNEL_TOOLS = [
    ("Parallel Channel", "channel", "Channels group → Parallel Channel. Three anchor clicks."),
    ("Regression Channel", "regressionCh", "Channels group → Regression Channel."),
    ("Flat Top/Bottom", "flatChannel", "Channels group → Flat Top/Bottom."),
    ("Disjoint Channel", "disjointCh", "Channels group → Disjoint Channel."),
    ("Pitchfork — Original", "pitchfork", "Channels group → Pitchfork. Test Original variant in settings if listed."),
    ("Pitchfork — Schiff", "pitchfork", "Place pitchfork → Settings → change variant to Schiff."),
    ("Pitchfork — Modified Schiff", "pitchfork", "Place pitchfork → Settings → Modified Schiff variant."),
    ("Pitchfork — Inside", "pitchfork", "Place pitchfork → Settings → Inside variant."),
]

BRUSH_TOOLS = [
    ("Brush", "draw", "Brushes group → Brush. Freehand stroke on chart."),
    ("Highlighter", "brush", "Brushes group → Highlighter. Semi-transparent stroke."),
]

FIB_GANN_TOOLS = [
    ("Fib Retracement", "fib", "Fib group → Fib Retracement (shortcut F). Verify 23.6/38.2/50/61.8% levels."),
    ("Fib Extension", "fibExtension", "Fib group → Trend-Based Fib Extension."),
    ("Fib Channel", "fibChannel", "Fib group → Fib Channel."),
    ("Fib Time Zone", "fibTimeZone", "Fib group → Fib Time Zone. Vertical time divisions."),
    ("Fib Speed Resistance Fan", "fibFan", "Fib group → Speed Resistance Fan."),
    ("Trend-Based Fib Time", "fibTime", "Fib group → Trend Fib Time."),
    ("Fib Circles", "fibCircles", "Fib group → Fib Circles."),
    ("Fib Spiral", "fibSpiral", "Fib group → Fib Spiral."),
    ("Fib Speed Resistance Arcs", "fibArcs", "Fib group → Fib Arcs."),
    ("Fib Wedge", "fibWedge", "Fib group → Fib Wedge."),
    ("Gann Box", "gannBox", "Fib/Gann group → Gann Box grid."),
    ("Gann Square Fixed", "gannSquare", "Fib/Gann group → Gann Square Fixed."),
    ("Gann Fan", "gannFan", "Fib/Gann group → Gann Fan."),
]

PATTERN_TOOLS = [
    ("Elliott Impulse (12345)", "elliott5", "Patterns group → Elliott Impulse wave labels."),
    ("Elliott Correction (ABC)", "elliottABC", "Patterns group → Elliott Correction."),
    ("Elliott Triangle (ABCDE)", "elliottTri", "Patterns group → Elliott Triangle."),
    ("Elliott Double Combo (WXY)", "elliottWXY", "Patterns group → Double Combo."),
    ("Elliott Triple Combo (WXYXZ)", "elliottWXYXZ", "Patterns group → Triple Combo."),
    ("XABCD Pattern", "xabcd", "Patterns group → XABCD harmonic."),
    ("Head & Shoulders", "headShoulders", "Patterns group → Head & Shoulders template."),
    ("ABCD Pattern", "abcdPattern", "Patterns group → ABCD."),
    ("Triangle Pattern", "triPattern", "Patterns group → Triangle Pattern tool."),
    ("Three Drives", "threeDrives", "Patterns group → Three Drives."),
    ("Cypher Pattern", "cypher", "If in menu: Cypher Pattern (engine tool)."),
    ("Cyclic Lines", "cyclic", "If in menu: Cyclic Lines."),
    ("Time Cycles", "timeCycles", "If in menu: Time Cycles."),
    ("Sine Line", "sine", "If in menu: Sine Line."),
    ("Pitchfan", "pitchfan", "If in menu: Pitchfan."),
]

MEASURE_TOOLS = [
    ("Long Position (RR box)", "longPos", "Measure group → Long Position. Entry/SL/TP zones + stats."),
    ("Short Position (RR box)", "shortPos", "Measure group → Short Position."),
    ("Date-Price Range / Measure", "measure", "Measure group → Range Tool. Shows bars, price distance, time span."),
]

TEXT_TOOLS = [
    ("Text", "text", "Text group → Text. Type label; verify on-chart editor."),
    ("Note", "note", "Text group → Note with connector."),
    ("Price Note", "priceNote", "Text group → Price Note linked to price."),
    ("Callout", "callout", "Text group → Callout bubble."),
    ("Comment", "comment", "Text group → Comment."),
    ("Pin", "pin", "Text group → Pin marker."),
    ("Price Label", "priceLabel", "Text group → Price Label."),
    ("Signpost", "signpost", "Text group → Signpost."),
    ("Flag Mark", "flag", "Text group → Flag Mark."),
    ("Image", "image", "Text group → Image. Upload file or paste URL."),
    ("Emoji / Sticker", "emoji", "Text group → Emoji. Browse categories (Smileys, People, etc.)."),
    ("Anchored Text", "anchoredText", "If available: Anchored Text with anchor line."),
    ("Note Box", "notebox", "If available: Note Box with border."),
    ("Label / Marker", "label", "If available: Label marker tool."),
]

VOLUME_TOOLS = [
    ("Anchored VWAP", "vwap", "Volume group (bottom of rail) → Anchored VWAP. Anchor on a bar."),
    ("Fixed Range Volume Profile", "volProfile", "Volume group → Fixed Range VP. Select price/time range."),
    ("Anchored Volume Profile", "anchoredVol", "Volume group → Anchored VP."),
]

INDICATORS = [
    ("SMA", "Simple Moving Average overlay — change period in settings."),
    ("EMA", "Exponential Moving Average overlay."),
    ("WMA", "Weighted Moving Average."),
    ("DEMA", "Double EMA."),
    ("TEMA", "Triple EMA."),
    ("HMA", "Hull Moving Average."),
    ("Supertrend", "Supertrend overlay with ATR bands."),
    ("Bollinger Bands", "BB overlay — upper/mid/lower bands."),
    ("Donchian Channel", "DC — high/low channel."),
    ("Keltner Channel", "KC — ATR envelope."),
    ("Parabolic SAR", "PSAR dots on chart."),
    ("Aroon", "Aroon Up/Down in separate panel."),
    ("RSI", "RSI oscillator 0–100 panel. Test overbought/oversold levels in Style."),
    ("MACD", "MACD histogram + signal lines."),
    ("Stochastic", "Stochastic %K/%D oscillator."),
    ("Stochastic RSI", "Stochastic RSI — more sensitive."),
    ("CCI", "Commodity Channel Index."),
    ("Momentum", "Raw momentum oscillator."),
    ("Rate of Change", "ROC percentage oscillator."),
    ("Williams %R", "Williams %R oscillator."),
    ("DPO", "Detrended Price Oscillator."),
    ("Awesome Oscillator", "AO histogram."),
    ("Ultimate Oscillator", "UO multi-period."),
    ("TRIX", "TRIX triple-smoothed."),
    ("Coppock Curve", "Coppock long-term momentum."),
    ("RVI", "Relative Vigor Index."),
    ("Elder Ray", "Bull/Bear power lines."),
    ("Mass Index", "Mass Index reversal detector."),
    ("ATR", "Average True Range panel."),
    ("Standard Deviation", "StdDev volatility."),
    ("Vortex Indicator", "Vortex VI+ / VI-."),
    ("VWAP", "VWAP benchmark overlay."),
    ("Volume", "Volume bars + optional MA."),
    ("OBV", "On Balance Volume."),
    ("CMF", "Chaikin Money Flow."),
    ("MFI", "Money Flow Index."),
    ("ADX", "Average Directional Index."),
    ("ADR", "Average Daily Range."),
    ("Session Boxes", "Session shading for major sessions."),
    ("ICT Kill Zones", "ICT kill zone windows."),
    ("ICT Everything", "Full ICT bundle — sessions, FVG, etc."),
    ("Opening Range", "Opening range high/low."),
    ("ICT PD (Previous Day)", "If in ICT submenu: Previous Day levels."),
    ("ICT Asian Range", "If in menu: Asian range."),
    ("ICT OTE Zone", "If in menu: OTE zone."),
    ("ICT Fair Value Gaps", "If in menu: FVG boxes."),
    ("ICT Session PD", "If in menu: Session PD."),
    ("ICT Equal Highs/Lows", "If in menu: Equal highs/lows liquidity."),
    ("COT Net Commercial", "COT positioning (requires data feed)."),
    ("Seasonality", "If in menu: Seasonality overlay."),
    ("SMA Envelope", "If in menu: SMA Envelope bands."),
    ("Custom JS Indicator", "Custom/Script category — add sandbox indicator if enabled."),
]

CHART_TYPES = [
    ("Candlestick", "Top bar chart type → Candlestick. Body + wick visible."),
    ("Hollow Candles", "Chart type → Hollow Candles."),
    ("Heikin Ashi", "Chart type → Heikin Ashi smoothed candles."),
    ("Bar (OHLC)", "Chart type → Bar/OHLC ticks."),
    ("Line", "Chart type → Line on close."),
    ("Area", "Chart type → Area fill under line."),
    ("Baseline", "Chart type → Baseline two-tone fill."),
]

TIMEFRAMES = ["1m", "5m", "15m", "30m", "1h", "4h", "1D", "1W", "1M"]

LAYOUT_COUNTS = [
    (1, "Single full-width chart"),
    (2, "2 panels — test vertical AND horizontal variants"),
    (3, "3 panels — cycle through layout variants"),
    (4, "4 panels (2×2 grid)"),
    (5, "5 panels"),
    (6, "6 panels"),
    (7, "7 panels"),
    (8, "8 panels"),
]

SYNC_TOGGLES = [
    ("Symbol", "Same symbol on all panels when ON."),
    ("Interval", "Same timeframe on all panels when ON."),
    ("Crosshair", "Crosshair moves together across panels."),
    ("Time", "Pan/zoom time axis together."),
    ("Date Range", "Visible bar range matches."),
    ("Drawings", "Drawings sync across panels (verify expected behavior)."),
    ("Indicators", "Indicator changes apply to focused/synced panels."),
    ("Chart Type", "Candle/line type matches all panels."),
]


def item(category, name, steps, expected, priority="P1", precond="Backtest session loaded"):
    return (category, name, steps, expected, priority, precond)


def drawing_place_tests(group_name, tools):
    rows = []
    for name, _id, how in tools:
        rows.append(item(
            f"Drawing — {group_name}",
            f"{name} — PLACE",
            how,
            f"{name} appears on chart at clicked coordinates; no error toast.",
            "P0" if name in ("Trend Line", "Fib Retracement", "Rectangle", "Text") else "P1",
        ))
        rows.append(item(
            f"Drawing — {group_name}",
            f"{name} — SELECT & MOVE",
            f"After placing {name}: click it → handles appear. Drag body to new location.",
            "Selection handles visible; drawing moves smoothly; price/time updates.",
        ))
        rows.append(item(
            f"Drawing — {group_name}",
            f"{name} — QUICK TOOLBAR",
            f"Select {name} → floating quick toolbar → change color, line style, or thickness.",
            "Visual update on chart immediately.",
        ))
        rows.append(item(
            f"Drawing — {group_name}",
            f"{name} — SETTINGS DIALOG",
            f"Select {name} → gear/settings → open Style tab. Change color, width, dash. Check Input/Coordinates/Visibility tabs if present.",
            "All tabs open without error; changes apply on OK/close.",
        ))
        rows.append(item(
            f"Drawing — {group_name}",
            f"{name} — DELETE",
            f"Select {name} → press Delete/Backspace OR use eraser OR trash menu.",
            "Drawing removed from chart and Objects Tree.",
        ))
    return rows


def indicator_tests():
    rows = []
    rows.append(item(
        "Indicators — Menu",
        "Open indicator search ( / or Indicators button )",
        "Press / or click Indicators in top bar.",
        "Search panel opens with categories: Favorites, Technicals, Sessions, ICT, Custom.",
        "P0",
    ))
    rows.append(item(
        "Indicators — Menu",
        "Search filter works",
        "Type 'RSI' in search box.",
        "List filters to matching indicators only.",
    ))
    rows.append(item(
        "Indicators — Menu",
        "Star / Favorite indicator",
        "Star an indicator → check Favorites category.",
        "Starred indicator appears under Favorites.",
    ))
    rows.append(item(
        "Indicators — Menu",
        "Max 10 active indicators",
        "Add 10 indicators, then try an 11th.",
        "Warning or block at 10; chart stays stable.",
        "P0",
    ))
    for name, note in INDICATORS:
        rows.append(item(
            "Indicators — Add & Display",
            f"{name} — ADD",
            f"Indicators menu → find {name} → Add. {note}",
            f"{name} renders on chart (overlay or lower panel as designed). Legend chip shows name + values.",
            "P0" if name in ("RSI", "SMA", "EMA", "MACD", "Volume", "VWAP") else "P1",
        ))
        rows.append(item(
            "Indicators — Settings",
            f"{name} — INPUT settings",
            f"Click {name} legend/settings → Input tab → change period/length/source (e.g. Close vs HL2).",
            "Indicator recalculates; line/bars update without crash.",
        ))
        rows.append(item(
            "Indicators — Settings",
            f"{name} — STYLE settings",
            f"{name} settings → Style tab → change line color, thickness, plot type (Line/Histogram/Area).",
            "Visual style updates on chart.",
        ))
        rows.append(item(
            "Indicators — Manage",
            f"{name} — HIDE / SHOW",
            f"Hide {name} via eye menu or legend; show again.",
            "Indicator hidden then restored correctly.",
        ))
        rows.append(item(
            "Indicators — Manage",
            f"{name} — REMOVE",
            f"Remove {name} via trash menu or legend X.",
            "Fully removed; legend gone; panel collapses if empty.",
        ))
    rows.append(item(
        "Indicators — Bulk",
        "Hide all indicators",
        "Visibility menu → Hide Indicators.",
        "All indicators hidden; chart price still visible.",
    ))
    rows.append(item(
        "Indicators — Bulk",
        "Delete all indicators",
        "Trash menu → Delete Indicators only.",
        "All indicators removed; drawings untouched.",
    ))
    rows.append(item(
        "Indicators — Timeframe",
        "Indicators survive timeframe change",
        "With 3+ indicators active, switch 15m → 1h → 1D.",
        "No blank chart; values recalculate.",
        "P0",
    ))
    return rows


PHASES = []


def add_phase(num, name, goal, items):
    PHASES.append({"num": num, "name": name, "goal": goal, "items": items})


# ═══ PHASE 1 — Access & Main Chrome ═══
add_phase(1, "Access, Loading & Main Chrome", "Confirm login, session load, and every major button/panel in the shell opens correctly.", [
    item("Access", "Open chart from backtest session", "From Sessions/Backtest page, click Start/Resume on a session.", "Chart page loads; loader disappears; no blank white screen.", "P0"),
    item("Access", "Loader screen", "Watch loading screen on first open.", "Logo, progress bar, and quote display; then chart appears.", "P1"),
    item("Access", "Session HUD — symbol", "Check top bar shows current pair/symbol.", "Symbol name matches session (e.g. EUR/USD).", "P0"),
    item("Access", "Session HUD — balance/equity", "Check balance and equity display in header.", "Numbers show; update after trades.", "P0"),
    item("Access", "Session HUD — hide balance toggle", "Toggle eye icon to hide account values.", "Balance masks to dots when hidden.", "P1"),
    item("Access", "Session HUD — P&L display", "After a closed trade, check session P&L in HUD.", "P&L reflects closed trade result.", "P1"),
    item("Access", "Logo menu opens", "Click Talaria logo (top-left).", "Dropdown menu opens with Profile, Settings, FAQ, etc.", "P0"),
    item("Access", "Logo menu — Profile", "Logo menu → Profile.", "Profile panel/modal opens (account, billing, password).", "P1"),
    item("Access", "Logo menu — Settings", "Logo menu → Settings.", "Settings drawer/modal opens.", "P0"),
    item("Access", "Logo menu — FAQ / Help", "Logo menu → FAQ or Help.", "FAQ panel with searchable questions opens.", "P2"),
    item("Access", "Logo menu — Education", "Logo menu → Education (if listed).", "Education/lesson cards open.", "P2"),
    item("Access", "Support chat widget", "Click support/chat icon (if visible).", "Support chat panel opens; can view threads.", "P2"),
    item("Access", "Theme toggle", "Toggle dark theme in sidebar/logo area.", "UI stays readable; colors update.", "P1"),
    item("Access", "Fullscreen button", "Right sidebar → Expand/Fullscreen.", "Chart fills screen; exit restores layout.", "P1"),
    item("Access", "Bottom tab bar — All trades", "Click bottom tab 'All' or trades tab.", "Trade list/table opens below chart.", "P0"),
    item("Access", "Bottom tab bar — Analytics", "Click Analytics tab.", "Journal analytics / stats panel loads.", "P1"),
    item("Access", "Bottom tab bar — resize", "Drag divider between chart and bottom panel.", "Panel height resizes smoothly.", "P2"),
    item("Access", "Right sidebar — Place Order", "Click Place Order button (blue).", "Order panel opens on right (336px).", "P0"),
    item("Access", "Right sidebar — Layouts", "Click Layouts icon.", "Layouts panel opens (not order panel).", "P0"),
    item("Access", "Right sidebar — News", "Click News icon.", "News / economic calendar panel opens.", "P1"),
    item("Access", "Right sidebar — Objects Tree", "Click Objects Tree (layers) icon.", "Objects Tree panel lists drawings/indicators.", "P0"),
    item("Access", "Right sidebar — close panel", "Open any right panel → click X or toggle same icon.", "Panel closes; chart expands.", "P1"),
    item("Access", "Top bar — Screenshot camera", "Click camera icon in top bar.", "Screenshot capture panel opens.", "P1"),
    item("Access", "Top bar — Alerts", "Click alerts/bell icon (if present).", "Alerts list or create-alert UI opens.", "P1"),
    item("Access", "Top bar — Compare", "Click Compare or + overlay button.", "Compare symbol modal opens.", "P1"),
    item("Access", "Browser refresh persistence", "Draw one line, refresh page (F5).", "Session reloads; drawings restore (or cloud save prompt).", "P0"),
    item("Access", "PWA install prompt", "If browser shows install app prompt, install.", "App opens standalone without browser chrome.", "P2"),
])

# ═══ PHASE 2 — Symbol, Timeframe & Chart Type ═══
p2 = [
    item("Symbol", "Symbol search dropdown opens", "Click symbol name in top bar.", "Search list of session pairs opens.", "P0"),
    item("Symbol", "Switch symbol in session", "Pick different pair from list (e.g. EUR/USD → GBP/USD).", "Chart reloads new data; symbol badge updates.", "P0"),
    item("Symbol", "Type-to-search symbol", "With chart focused, type letters (e.g. GBP).", "Symbol search buffer opens/filters.", "P2"),
    item("Symbol", "Quick search Ctrl+K", "Press Ctrl+K.", "Quick search overlay opens.", "P1"),
]
for ct, steps in CHART_TYPES:
    p2.append(item("Chart Type", ct, steps, f"{ct} renders correctly with proper colors/shape.", "P0" if ct == "Candlestick" else "P1"))
p2 += [
    item("Timeframe", "Timeframe dropdown opens", "Click timeframe badge (e.g. 1D) in top bar.", "Dropdown with Minutes/Hours/Days/Weeks/Months sections.", "P0"),
]
for tf in TIMEFRAMES:
    p2.append(item("Timeframe", f"Switch to {tf}", f"Select {tf} from dropdown.", f"Chart resamples to {tf}; candles load without gap.", "P0" if tf in ("15m", "1h", "1D") else "P1"))
p2 += [
    item("Timeframe", "Favorite timeframe (star)", "Star a timeframe in dropdown.", "Starred TF appears in favorites/quick access.", "P1"),
    item("Timeframe", "Add custom timeframe", "Add custom TF (e.g. 7m or 3H) if option exists.", "Custom TF works and persists.", "P2"),
    item("Timeframe", "Type interval shortcut", "Type '15,' or digits to change interval.", "Interval changes to typed value.", "P2"),
    item("Timeframe", "TF change during replay", "During replay, switch 15m → 1h.", "Replay continues; data resamples without freeze.", "P0"),
    item("Compare", "Compare modal — open", "Open Compare modal from top bar.", "Modal with symbol search list.", "P1"),
    item("Compare", "Compare modal — add symbol", "Search and add second symbol as overlay.", "Second line/series on chart (same scale).", "P1"),
    item("Compare", "Compare modal — remove symbol", "Remove compare overlay from modal or legend.", "Overlay disappears.", "P1"),
    item("Compare", "Compare — new pane settings", "If new-pane overlay settings exist, open and change colors.", "Overlay colors update.", "P2"),
]
add_phase(2, "Symbol, Timeframe & Chart Type", "Every top-bar control for symbol, interval, and chart type.", p2)

# ═══ PHASE 3 — Chart Display & Settings Modal ═══
add_phase(3, "Chart Display, Scales & Settings", "Every visual option and every Settings section/tab.", [
    item("Price Scale", "Linear scale", "Settings or scale menu → Normal/Linear (Alt+I off for log).", "Y-axis linear spacing.", "P0"),
    item("Price Scale", "Log scale", "Enable Log scale (Alt+L).", "Logarithmic compression at high prices.", "P1"),
    item("Price Scale", "Percent scale", "Enable Percent (Alt+P).", "% change from reference visible.", "P1"),
    item("Price Scale", "Auto-scale toggle", "Toggle Auto-scale (Alt+A).", "Manual Y drag sticks when off; auto fits when on.", "P0"),
    item("Price Scale", "Invert scale", "Invert scale (Alt+I).", "High prices at bottom.", "P2"),
    item("Price Scale", "Drag price scale", "Drag Y-axis to stretch/compress.", "Vertical zoom works.", "P0"),
    item("Price Scale", "Drag time scale", "Drag X-axis.", "Horizontal zoom works.", "P0"),
    item("Display", "Background color", "Settings → Chart → Background color.", "Chart background updates live.", "P1"),
    item("Display", "Grid lines on/off", "Settings → Grid toggle.", "Grid shows/hides.", "P1"),
    item("Display", "Grid color/style/thickness", "Settings → Grid style options.", "Grid appearance changes.", "P2"),
    item("Display", "Crosshair on/off", "Settings → Crosshair toggle.", "Crosshair shows/hides on hover.", "P0"),
    item("Display", "Crosshair color/style", "Settings → Crosshair color and line style.", "Crosshair appearance updates.", "P2"),
    item("Display", "Crosshair lock", "Enable crosshair lock if available.", "Crosshair stays fixed while panning.", "P2"),
    item("Display", "Last price line", "Toggle last price line.", "Dashed line at current price.", "P1"),
    item("Display", "Candle body colors", "Settings → Bull/Bear body colors.", "Up/down candles recolor.", "P1"),
    item("Display", "Candle border colors", "Settings → Border colors.", "Borders recolor.", "P2"),
    item("Display", "Candle wick colors", "Settings → Wick colors.", "Wicks recolor.", "P2"),
    item("Display", "Show/hide wick/body/border", "Toggle each candle part off/on.", "Parts hide/show independently.", "P2"),
    item("Display", "Volume colors", "Settings → Volume up/down colors.", "Volume bars recolor.", "P2"),
    item("Display", "OHLC legend bar", "Enable OHLC values on hover.", "O/H/L/C and change % display.", "P1"),
    item("Display", "Symbol title format", "Settings → Title: Description / Ticker / Both.", "Header label format changes.", "P2"),
    item("Display", "Scale text color & size", "Settings → Scale font size/color.", "Axis labels update.", "P2"),
    item("Display", "Right offset (padding)", "Settings → Right offset bars.", "Last candle moves from right edge.", "P2"),
    item("Display", "Decimal precision", "Settings → Precision / decimals.", "Prices show correct decimals.", "P1"),
    item("Display", "Timezone (Settings)", "Settings → Time zone dropdown (26 zones).", "X-axis times shift to selected zone.", "P0"),
    item("Display", "12h vs 24h time", "Settings → Time format.", "Axis labels switch format.", "P1"),
    item("Display", "Watermark / brand link", "Check bottom-left watermark (hidden in multichart tiles).", "Logo/link visible in single chart.", "P2"),
    item("Settings — General", "Open General section", "Settings → General.", "General section loads.", "P0"),
    item("Settings — General", "Scroll speed slider", "Adjust scroll speed.", "Pan/wheel feels faster/slower.", "P1"),
    item("Settings — General", "Mouse sensitivity", "Adjust mouse sensitivity.", "Hit detection / interaction changes.", "P2"),
    item("Settings — General", "Magnet sensitivity", "Adjust magnet sensitivity.", "Snap-to-OHLC strength changes.", "P2"),
    item("Settings — General", "Magnet crosshair mode", "Toggle magnet crosshair.", "Magnet behavior updates.", "P2"),
    item("Settings — General", "Order placement — Instant", "Set Instant order mode.", "Orders fire without confirmation modal.", "P0"),
    item("Settings — General", "Order placement — Confirmation", "Set Confirmation mode.", "Modal appears before each order.", "P0"),
    item("Settings — General", "Show order history on chart", "Toggle show order history.", "Closed trade markers show/hide.", "P1"),
    item("Settings — General", "Show open orders on chart", "Toggle show open orders.", "Open position lines show/hide.", "P0"),
    item("Settings — General", "Sound on execution + volume", "Enable sound; adjust volume slider.", "Fill plays sound at set volume.", "P1"),
    item("Settings — Chart", "Open Chart section", "Settings → Chart.", "All chart visual options listed.", "P0"),
    item("Settings — Project", "Open Project section", "Settings → Project.", "Timezone, DST, session close, initial balance.", "P1"),
    item("Settings — Project", "Initial balance save", "Change initial balance → Save.", "Balance updates in HUD.", "P1"),
    item("Settings — Project", "Session close time", "Set session close time preset.", "Setting saves without error.", "P2"),
    item("Settings — Leverage", "Open Leverage section", "Settings → Leverage.", "Leverage slider 1:1 to 1:500.", "P1"),
    item("Settings — Leverage", "Save leverage", "Set 1:100 → Save.", "Margin calculations use new leverage.", "P1"),
    item("Settings — Symbol", "Open Symbol Properties", "Settings → Symbol Properties.", "Symbol-specific pip/lot settings.", "P1"),
    item("Settings — Commissions", "Open Commissions section", "Settings → Commissions.", "Spread, slippage, commission, swap fields.", "P1"),
    item("Settings — Commissions", "Apply commission on trade", "Set commission → place/close trade.", "Commission deducted from balance.", "P0"),
    item("Settings — Alerts", "Open Alerts section", "Settings → Alerts.", "Alert defaults and sound settings.", "P2"),
    item("Settings — Templates", "Save chart template", "Settings → Save as template.", "Template appears in custom list.", "P1"),
    item("Settings — Templates", "Load chart template", "Load saved template.", "Colors/type/settings restore.", "P1"),
    item("Settings — Persist", "Settings survive reload", "Change 3 settings → refresh browser.", "All settings restored.", "P0"),
])

# ═══ PHASE 4 — Navigation & Shortcuts ═══
add_phase(4, "Navigation, Go To & Keyboard Shortcuts", "Pan, zoom, Go To menu, and all keyboard shortcuts.", [
    item("Pan/Zoom", "Mouse drag pan", "Click-drag empty chart area left/right.", "Chart scrolls through history.", "P0"),
    item("Pan/Zoom", "Mouse wheel zoom", "Scroll wheel over chart.", "Time axis zooms in/out.", "P0"),
    item("Pan/Zoom", "Scroll speed feel", "Compare slow vs fast scroll setting.", "Noticeable speed difference.", "P2"),
    item("Pan/Zoom", "Inertia / coast", "Pan quickly and release.", "Chart coasts slightly (if enabled).", "P2"),
    item("Keyboard", "← / → one bar", "Press arrow keys.", "Moves one bar.", "P0"),
    item("Keyboard", "Ctrl+← / Ctrl+→ ten bars", "Hold Ctrl + arrows.", "Jumps ~10 bars.", "P1"),
    item("Keyboard", "Ctrl+↑/↓ zoom", "Ctrl + up/down.", "Zoom in/out.", "P1"),
    item("Keyboard", "+ / - zoom", "Press + and - keys.", "Zoom in/out.", "P1"),
    item("Keyboard", "Home — latest bar", "Press Home.", "Jumps to latest bar.", "P0"),
    item("Keyboard", "End — dataset start", "Press End.", "Jumps to oldest loaded bar.", "P1"),
    item("Keyboard", "Alt+R reset chart", "Press Alt+R.", "View resets to default.", "P1"),
    item("Keyboard", "Alt+G Go To date", "Press Alt+G.", "Go To menu opens.", "P1"),
    item("Keyboard", "Alt+S snapshot", "Press Alt+S.", "Screenshot panel opens.", "P2"),
    item("Keyboard", "Alt+Z keyboard nav mode", "Press Alt+Z.", "Keyboard navigation mode toggles.", "P2"),
    item("Keyboard", "Ctrl+S save layout", "Press Ctrl+S.", "Layout save OR screenshot (verify which in your build).", "P1"),
    item("Keyboard", "Ctrl+L load layout", "Press Ctrl+L.", "Load layout dialog/list.", "P1"),
    item("Keyboard", "Ctrl+Z / Ctrl+Y undo redo", "Draw line → Ctrl+Z → Ctrl+Y.", "Undo removes; redo restores.", "P0"),
    item("Keyboard", "Escape cancel drawing", "Start drawing → Esc.", "Placement cancels.", "P0"),
    item("Keyboard", "? shortcuts help", "Press ?.", "Shortcuts overlay lists all bindings.", "P1"),
    item("Keyboard", "/ open indicators", "Press /.", "Indicator search opens.", "P0"),
    item("Keyboard", "Drawing shortcuts T H V R F", "Press T, H, V, R, F.", "Activates Trend, Horizontal, Vertical, Rectangle, Fib.", "P1"),
    item("Keyboard", "Custom shortcut rebind", "Shortcuts help → edit a binding → save.", "New binding works; persists after reload.", "P2"),
    item("Go To", "Open Go To menu", "Top bar → Go To or Alt+G.", "Go To panel with tabs opens.", "P0"),
    item("Go To", "Calendar date picker", "Pick a date on calendar.", "Chart jumps to that date.", "P0"),
    item("Go To", "Preset — NY Open", "Select NY Open preset.", "Chart jumps to NY session open.", "P1"),
    item("Go To", "Preset — London Open", "Select London Open.", "Chart jumps correctly.", "P1"),
    item("Go To", "Preset — Tokyo Open", "Select Tokyo Open.", "Chart jumps correctly.", "P1"),
    item("Go To", "Preset — Sydney Open", "Select Sydney Open.", "Chart jumps correctly.", "P2"),
    item("Go To", "Preset — Day/Week/Month open", "Select Day open, Week open, Month open.", "Each jumps to correct boundary.", "P1"),
    item("Go To", "Preset — Previous day H/L", "Select Previous day high/low.", "Chart navigates to PDH/PDL.", "P1"),
    item("Go To", "Preset — Asian/London/NY session", "Select session presets.", "Session boundaries correct.", "P1"),
    item("Go To", "Pin bookmark", "Pin current datetime → revisit from list.", "Bookmark saves and jumps back.", "P1"),
    item("Go To", "Repeat rule daily/weekly", "Set repeat on bookmark.", "Repeat behavior works on replay advance.", "P2"),
    item("Go To", "Go to price level", "If price tab exists: enter price.", "Chart scrolls to that price.", "P2"),
])

# ═══ PHASE 5 — Drawing utilities & Lines ═══
p5 = [
    item("Drawing — Cursor", "Crosshair cursor (default)", "Left rail → Crosshair.", "Crosshair on hover.", "P0"),
    item("Drawing — Cursor", "Dot cursor", "Left rail → Dot cursor.", "Dot cursor mode active.", "P2"),
    item("Drawing — Cursor", "Arrow cursor", "Left rail → Arrow cursor.", "Arrow cursor mode active.", "P2"),
    item("Drawing — Magnet", "Magnet Off", "Magnet button → Off (or M key).", "Free placement without snap.", "P1"),
    item("Drawing — Magnet", "Magnet Weak", "Magnet → Weak.", "Mild snap to OHLC.", "P1"),
    item("Drawing — Magnet", "Magnet Strong", "Magnet → Strong.", "Strong snap to OHLC.", "P1"),
    item("Drawing — Magnet", "Temporary magnet Ctrl", "Hold Ctrl while placing line.", "Temporary strong snap.", "P2"),
    item("Drawing — Utility", "Eraser tool", "Select Eraser → click drawing.", "Clicked drawing deleted.", "P0"),
    item("Drawing — Utility", "Lock all drawings", "Lock all → try to move → Ctrl+U unlock.", "Locked until unlocked.", "P1"),
    item("Drawing — Utility", "Keep drawing mode", "Enable keep-drawing; place 3 lines without re-selecting tool.", "Tool stays armed.", "P1"),
    item("Drawing — Utility", "Undo / Redo", "Place → Ctrl+Z → Ctrl+Y.", "Undo/redo works.", "P0"),
    item("Drawing — Utility", "Visibility — Hide Drawings", "Eye menu → Hide Drawings.", "All drawings hidden.", "P1"),
    item("Drawing — Utility", "Visibility — Hide Indicators", "Eye menu → Hide Indicators.", "Indicators hidden.", "P1"),
    item("Drawing — Utility", "Visibility — Hide Positions", "Eye menu → Hide Positions.", "Order lines hidden.", "P1"),
    item("Drawing — Utility", "Visibility — Hide All", "Eye menu → Hide All.", "Everything hidden.", "P1"),
    item("Drawing — Utility", "Delete — Drawings only", "Trash → Delete drawings only.", "Drawings gone; indicators remain.", "P1"),
    item("Drawing — Utility", "Delete — Indicators only", "Trash → Delete indicators only.", "Indicators gone; drawings remain.", "P1"),
    item("Drawing — Utility", "Delete — All objects", "Trash → Delete all.", "All objects removed.", "P1"),
    item("Drawing — Utility", "Pinned favorites bar", "Star a tool → check favorites bar.", "Tool appears in pinned bar.", "P1"),
    item("Drawing — Utility", "Multi-select rectangle", "Drag rectangle on empty canvas.", "Multiple drawings selected.", "P1"),
    item("Drawing — Utility", "Shift+click stacked lines", "Stack 3+ lines → Shift+click.", "All stacked lines selected.", "P2"),
    item("Drawing — Utility", "Copy / Paste drawing", "Select → copy → paste.", "Duplicate appears nearby.", "P1"),
    item("Drawing — Utility", "Clone drawing", "Clone from context/quick menu if available.", "Clone created.", "P2"),
    item("Drawing — Utility", "Drawing template save/apply", "Style a trend line → save template → apply to new line.", "Template applies colors/width.", "P1"),
    item("Drawing — Utility", "Per-drawing timeframe visibility", "Drawing settings → Visibility → show only on 1h.", "Drawing hides on other TFs.", "P2"),
    item("Drawing — Utility", "Per-drawing lock", "Lock single drawing from quick bar.", "That drawing cannot move.", "P1"),
]
p5 += drawing_place_tests("Lines", LINE_TOOLS)
add_phase(5, "Drawing Tools — Utilities & Lines", "Cursor, magnet, visibility, undo, and every line tool (place/edit/delete).", p5)

# ═══ PHASE 6 — Shapes, Channels, Brushes ═══
p6 = drawing_place_tests("Shapes", SHAPE_TOOLS)
p6 += drawing_place_tests("Channels", CHANNEL_TOOLS)
p6 += drawing_place_tests("Brushes", BRUSH_TOOLS)
p6.append(item("Drawing — Shapes", "Rotated Rectangle", "If in Shapes submenu: Rotated Rectangle.", "Rotated box places and rotates via handles.", "P2"))
add_phase(6, "Drawing Tools — Shapes, Channels & Brushes", "Every shape, channel, pitchfork variant, and brush tool.", p6)

# ═══ PHASE 7 — Fib, Gann, Patterns, Measure ═══
p7 = drawing_place_tests("Fibonacci & Gann", FIB_GANN_TOOLS)
p7 += drawing_place_tests("Patterns & Elliott", PATTERN_TOOLS)
p7 += drawing_place_tests("Measure & Positions", MEASURE_TOOLS)
p7 += [
    item("Drawing — Fib levels editor", "Fib Retracement — edit levels list", "Select Fib → Settings → Input → enable/disable levels, change % values.", "Levels show/hide; custom % work.", "P0"),
    item("Drawing — Fib levels editor", "Fib — change level colors", "Fib settings → Style → per-level colors.", "Each level recolors.", "P1"),
    item("Drawing — Gann editor", "Gann Box — grid levels", "Select Gann Box → edit price/time grid levels.", "Grid updates correctly.", "P1"),
    item("Drawing — Ruler tool", "Ruler / Measure (engine)", "If Ruler in menu: measure distance.", "Bar count, price, time stats show.", "P2"),
]
add_phase(7, "Drawing Tools — Fibonacci, Gann, Patterns & Measure", "All advanced drawing tools traders use daily.", p7)

# ═══ PHASE 8 — Text, Volume, Edit UI, Objects Tree ═══
p8 = drawing_place_tests("Text & Annotations", TEXT_TOOLS)
p8 += drawing_place_tests("Volume Tools", VOLUME_TOOLS)
p8 += [
    item("Drawing — Text", "Text — Bold / Italic", "Select text → quick bar Bold/Italic.", "Formatting applies.", "P1"),
    item("Drawing — Text", "Text — Font size", "Change font size in text settings.", "Size updates on chart.", "P1"),
    item("Drawing — Text", "Text — Arabic/RTL", "Type Arabic text in Text tool.", "RTL renders correctly.", "P2"),
    item("Drawing — Emoji", "Emoji — each category", "Emoji tool → browse Smileys, People, Animals, Food, Activity, Travel, Objects, Symbols.", "Emoji from each category places on chart.", "P2"),
    item("Drawing — Image", "Image — opacity", "Place image → reduce opacity in settings.", "Image fades correctly.", "P2"),
    item("Drawing — Settings Dialog", "Style tab — line color picker", "Any line drawing → Settings → Style → color.", "Color updates.", "P0"),
    item("Drawing — Settings Dialog", "Style tab — line width 1-4", "Change thickness.", "Width updates.", "P0"),
    item("Drawing — Settings Dialog", "Style tab — dash solid/dashed/dotted", "Change line type.", "Dash pattern updates.", "P1"),
    item("Drawing — Settings Dialog", "Input tab — tool-specific fields", "Fib/Channel/VP tools → Input tab fields.", "Fields editable; chart updates.", "P1"),
    item("Drawing — Settings Dialog", "Coordinates tab — price/time", "Edit exact price and bar index.", "Drawing jumps to exact coordinates.", "P1"),
    item("Drawing — Settings Dialog", "Visibility tab — timeframes", "Check/uncheck timeframe boxes.", "Drawing shows only on checked TFs.", "P2"),
    item("Drawing — Settings Dialog", "Custom display name", "Rename drawing in settings or tree.", "Custom name shows in Objects Tree.", "P2"),
    item("Drawing — Floating Bar", "Quick bar — color", "Select drawing → quick bar color.", "Instant recolor.", "P0"),
    item("Drawing — Floating Bar", "Quick bar — lock/delete", "Quick bar lock and trash icons.", "Lock/delete work.", "P0"),
    item("Drawing — Floating Bar", "Quick bar — clone/copy", "Clone from more menu on quick bar.", "Duplicate created.", "P2"),
    item("Drawing — Floating Bar", "More menu → Object Tree", "More menu → Open Object Tree.", "Objects Tree panel opens.", "P1"),
    item("Objects Tree", "Panel opens", "Right sidebar → Objects Tree.", "List of all drawings/indicators.", "P0"),
    item("Objects Tree", "Search filter", "Type name in search box.", "List filters.", "P1"),
    item("Objects Tree", "Eye toggle per item", "Click eye on one item.", "Item hides/shows on chart.", "P0"),
    item("Objects Tree", "Select item from tree", "Click item name in tree.", "Selects on chart with handles.", "P0"),
    item("Objects Tree", "Rename item", "Double-click or edit name.", "Name updates.", "P2"),
    item("Objects Tree", "Delete from tree", "Delete icon on tree item.", "Removed from chart.", "P1"),
    item("Objects Tree", "Lists open orders", "With open trade, check tree includes orders.", "Orders visible in tree.", "P1"),
    item("Drawings — Persist", "Save layout Ctrl+S", "Draw 5 objects → Ctrl+S save layout.", "Save confirmation or indicator.", "P0"),
    item("Drawings — Persist", "Load layout Ctrl+L", "Load saved layout.", "All objects restore.", "P0"),
    item("Drawings — Persist", "Cloud save indicator", "After edit, check save-to-cloud indicator.", "Save status shows synced/pending.", "P2"),
]
add_phase(8, "Text, Volume Tools, Edit UI & Objects Tree", "Annotations, volume profile, settings dialog, floating bar, and object manager.", p8)

# ═══ PHASE 9 — Indicators ═══
add_phase(9, "Indicators — Full Catalog", "Every indicator in the menu: add, configure, hide, remove.", indicator_tests())

# ═══ PHASE 10 — Layouts & Multichart ═══
p10 = [
    item("Layouts Panel", "Open Layouts panel", "Right sidebar → Layouts.", "Layout grid picker opens.", "P0"),
]
for n, desc in LAYOUT_COUNTS:
    p10.append(item("Layouts", f"{n}-panel layout", f"Layouts panel → select {n}-panel layout. {desc}", f"{n} chart tiles render with data.", "P0" if n <= 4 else "P1"))
p10 += [
    item("Layouts", "Cycle layout variants", "For 3-panel: click through all variant icons.", "Each variant arranges tiles differently.", "P1"),
    item("Layouts", "Focus panel click", "Click each tile.", "Focused tile highlighted; tools apply to it.", "P0"),
    item("Layouts", "Per-panel symbol badge", "Each tile shows its symbol.", "Badges correct per tile.", "P1"),
    item("Layouts", "Independent symbol per panel", "Symbol sync OFF → different symbol per tile.", "Each tile loads different pair.", "P1"),
    item("Layouts", "Independent timeframe per panel", "Interval sync OFF → 1h + 15m + 1D.", "Each tile different TF.", "P0"),
    item("Layouts", "Resize panel dividers", "Drag dividers between tiles.", "Smooth resize.", "P1"),
    item("Layouts", "Return to 1 panel", "Select 1-panel layout.", "Single chart; no ghost tiles.", "P0"),
]
for name, desc in SYNC_TOGGLES:
    p10.append(item("Layouts — Sync", f"Sync toggle — {name}", f"Layouts panel → toggle {name} sync ON/OFF. {desc}", f"Sync behavior matches description.", "P0" if name in ("Symbol", "Interval", "Crosshair") else "P1"))
p10 += [
    item("Layouts — Persist", "Save multichart layout", "Set 2×2 + drawings → save layout.", "Layout saves.", "P0"),
    item("Layouts — Persist", "Load multichart layout", "Load saved layout.", "Panels + objects restore.", "P0"),
    item("Layouts — Replay", "Replay sync across tiles", "4-panel + replay play.", "Playhead advances on all tiles together.", "P0"),
    item("Layouts — Drawing", "Drawing on focused panel", "Multichart → draw on panel B only.", "Drawing appears on correct tile.", "P0"),
    item("Layouts — Indicators", "Different indicators per panel", "Sync OFF → SMA on A, RSI on B.", "Each panel correct indicator set.", "P1"),
]
add_phase(10, "Layouts & Multi-Chart", "Every layout count, variant, sync toggle, and persistence.", p10)

# ═══ PHASE 11 — Orders & Trading ═══
p11 = [
    item("Order Panel", "Open Place Order panel", "Click Place Order (blue sidebar button).", "Order panel opens on right.", "P0"),
    item("Order Panel", "Close order panel", "Click Place Order again or X.", "Panel closes.", "P1"),
    item("Order Panel", "Detach order panel", "If detach option exists, detach panel.", "Panel floats/docks correctly.", "P2"),
    item("Order Panel", "BUY side selected", "Click BUY tab.", "Long direction; green styling.", "P0"),
    item("Order Panel", "SELL side selected", "Click SELL tab.", "Short direction; red styling.", "P0"),
    item("Order Panel", "Order type — Market", "Select Market.", "No entry price required for instant fill.", "P0"),
    item("Order Panel", "Order type — Limit", "Select Limit.", "Entry price editable; pending until filled.", "P0"),
    item("Order Panel", "Order type — Stop", "Select Stop.", "Stop entry above/below market per side.", "P0"),
    item("Order Panel", "Size mode — Dollar ($)", "Switch size to $ mode.", "Risk/margin preview in dollars.", "P0"),
    item("Order Panel", "Size mode — Percent (%)", "Switch size to %.", "Size as % of balance.", "P0"),
    item("Order Panel", "Size mode — Lots (#)", "Switch size to lots/contracts.", "Lot count editable.", "P0"),
    item("Order Panel", "Size stepper +/−", "Use +/- stepper.", "Size increments/decrements.", "P1"),
    item("Order Panel", "Entry price manual edit", "Limit order → edit entry price field.", "Price updates; chart preview line moves.", "P0"),
    item("Order Panel", "Stop Loss toggle + price", "Enable SL → set price.", "SL distance and risk preview show.", "P0"),
    item("Order Panel", "Take Profit toggle + price", "Enable TP → set price.", "Reward preview shows.", "P0"),
    item("Order Panel", "Risk/Reward bar visual", "Set SL and TP.", "Risk vs reward bar proportional.", "P1"),
    item("Order Panel", "Reward / Risk / Margin summary", "Check summary numbers.", "Values match SL/TP/size.", "P0"),
    item("Order Panel", "Advanced section expand", "Toggle Advanced order.", "Multi-entry, BE, trailing UI expands.", "P1"),
    item("Order Panel", "Place order — confirmation modal", "Confirmation mode → Place Order.", "Modal asks confirm before send.", "P0"),
    item("Order Panel", "Place order — instant", "Instant mode → Place Order.", "Order fires immediately.", "P0"),
    item("Orders — Market", "Market BUY fill", "Replay → Market BUY.", "Long opens; entry line on chart.", "P0"),
    item("Orders — Market", "Market SELL fill", "Replay → Market SELL.", "Short opens; entry line on chart.", "P0"),
    item("Orders — Pending", "Limit BUY below market", "Place limit buy below price → advance replay.", "Fills when price touches.", "P0"),
    item("Orders — Pending", "Limit SELL above market", "Place limit sell above → advance.", "Fills at limit.", "P0"),
    item("Orders — Pending", "Stop BUY above market", "Stop buy above → advance.", "Triggers on cross.", "P0"),
    item("Orders — Pending", "Stop SELL below market", "Stop sell below → advance.", "Triggers on cross.", "P0"),
    item("Orders — Chart lines", "Entry line label", "After fill, check entry line P&L label.", "Entry price and side shown.", "P0"),
    item("Orders — Chart lines", "Drag SL on chart", "Drag SL line.", "SL price updates in panel.", "P0"),
    item("Orders — Chart lines", "Drag TP on chart", "Drag TP line.", "TP price updates in panel.", "P0"),
    item("Orders — Advanced", "Multi-entry / scale-in", "Advanced → multiple entry levels.", "Weighted avg entry line on chart.", "P1"),
    item("Orders — Advanced", "Multi-TP partial exits", "Set 2+ TP levels with size split.", "Partial closes at each TP.", "P1"),
    item("Orders — Advanced", "Breakeven rule", "Enable BE (RR or pips mode) → advance.", "SL moves to BE after trigger.", "P1"),
    item("Orders — Advanced", "Trailing stop", "Enable trailing → advance favorable move.", "Trail follows price; exits on reversal.", "P1"),
    item("Orders — Close", "Close at market", "Close full position from chart or panel.", "Position closed; P&L recorded.", "P0"),
    item("Orders — Close", "Partial close", "Partial close 50%.", "Size halved; marker on chart.", "P1"),
    item("Orders — Exit", "TP hit — full close", "Advance until TP.", "Position closes; TP marker.", "P0"),
    item("Orders — Exit", "SL hit — full close", "Advance until SL.", "Loss recorded; SL marker.", "P0"),
    item("Orders — Exit", "Stop-out / margin call", "Over-leverage until stop-out.", "Forced liquidation with marker.", "P1"),
    item("Orders — Account", "Balance updates", "After trade, check balance.", "Balance = previous + P&L - costs.", "P0"),
    item("Orders — Account", "Equity updates", "With open trade, equity floats.", "Equity = balance + open P&L.", "P0"),
    item("Orders — Account", "Commission deducted", "With commission set, check fill.", "Commission in trade details.", "P1"),
    item("Orders — Account", "Spread on fill", "Compare fill vs mid price.", "Spread applied realistically.", "P1"),
    item("Orders — Trades panel", "Open positions table", "Bottom panel → open trades.", "Lists side, size, entry, P&L.", "P0"),
    item("Orders — Trades panel", "Closed trades history", "Bottom panel → closed trades.", "Exit reason and P&L shown.", "P0"),
    item("Orders — Trade card", "Open trade card", "Click closed trade row.", "Card: entry, exit, R, duration, notes.", "P1"),
    item("Orders — Trade card", "PRE/POST screenshots on card", "Attach screenshots to trade card.", "Screenshots display on card.", "P2"),
    item("Orders — Trade card", "Trade notes / journal", "Add notes on trade card → save.", "Notes persist after reload.", "P1"),
    item("Orders — Prop firm", "Profit target progress", "Prop firm session → winning trades.", "Target progress bar updates.", "P1"),
    item("Orders — Prop firm", "Max drawdown rule", "Losing streak triggers drawdown limit.", "Warning or session fail per rules.", "P1"),
    item("Orders — Prop firm", "Daily loss cap", "Hit daily loss limit.", "Trading blocked or flagged.", "P1"),
    item("Orders — Prop firm", "Consistency rule", "If enabled, verify consistency tracking.", "Rule displays in HUD.", "P2"),
    item("Orders — Sound", "Execution sound", "Fill order with sound on.", "Audio plays.", "P1"),
    item("Orders — Replay", "Replay pauses on fill", "Enable pause-on-fill → place order.", "Replay pauses at fill (if configured).", "P2"),
]
add_phase(11, "Orders, Positions & Trade Journal", "Every order panel field, order type, advanced features, and trade tracking.", p11)

# ═══ PHASE 12 — Replay, Alerts, News, Screenshot, Final ═══
add_phase(12, "Replay, Alerts, News, Screenshot & Final QA", "Backtest replay workflow and remaining panels.", [
    item("Replay Bar", "Replay bar visible in session", "Load backtest session.", "Replay controls visible (play, speed, step).", "P0"),
    item("Replay Bar", "Play", "Click Play or Space.", "Chart advances through history.", "P0"),
    item("Replay Bar", "Pause", "Click Pause or Space.", "Advancement stops.", "P0"),
    item("Replay Bar", "Step forward", "Step forward one bar (Shift+→ or .).", "One bar advances.", "P0"),
    item("Replay Bar", "Step backward", "Step back one bar (Shift+← or ,).", "One bar back (if allowed).", "P1"),
    item("Replay Bar", "Speed slider", "Set 1×, 5×, 20× speed.", "Faster advance at higher speed.", "P0"),
    item("Replay Bar", "Tick playback mode", "Switch to Tick mode.", "Intra-bar animation between OHLC.", "P1"),
    item("Replay Bar", "Candle playback mode", "Switch to Candle mode.", "Jumps bar-to-bar.", "P0"),
    item("Replay Bar", "Step timeframe override", "Set step TF Auto / 1m / 5m.", "Step size changes accordingly.", "P2"),
    item("Replay Bar", "Pick-point rollback", "Click chart to set rollback line.", "Replay restarts from that bar.", "P0"),
    item("Replay Bar", "Rollback badge in header", "After rollback, check OHLC badge.", "Badge shows rollback active.", "P1"),
    item("Replay Bar", "Rollback blocked", "If session forbids rewind, try rollback.", "Blocked with clear message.", "P2"),
    item("Replay Bar", "Drag replay bar position", "Drag replay toolbar.", "Position moves; persists.", "P2"),
    item("Replay Bar", "Forward data prefetch", "Replay to edge of loaded data.", "New bars load without freeze.", "P1"),
    item("Alerts", "Create price alert", "Create alert at price (horizontal line).", "Alert line on chart.", "P1"),
    item("Alerts", "Condition — crossing", "Alert: crossing price.", "Triggers when crossed in replay.", "P1"),
    item("Alerts", "Condition — crossing up/down", "Alert: crossing up / crossing down.", "Directional trigger works.", "P2"),
    item("Alerts", "Condition — greater/less than", "Alert: price greater than X.", "Triggers correctly.", "P2"),
    item("Alerts", "Expiration — once / every time", "Test once vs every time.", "Behavior matches setting.", "P2"),
    item("Alerts", "Alert list — edit/delete", "Alerts panel → edit and delete.", "Changes apply.", "P1"),
    item("Alerts", "Drag alert line", "Drag alert to new price.", "Price updates.", "P1"),
    item("Alerts", "Sound on trigger", "Alert fires during replay.", "Sound plays; badge updates.", "P2"),
    item("News Panel", "News panel opens", "Right sidebar → News.", "Economic calendar loads.", "P1"),
    item("News Panel", "Impact filter H/M/L", "Toggle High/Medium/Low impact filters.", "Event list filters.", "P1"),
    item("News Panel", "Pair-only filter", "Show events for current pair only.", "Irrelevant events hidden.", "P2"),
    item("News Panel", "Country filter", "Filter by country.", "List updates.", "P2"),
    item("News Panel", "Axis markers on chart", "High-impact event → marker on time axis.", "Marker at event time.", "P1"),
    item("News Panel", "Click event scrolls chart", "Click event in list.", "Chart jumps to event time.", "P2"),
    item("Screenshot", "Open screenshot panel", "Camera icon or Alt+S.", "Screenshot panel opens.", "P1"),
    item("Screenshot", "Capture preview", "Capture chart.", "Preview image appears.", "P1"),
    item("Screenshot", "Download PNG", "Download PNG.", "File saves correctly.", "P1"),
    item("Screenshot", "Copy image to clipboard", "Copy image.", "Pastes in image editor.", "P2"),
    item("Screenshot", "Copy link", "Copy link if available.", "Link works.", "P2"),
    item("Screenshot", "Link to journal trade", "Link screenshot to trade in journal.", "Screenshot on trade card.", "P2"),
    item("Screenshot", "Logo watermark on capture", "Capture with watermark setting.", "Logo visible on PNG.", "P2"),
    item("Analytics", "Analytics tab loads stats", "Bottom → Analytics tab.", "Win rate, P&L stats, charts load.", "P1"),
    item("Analytics", "Analytics — by symbol", "Filter stats by symbol.", "Filtered results correct.", "P2"),
    item("Analytics", "Analytics — by strategy", "If strategies tagged, filter by strategy.", "Filtered results correct.", "P2"),
    item("Final — Stability", "30-minute stress test", "30 min: pan, zoom, draw, 5 orders, replay, 3 TF changes.", "No crash, freeze, or memory leak symptoms.", "P0"),
    item("Final — Stability", "Multichart + orders + replay", "4-panel + 2 orders + replay 10 min.", "All tiles stable.", "P0"),
    item("Final — Stability", "10 drawings + 5 indicators + reload", "Heavy chart → F5 reload.", "Everything restores.", "P0"),
    item("Final — Blockers log", "Review all ⚠ Partial and Fail rows", "Re-test any failed items; file BLOCKER list.", "Zero P0 failures before launch.", "P0"),
])


def style_header(ws, row, cols, color="1F4E79"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_list_validation(ws, col, r1, r2, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}{r1}:{col}{r2}")


def build():
    wb = Workbook()
    headers = [
        "Phase", "Phase Name", "Priority", "Area", "Test ID", "What to Test",
        "Steps (how to test)", "Expected Result", "Pre-condition",
        "Result", "Issue?", "Comments", "Tester", "Date",
    ]

    # Instructions
    ws_i = wb.active
    ws_i.title = "Instructions"
    ws_i["A1"] = "TALARIA CHART — COMPREHENSIVE UAT TRACKER"
    ws_i["A1"].font = Font(bold=True, size=14, color="1F4E79")
    lines = [
        ("", ""),
        ("Who", "Traders/backtesters who know TradingView. No coding needed."),
        ("Scope", "Every tool, indicator, panel, window, and setting in the chart app."),
        ("Phases", f"{len(PHASES)} phases — complete in order 1 → {len(PHASES)}."),
        ("Columns", "Result: ☑ Pass / ☐ Fail / — Not tested / ⚠ Partial. Issue?: ☑ Yes if any bug."),
        ("Priority", "P0 = launch blocker. P1 = important. P2 = edge/nice-to-have."),
        ("Comments", "Always write what you saw vs what you expected. Mark BLOCKER for show-stoppers."),
        ("Time", "~3–6 hours per phase. Full pass may take 40–60 hours with 2 testers."),
        ("Sheet", "Use QA Tracker tab; filter by Phase or Priority."),
        ("", "PHASE OVERVIEW"),
    ]
    for p in PHASES:
        lines.append((f"Phase {p['num']}: {p['name']}", p["goal"]))
    ws_i.column_dimensions["A"].width = 22
    ws_i.column_dimensions["B"].width = 95
    for r, (a, b) in enumerate(lines, 2):
        ws_i.cell(r, 1, a).font = Font(bold=True) if a else Font()
        ws_i.cell(r, 2, b).alignment = Alignment(wrap_text=True, vertical="top")

    # Master tracker
    ws = wb.create_sheet("QA Tracker")
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(headers))

    row = 2
    counters = {}
    for phase in PHASES:
        counters[phase["num"]] = 0
        for cat, name, steps, expected, pri, pre in phase["items"]:
            counters[phase["num"]] += 1
            tid = f"P{phase['num']}-{counters[phase['num']]:04d}"
            ws.cell(row, 1, phase["num"])
            ws.cell(row, 2, phase["name"])
            ws.cell(row, 3, pri)
            ws.cell(row, 4, cat)
            ws.cell(row, 5, tid)
            ws.cell(row, 6, name)
            ws.cell(row, 7, steps)
            ws.cell(row, 8, expected)
            ws.cell(row, 9, pre)
            ws.cell(row, 10, "— Not tested")
            ws.cell(row, 11, "— Not tested")
            for col in range(1, 13):
                ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
            row += 1

    last = row - 1
    add_list_validation(ws, "J", 2, last, STATUS_LIST)
    add_list_validation(ws, "K", 2, last, ISSUE_LIST)

    widths = [7, 26, 8, 22, 11, 32, 48, 36, 22, 14, 10, 36, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"

    # Summary
    ws_s = wb.create_sheet("Summary")
    sh = ["Phase", "Phase Name", "Tests", "P0 count", "Pass", "Fail", "Partial", "Issues", "Not tested", "Done %"]
    for c, h in enumerate(sh, 1):
        ws_s.cell(1, c, h)
    style_header(ws_s, 1, len(sh), "2E7D32")
    tr = "'QA Tracker'"
    for i, p in enumerate(PHASES, 2):
        n = len(p["items"])
        ws_s.cell(i, 1, p["num"])
        ws_s.cell(i, 2, p["name"])
        ws_s.cell(i, 3, n)
        ws_s.cell(i, 4, f'=COUNTIFS({tr}!A:A,{p["num"]},{tr}!C:C,"P0")')
        ws_s.cell(i, 5, f'=COUNTIFS({tr}!A:A,{p["num"]},{tr}!J:J,"☑ Pass")')
        ws_s.cell(i, 6, f'=COUNTIFS({tr}!A:A,{p["num"]},{tr}!J:J,"☐ Fail")')
        ws_s.cell(i, 7, f'=COUNTIFS({tr}!A:A,{p["num"]},{tr}!J:J,"⚠ Partial")')
        ws_s.cell(i, 8, f'=COUNTIFS({tr}!A:A,{p["num"]},{tr}!K:K,"☑ Yes")')
        ws_s.cell(i, 9, f'=COUNTIFS({tr}!A:A,{p["num"]},{tr}!J:J,"— Not tested")')
        ws_s.cell(i, 10, f'=IF(C{i}=0,0,ROUND((E{i}+G{i})/C{i}*100,0)&"%")')
    tot = len(PHASES) + 2
    ws_s.cell(tot, 2, "TOTAL").font = Font(bold=True)
    for col, formula in [(3, f"SUM(C2:C{len(PHASES)+1})"), (4, f"SUM(D2:D{len(PHASES)+1})"),
                         (5, f"SUM(E2:E{len(PHASES)+1})"), (6, f"SUM(F2:F{len(PHASES)+1})"),
                         (7, f"SUM(G2:G{len(PHASES)+1})"), (8, f"SUM(H2:H{len(PHASES)+1})"),
                         (9, f"SUM(I2:I{len(PHASES)+1})")]:
        ws_s.cell(tot, col, formula)

    # Checklist index
    ws_c = wb.create_sheet("Coverage Index")
    ws_c.cell(1, 1, "Area")
    ws_c.cell(1, 2, "Items in tracker")
    ws_c.cell(1, 3, "Phase(s)")
    style_header(ws_c, 1, 3, "5D4037")
    coverage = [
        ("Drawing line tools (×5 checks each)", len(LINE_TOOLS) * 5, "5"),
        ("Drawing shape tools (×5)", len(SHAPE_TOOLS) * 5, "6"),
        ("Drawing channel tools (×5)", len(CHANNEL_TOOLS) * 5, "6"),
        ("Drawing brush tools (×5)", len(BRUSH_TOOLS) * 5, "6"),
        ("Fib/Gann tools (×5)", len(FIB_GANN_TOOLS) * 5, "7"),
        ("Pattern/Elliott tools (×5)", len(PATTERN_TOOLS) * 5, "7"),
        ("Measure/position tools (×5)", len(MEASURE_TOOLS) * 5, "7"),
        ("Text/annotation tools (×5)", len(TEXT_TOOLS) * 5, "8"),
        ("Volume profile tools (×5)", len(VOLUME_TOOLS) * 5, "8"),
        ("Indicators (×5 checks each)", len(INDICATORS) * 5 + 8, "9"),
        ("Chart types", len(CHART_TYPES), "2"),
        ("Timeframes", len(TIMEFRAMES) + 4, "2"),
        ("Layout counts + sync toggles", len(LAYOUT_COUNTS) + len(SYNC_TOGGLES) + 8, "10"),
        ("Settings sections", 20, "3"),
        ("Order panel & trading", 55, "11"),
        ("Replay / alerts / news / screenshot", 40, "12"),
        ("UI panels & windows", 30, "1"),
    ]
    for r, (area, count, ph) in enumerate(coverage, 2):
        ws_c.cell(r, 1, area)
        ws_c.cell(r, 2, count)
        ws_c.cell(r, 3, ph)
    ws_c.column_dimensions["A"].width = 40
    ws_c.column_dimensions["B"].width = 18
    ws_c.column_dimensions["C"].width = 12

    # Per-phase sheets (cap name length)
    names = {
        1: "P1 Access Chrome", 2: "P2 Symbol TF Type", 3: "P3 Display Settings",
        4: "P4 Navigation", 5: "P5 Draw Lines", 6: "P6 Draw Shapes",
        7: "P7 Fib Patterns", 8: "P8 Text Edit Tree", 9: "P9 Indicators",
        10: "P10 Layouts", 11: "P11 Orders", 12: "P12 Replay Final",
    }
    phase_headers = headers[:12]
    for phase in PHASES:
        wsp = wb.create_sheet(names[phase["num"]])
        for c, h in enumerate(phase_headers, 1):
            wsp.cell(1, c, h)
        style_header(wsp, 1, len(phase_headers), "37474F")
        r = 2
        idx = 0
        for cat, name, steps, expected, pri, pre in phase["items"]:
            idx += 1
            wsp.cell(r, 1, phase["num"])
            wsp.cell(r, 2, phase["name"])
            wsp.cell(r, 3, pri)
            wsp.cell(r, 4, cat)
            wsp.cell(r, 5, f"P{phase['num']}-{idx:04d}")
            wsp.cell(r, 6, name)
            wsp.cell(r, 7, steps)
            wsp.cell(r, 8, expected)
            wsp.cell(r, 9, pre)
            wsp.cell(r, 10, "— Not tested")
            wsp.cell(r, 11, "— Not tested")
            r += 1
        add_list_validation(wsp, "J", 2, r - 1, STATUS_LIST)
        add_list_validation(wsp, "K", 2, r - 1, ISSUE_LIST)
        wsp.freeze_panes = "A2"
        for i, w in enumerate(widths[:12], 1):
            wsp.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT)
    total = sum(len(p["items"]) for p in PHASES)
    print(f"Created {OUTPUT}")
    print(f"Phases: {len(PHASES)} | Total tests: {total}")


if __name__ == "__main__":
    build()
