"""Build a simple two-status Excel sheet (Solved / On the way) for all tester tickets."""
import csv
import sys

CSV_PATH = "tickets/support-export-full-16-07-26/messages.csv"
OUT_XLSX = "docs/tickets-overhaul/TICKET-STATUS-SIMPLE.xlsx"
OUT_CSV = "docs/tickets-overhaul/TICKET-STATUS-SIMPLE.csv"

# Solved = fix implemented (staged/pending deploy) or closed outright.
SOLVED = {
    "TAL-01581", "TAL-01582",  # D-009 replay mode/cadence fixes staged
    "TAL-01585",               # T5 anchoring phases landed (RC-3 resolved)
    "TAL-01588",               # fixed directly by PO
    "TAL-01590",               # D-015 edge-park fix, PO-confirmed staging a4
    "TAL-01596",               # closed by tester
    "TAL-01600",               # D-015 + D-016 staged
    "TAL-01602",               # A6-1 apply-on-release landed
    "TAL-01609", "TAL-01610",  # D-015 staged
    "TAL-01611", "TAL-01612",  # D-009 staged
    "TAL-01616",               # A6-2 persistence landed
    "TAL-01626",               # refresh-persistence staged
    "TAL-01629", "TAL-01631",  # replay re-render family staged
    "TAL-01638",               # T4 order-type reclassify staged
    "TAL-01647",               # D-009 mode routing staged
    "TAL-01650",               # D-015/D-016 + refresh-persistence staged
    "TAL-01653",               # A6-1 covers it
}

rows = list(csv.DictReader(open(CSV_PATH, encoding="utf-8")))
tickets = {}
for r in rows:
    ref = r["ticket_ref"]
    if ref not in tickets:
        tickets[ref] = r["subject"]

refs = sorted(tickets)
print(f"tickets: {len(refs)}  first: {refs[0]}  last: {refs[-1]}")
missing_1617 = "TAL-01617" not in tickets
print("TAL-01617 present:", not missing_1617)

data = [(ref, "Solved" if ref in SOLVED else "On the way") for ref in refs]

with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["Ticket", "Status"])
    w.writerows(data)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("openpyxl not available; CSV written only")
    sys.exit(0)

wb = Workbook()
ws = wb.active
ws.title = "Ticket Status"
ws.append(["Ticket", "Status"])
green = PatternFill("solid", start_color="C6EFCE")
amber = PatternFill("solid", start_color="FFEB9C")
for ref, status in data:
    ws.append([ref, status])
    cell = ws.cell(row=ws.max_row, column=2)
    cell.fill = green if status == "Solved" else amber
for c in ("A1", "B1"):
    ws[c].font = Font(bold=True)
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 14
ws.auto_filter.ref = f"A1:B{ws.max_row}"
wb.save(OUT_XLSX)
solved = sum(1 for _, s in data if s == "Solved")
print(f"xlsx written: {OUT_XLSX}  Solved={solved}  On the way={len(data)-solved}")
