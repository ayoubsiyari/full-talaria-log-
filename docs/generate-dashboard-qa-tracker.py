#!/usr/bin/env py -3
"""Exhaustive Dashboard UAT tracker — Strategies, Sessions, Trades, Sources."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

OUTPUT = Path(__file__).resolve().parent / "Talaria-Dashboard-QA-Tracker.xlsx"
RESULT = '"☑ Pass,☐ Fail,⚠ Partial,— Not tested"'
ISSUE = '"☑ Yes,☐ No,— Not tested"'

STEPS = []  # dicts


def add(section, phase, phase_name, category, instruction, test_input, expected, priority="P1"):
    STEPS.append({
        "section": section,
        "phase": phase,
        "phase_name": phase_name,
        "category": category,
        "instruction": instruction,
        "test_input": test_input,
        "expected": expected,
        "priority": priority,
    })


def bulk(section, phase, phase_name, category, items, priority="P1"):
    for row in items:
        if len(row) == 2:
            instr, expected = row
            add(section, phase, phase_name, category, instr, "—", expected, priority)
        else:
            instr, tin, expected, *rest = row
            pri = rest[0] if rest else priority
            add(section, phase, phase_name, category, instr, tin, expected, pri)


S1 = "1 — Strategies"
S2 = "2 — Sessions"
S3 = "3 — Trades"
S4 = "4 — Sources"

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 1 — STRATEGIES
# ═══════════════════════════════════════════════════════════════════════════

bulk(S1, 1, "Strategy Bank — Access & Shell", "Structure", [
    ("Open /dashboard/ while logged in.", "—", "Dashboard shell loads; left nav visible.", "P0"),
    ("Navigate to Strategies via left nav.", "—", "Strategy Bank page; header title visible.", "P0"),
    ("Refresh browser on Strategy Bank.", "—", "Page reloads; strategies list persists.", "P1"),
    ("Resize browser: 1920px, 1366px, 1024px widths.", "3 widths", "Layout does not overlap or clip header/controls.", "P1"),
    ("Zoom browser 90% and 110%.", "—", "Cards/rows remain readable; no broken dropdowns.", "P2"),
])

bulk(S1, 2, "Strategy Bank — Controls & Empty States", "Normal", [
    ("Click Build Strategy.", "—", "Template picker OR builder opens.", "P0"),
    ("Toggle Cards view.", "—", "Grid of cards with consistent height.", "P1"),
    ("Toggle Rows view.", "—", "Table rows align; columns readable.", "P1"),
    ("Search partial strategy name.", "e.g. 'mom'", "Only matching strategies shown.", "P1"),
    ("Search by tag text.", "e.g. 'Scalping'", "Strategies with that tag appear.", "P1"),
    ("Search gibberish.", "zzzznotfound", "Empty/no-match message; no crash.", "P1"),
    ("Clear search with ×.", "—", "Full list restored.", "P2"),
])

bulk(S1, 2, "Strategy Bank — Controls & Empty States", "Edge", [
    ("Search with 1 character.", "a", "Filter applies or shows all if too short.", "P2"),
    ("Search with 100+ characters.", "long string", "Input capped or handled; no freeze.", "P2"),
    ("Search special characters.", "<script>, ', \"", "No script execution; safe display.", "P0"),
    ("Search Arabic/Unicode.", "استراتيجية", "Filter works or safely shows none.", "P2"),
    ("Rapidly toggle Cards ↔ Rows 10 times.", "—", "No duplicate cards or layout break.", "P2"),
])

bulk(S1, 3, "Strategy Bank — Card/Row Content & Visuals", "Structure", [
    ("Verify card shows: icon/emoji, name, description snippet.", "—", "All zones render.", "P1"),
    ("Verify markets, timeframes, tags pills.", "—", "Pills truncate gracefully if many.", "P1"),
    ("Verify backtest metrics: sessions, P&L, win rate.", "—", "Numbers or em-dash when N/A.", "P1"),
    ("Hover card.", "—", "Border/shadow highlight.", "P2"),
    ("Compare card at min vs max content (long name, 10 tags).", "2 strategies", "No overflow breaking card height (342px).", "P1"),
])

bulk(S1, 3, "Strategy Bank — Card/Row Content & Visuals", "Logic", [
    ("Double-click owned strategy.", "—", "Edit builder opens with data.", "P0"),
    ("Open ⋯ menu → New Session.", "—", "Session modal; strategy pre-linked.", "P0"),
    ("Open ⋯ → Dashboard.", "—", "Strategy performance view.", "P1"),
    ("Open ⋯ → Edit.", "—", "Builder edit mode.", "P0"),
    ("Open ⋯ → Copy.", "—", "Duplicate with (copy) suffix.", "P1"),
    ("Open ⋯ → Delete → Cancel.", "—", "Strategy remains.", "P1"),
    ("Open ⋯ → Delete → Confirm.", "—", "Strategy removed.", "P1"),
    ("Repeat all ⋯ actions in Rows view.", "—", "Parity with Cards.", "P1"),
])

bulk(S1, 4, "Template Picker", "Normal", [
    ("Build Strategy → pick each visible template one by one.", "all templates", "Each opens builder with distinct flow/groups.", "P1"),
    ("Cancel template picker.", "—", "Returns to bank without new strategy.", "P1"),
    ("Pick template → verify Step 2 canvas has groups/conditions.", "any template", "Pre-filled nodes visible.", "P0"),
])

bulk(S1, 4, "Template Picker", "Edge", [
    ("Pick template → change name → save → verify not confused with template preview.", "—", "Saved as user strategy in My Strategies.", "P1"),
    ("Re-open TEMPLATES from builder mid-edit.", "—", "Picker opens; confirm overwrite behavior if switching template.", "P2"),
])

# Builder shell
bulk(S1, 5, "Strategy Builder — Wizard Shell", "Structure", [
    ("Verify modal size ~1400px max width, ~90vh height.", "—", "Centered; scroll inside content.", "P2"),
    ("Verify 4 step tabs with hints.", "—", "General Info, Strategy Flow, Trade Tags, Review.", "P0"),
    ("Verify footer: Back, Next, Save Strategy (step 4).", "—", "Buttons visible per step.", "P0"),
    ("Click × close.", "—", "Modal closes.", "P1"),
    ("Click dark backdrop.", "—", "Modal stays open (backdrop ignored).", "P2"),
])

bulk(S1, 5, "Strategy Builder — Wizard Shell", "Logic", [
    ("Step 1 empty → click Next.", "—", "Blocked; red hints: name, markets, timeframes.", "P0"),
    ("Step 1 valid → jump to Step 4 tab.", "—", "Allowed when step 1 complete.", "P1"),
    ("Step 1 invalid → click Step 3 tab.", "—", "Forced back to Step 1.", "P0"),
    ("Back from Step 3 → Step 2 → data preserved.", "—", "No data loss.", "P0"),
])

# General Info — name
NAME_TESTS = [
    ("Leave name empty → Next.", "", "Blocked; red border on name field.", "P0"),
    ("Name = 1 character.", "A", "Accepts; can proceed if markets+TF set.", "P1"),
    ("Name = 80 characters (max).", "80×A", "Accepts; no truncate in input.", "P1"),
    ("Name = 81+ characters.", "81×A", "Input stops at 80 (maxLength).", "P1"),
    ("Name = only spaces.", "   ", "Treated as empty; blocked.", "P0"),
    ("Name with emoji inline.", "Rocket 🚀 Scalp", "Displays correctly on card.", "P2"),
    ("Name special chars.", "Test <>&\"'", "Displays literally; no HTML injection.", "P0"),
    ("Name Arabic.", "استراتيجية الذهب", "RTL displays correctly.", "P2"),
    ("Name duplicate of existing.", "same as another", "Allowed (duplicates OK) or warn — note behavior.", "P2"),
]
for t in NAME_TESTS:
    add(S1, 6, "Builder Step 1 — Strategy Name", "Input matrix", t[0], t[1], t[2], t[3])

# Description
bulk(S1, 6, "Builder Step 1 — Description & Emoji", "Input matrix", [
    ("Description empty.", "—", "Allowed; optional field.", "P1"),
    ("Description 500 chars (max).", "500 chars", "Counter shows 500/500; accepts.", "P1"),
    ("Description 501+ chars.", "paste 501", "Blocked at 500 maxLength.", "P1"),
    ("Description with newlines.", "line1\\nline2", "Multiline displays on card/review.", "P2"),
    ("Open emoji picker → browse Finance category.", "—", "Emoji grid loads.", "P1"),
    ("Search emoji picker.", "chart", "Filters emoji list.", "P2"),
    ("Select emoji → appears on button.", "📈", "Emoji shows on picker button.", "P1"),
    ("Remove emoji via × on button.", "—", "Emoji clears.", "P2"),
    ("Each emoji category tab.", "9 categories", "All categories have icons.", "P2"),
])

# Tags
bulk(S1, 7, "Builder Step 1 — Tags", "Input matrix", [
    ("Add 0 tags.", "—", "Allowed.", "P1"),
    ("Add 1 tag from library.", "Scalping", "Tag chip appears.", "P1"),
    ("Add 10 tags (MAX_TAGS).", "10 distinct", "Counter shows 10/10.", "P1"),
    ("Try 11th tag.", "11th tag", "Blocked; opacity/disabled in dropdown.", "P0"),
    ("Remove tag via chip ×.", "—", "Tag removed; can add again.", "P1"),
    ("Add custom tag via input.", "MyCustomTag", "Custom tag added.", "P1"),
    ("Toggle same tag off in dropdown.", "—", "Removed from selection.", "P2"),
    ("Duplicate tag click.", "same tag twice", "No duplicate chips.", "P2"),
])

# Markets & instruments
for mkt in ["Forex", "Futures", "Crypto", "Stocks"]:
    add(S1, 8, "Builder Step 1 — Markets & Instruments", "Configuration",
         f"Select only **{mkt}** market.", mkt,
         f"Market chip active; instrument picker filters to {mkt} symbols.", "P1")

bulk(S1, 8, "Builder Step 1 — Markets & Instruments", "Configuration", [
    ("Select all 4 markets.", "Forex+Futures+Crypto+Stocks", "Label shows 'All markets' or all 4 chips.", "P1"),
    ("Deselect all markets → Next.", "none", "Blocked; markets required.", "P0"),
    ("Add 1 instrument.", "e.g. EURUSD", "Instrument chip appears.", "P1"),
    ("Add 10 instruments (max).", "10 symbols", "10 chips; cannot add 11th.", "P0"),
    ("Instrument search: type 'EUR'.", "EUR", "Filters instrument list.", "P1"),
    ("Instrument search: no match.", "ZZZZZ", "Empty list; no crash.", "P2"),
    ("Supporting instruments: add 2 secondary symbols.", "2 symbols", "Supporting list populated.", "P2"),
    ("Toggle Style: Scalping, Intraday, Swing each.", "3 options", "One active at a time.", "P2"),
    ("Toggle Direction: Both, Long Only, Short Only.", "3 options", "Selection persists to review.", "P2"),
])

# Timeframes
for tf in ["1m", "5m", "15m", "30m", "1H", "4H", "1D", "1W", "1M"]:
    add(S1, 9, "Builder Step 1 — Timeframes", "Configuration",
         f"Select timeframe **{tf}**.", tf, f"{tf} chip active; included in saved strategy.", "P1")

bulk(S1, 9, "Builder Step 1 — Timeframes", "Input matrix", [
    ("Select 6 timeframes (max).", "6 TFs", "6 chips active.", "P1"),
    ("Try 7th timeframe.", "7th TF", "Blocked at 6 max.", "P0"),
    ("Custom TF: add 7m.", "7 + minutes", "7m appears in list and selection.", "P1"),
    ("Custom TF: add 3H.", "3 + hours", "3H works.", "P1"),
    ("Custom TF: invalid 0 minutes.", "0m", "Rejected; not added.", "P1"),
    ("Custom TF: duplicate 5m.", "5m again", "Not duplicated.", "P2"),
    ("Deselect all timeframes → Next.", "none", "Blocked; timeframes required.", "P0"),
])

# Images
bulk(S1, 10, "Builder Step 1 — Cover Images", "Edge", [
    ("Upload 1 valid PNG/JPG.", "small image", "Thumbnail appears.", "P1"),
    ("Upload 6 images (max).", "6 images", "6 thumbnails.", "P1"),
    ("Try 7th image.", "7th file", "Blocked at 6 max.", "P1"),
    ("Upload very large image (>5MB).", "large file", "Compresses or shows error alert; no hang.", "P1"),
    ("Upload non-image file renamed .png.", "fake file", "Error message; rejected.", "P1"),
    ("Remove image from gallery.", "—", "Image removed.", "P2"),
    ("Click image for preview if supported.", "—", "Preview modal opens.", "P2"),
])

# Strategy Flow canvas
bulk(S1, 11, "Builder Step 2 — Canvas Structure", "Structure", [
    ("Empty canvas initial state.", "—", "Grid/background; palette visible.", "P0"),
    ("Collapse palette panel.", "—", "More canvas space; re-expand works.", "P2"),
    ("Collapse inspector panel.", "—", "Toggle works.", "P2"),
    ("Collapse minimap.", "—", "Minimap hides/shows.", "P2"),
])

bulk(S1, 11, "Builder Step 2 — Groups & Conditions", "Logic", [
    ("Add group from palette.", "—", "Group node on canvas.", "P0"),
    ("Rename group.", "new name", "Label updates on canvas.", "P1"),
    ("Add condition inside group.", "—", "Condition node appears.", "P0"),
    ("Set condition Mandatory.", "—", "Green/mandatory styling.", "P1"),
    ("Set condition Optional.", "—", "Optional styling.", "P1"),
    ("Set condition Invalidate.", "—", "Red/invalidate styling.", "P1"),
    ("Edit condition title (max ~70 chars).", "70 chars", "Accepts; displays on node.", "P1"),
    ("Edit condition description long text.", "paragraph", "Inspector scrolls; saves.", "P2"),
    ("Add image to condition.", "image", "Image attaches to condition.", "P2"),
    ("Delete condition.", "Delete key", "Node removed.", "P1"),
    ("Delete group with children.", "—", "Group and children removed or confirm.", "P1"),
    ("Connect two nodes with edge if supported.", "—", "Edge line visible.", "P2"),
    ("Undo last canvas action if available.", "Ctrl+Z", "Reverts.", "P2"),
    ("Print/Export flow.", "toolbar", "Print window or PDF opens.", "P2"),
])

bulk(S1, 12, "Builder Step 3 — Trade Tags", "Configuration", [
    ("Add pre-trade tag.", "name + type", "Appears in pre-trade list.", "P0"),
    ("Add post-trade tag.", "—", "Post-trade list populated.", "P0"),
    ("Add divider between tags.", "—", "Visual divider in list.", "P2"),
    ("Variable type: checkbox tag.", "checkbox", "Renders as checkbox in trade form later.", "P1"),
    ("Variable type: number tag.", "number", "Numeric input in trade form.", "P1"),
    ("Variable type: text tag.", "text", "Text input in trade form.", "P1"),
    ("Delete tag.", "—", "Removed from list.", "P1"),
    ("Reorder tags if drag supported.", "—", "Order persists on save.", "P2"),
])

bulk(S1, 13, "Builder Step 4 — Review & Persistence", "Logic", [
    ("Review: General Info block shows all Step 1 data.", "—", "Matches inputs.", "P0"),
    ("Review: flow group/condition counts.", "—", "Match canvas.", "P0"),
    ("Review: readiness checklist red/green.", "—", "Incomplete items flagged.", "P1"),
    ("Save Strategy → appears in bank.", "—", "New card visible.", "P0"),
    ("Reload page → edit saved strategy.", "—", "All 4 steps restore.", "P0"),
    ("Copy strategy → edit copy independently.", "—", "Changes don't affect original.", "P1"),
    ("Save with slow network (throttle DevTools).", "—", "Saving indicator; error if fail; no duplicate.", "P2"),
])

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2 — SESSIONS
# ═══════════════════════════════════════════════════════════════════════════

bulk(S2, 14, "Sessions Page — Shell & Stats", "Structure", [
    ("Open Backtest via left nav.", "—", "Backtesting Sessions page.", "P0"),
    ("Verify stats tile: Sessions & Mode (Std/Prop/Journal).", "—", "Counts match list.", "P1"),
    ("Hover Total Trades chart bar.", "—", "Tooltip with session details.", "P1"),
    ("Profitable Sessions % arc.", "—", "Percentage displayed.", "P2"),
    ("Days Tested dots/bars.", "—", "Visual renders.", "P2"),
    ("Tickers Tested count.", "—", "Matches unique tickers.", "P2"),
])

FILTERS = ["All", "Not Started", "Active", "Completed", "Standard", "Prop Firm", "Journal"]
for f in FILTERS:
    add(S2, 15, "Sessions — Filters & Sort", "Configuration",
         f"Click filter tab **{f}**.", f,
         f"Only {f} sessions shown; badge count correct.", "P1")

SORT_COLS = ["Name", "Strategy", "Date Range", "Balance", "Net P&L", "Win Rate", "Avg R:R", "Trades", "Progress"]
for col in SORT_COLS:
    add(S2, 15, "Sessions — Filters & Sort", "Configuration",
         f"Sort cards by **{col}** ascending then descending.", col,
         "Order changes correctly both directions.", "P2")

bulk(S2, 15, "Sessions — Filters & Sort", "Edge", [
    ("Search session name partial.", "part of name", "Filters list.", "P1"),
    ("Search by ticker symbol.", "EURUSD", "Sessions containing ticker shown.", "P1"),
    ("Search by strategy name.", "strategy", "Filters correctly.", "P1"),
    ("Escape clears search.", "Esc", "Search cleared.", "P2"),
    ("Combine filter Active + search.", "both", "Intersection of filters.", "P2"),
])

bulk(S2, 16, "Sessions — Row Actions by State", "Logic", [
    ("Not Started session: Resume ▶.", "progress=0", "Opens chart at start.", "P0"),
    ("Active session: Resume ▶.", "0<progress<100", "Continues from saved bar.", "P0"),
    ("Completed session: Resume ▶.", "progress=100", "Opens at end or review mode — note behavior.", "P1"),
    ("⋯ → Edit on Not Started.", "—", "Session modal edit mode.", "P0"),
    ("⋯ → Edit on Active (started).", "—", "Edit DISABLED.", "P0"),
    ("⋯ → Duplicate.", "—", "Copy created.", "P1"),
    ("⋯ → Delete → cancel.", "—", "Session kept.", "P1"),
    ("⋯ → Delete → confirm.", "—", "Session removed.", "P1"),
    ("Journal session ⋯ → Add Trade.", "—", "Opens add trade for journal.", "P1"),
    ("Journal ⋯ → Edit Journal (no live account).", "—", "Disabled if no account.", "P2"),
])

bulk(S2, 17, "New Session — Kind Picker", "Configuration", [
    ("New Session → Backtest.", "—", "Backtest modal opens.", "P0"),
    ("New Session → Journal → Real account.", "—", "Journal flow continues.", "P1"),
    ("New Session → Journal → Prop firm.", "—", "Prop journal path.", "P1"),
    ("Manual trades option.", "—", "Selectable.", "P1"),
    ("CSV import — Coming Soon.", "—", "Disabled or blocked.", "P2"),
    ("Link broker — Coming Soon.", "—", "Disabled or blocked.", "P2"),
])

# Session modal — name & strategy
bulk(S2, 18, "Add Session — Session Info Fields", "Input matrix", [
    ("Session name empty.", "", "Save/Start disabled.", "P0"),
    ("Session name 1 char.", "A", "Accepted.", "P2"),
    ("Session name very long (200+ chars).", "long", "Truncated or accepted — note limit.", "P2"),
    ("Session name special chars.", "Test & < >", "Displays safely.", "P1"),
    ("Strategy = None.", "None", "Session saves without linked strategy.", "P1"),
    ("Strategy = pick each of your strategies.", "each", "Strategy name links.", "P1"),
    ("New Strategy shortcut.", "—", "Navigates to strategy create.", "P2"),
    ("Description empty vs filled.", "paragraph", "Saves both cases.", "P2"),
])

# Asset classes & tickers
for ac in ["Forex", "Futures", "Stocks", "Crypto"]:
    add(S2, 19, "Add Session — Asset Class & Symbols", "Configuration",
         f"Set asset class **{ac}**; open symbol picker.", ac,
         f"Picker shows {ac} symbols; can add to session.", "P0")

bulk(S2, 19, "Add Session — Asset Class & Symbols", "Edge", [
    ("Add 1 primary ticker.", "EURUSD", "Chip appears; required met.", "P0"),
    ("Add tickers up to plan max (default 5).", "5 symbols", "Counter shows 5/5.", "P0"),
    ("Try 6th ticker on standard plan.", "6th", "Blocked or dimmed at cap.", "P0"),
    ("Remove ticker chip.", "× on chip", "Ticker removed.", "P1"),
    ("Random ticker picker — set count 3 → apply.", "random 3", "3 random symbols added.", "P2"),
    ("Symbol search in picker.", "NQ", "Filters list.", "P1"),
    ("Enable supporting tickers; add 2.", "2 support", "Supporting section populated.", "P2"),
    ("Supporting tickers at max cap.", "max support", "Cannot exceed plan supporting limit.", "P1"),
])

# Timeframes & replay
for tf in ["1m", "5m", "15m", "30m", "1H", "4H", "1D", "1W", "1M"]:
    add(S2, 20, "Add Session — Timeframe & Replay", "Configuration", f"Set timeframe {tf}.", tf, "Saves with session.", "P1")

bulk(S2, 20, "Add Session — Timeframe & Replay", "Configuration", [
    ("Replay mode: Candle.", "candle", "Saves.", "P1"),
    ("Replay mode: Tick.", "tick", "Saves.", "P1"),
    ("Replay speed minimum.", "1×", "Slider at low end.", "P2"),
    ("Replay speed maximum.", "60×+", "High speed accepted.", "P2"),
    ("Rollback ON.", "—", "Session allows rewind in chart.", "P1"),
    ("Rollback OFF.", "—", "Rewind blocked in chart.", "P1"),
    ("MFE/MAE tracking ON + 4 hours.", "4h", "Saves.", "P2"),
    ("MFE/MAE tracking OFF.", "—", "Fields disabled.", "P2"),
    ("Post-exit window: hours mode, 4 hours.", "4", "Saves.", "P2"),
    ("Post-exit window: candles mode, 50 candles.", "50", "Saves.", "P2"),
    ("Advanced order ON.", "—", "Chart order panel shows advanced.", "P1"),
    ("Advanced order OFF.", "—", "Basic orders only.", "P1"),
])

# Date range matrix
DATE_TESTS = [
    ("Valid range: 1 year.", "2023-01-01 to 2023-12-31", "Accepts; calendar highlights range.", "P0"),
    ("Same start and end date.", "one day", "Single-day session allowed or blocked — note.", "P1"),
    ("End before start.", "inverted", "Error or auto-swap.", "P0"),
    ("Range 10+ years.", "long range", "Accepts or warns on data availability.", "P2"),
    ("Start date in far past.", "1990", "Clamped to available data or error.", "P2"),
    ("End date in future.", "future", "Blocked or clamped to today.", "P1"),
    ("Manual type invalid date.", "2023-02-30", "Rejected.", "P1"),
    ("Quick preset: YTD.", "preset", "Range auto-fills.", "P1"),
    ("Quick preset: 1Y / 6M / 3M.", "each preset", "Each fills correct range.", "P1"),
    ("N-bars mode: 500 bars.", "500", "Uses last N bars instead of calendar.", "P2"),
    ("N-bars mode: 1 bar.", "1", "Extreme minimum.", "P2"),
    ("N-bars mode: 50000 bars.", "50000", "Extreme maximum — note performance.", "P2"),
]
for t in DATE_TESTS:
    add(S2, 21, "Add Session — Date Range", "Input matrix", t[0], t[1], t[2], t[3])

# Account settings matrix
CAPITAL_TESTS = [
    ("Capital = 100 (minimum realistic).", "100", "Accepts.", "P1"),
    ("Capital = 10000 (default).", "10000", "Accepts.", "P0"),
    ("Capital = 1000000.", "1000000", "Accepts; margin calc scales.", "P2"),
    ("Capital = 0.", "0", "Blocked or warning.", "P0"),
    ("Capital = negative.", "-1000", "Blocked.", "P0"),
    ("Capital = text abc.", "abc", "Rejected or ignored.", "P1"),
    ("Currency USD.", "USD", "Flag/icon USD.", "P1"),
    ("Currency EUR.", "EUR", "Changes currency symbol.", "P2"),
    ("Risk mode % ; value 1%.", "1%", "Risk per trade 1% of balance.", "P1"),
    ("Risk mode % ; value 100%.", "100%", "Extreme risk — accepts or caps.", "P2"),
    ("Risk mode fixed $ ; value 500.", "$500", "Fixed dollar risk.", "P1"),
    ("Leverage 1:1.", "1:1", "No leverage.", "P1"),
    ("Leverage 1:500.", "1:500", "Max leverage.", "P2"),
    ("Commission none.", "none", "Zero commission.", "P1"),
    ("Commission per lot $3.50.", "3.50", "Applied on trades.", "P1"),
    ("Slippage 0.", "0", "No slippage.", "P2"),
    ("Slippage 10 ticks.", "10", "Wide slippage.", "P2"),
]
for t in CAPITAL_TESTS:
    add(S2, 22, "Add Session — Account & Costs", "Input matrix", t[0], t[1], t[2], t[3])

bulk(S2, 22, "Add Session — Account & Costs", "Configuration", [
    ("Enable Trading Costs overrides.", "—", "Per-asset-class commission/leverage fields appear.", "P1"),
    ("Set Forex costs commission 7.00 leverage 1:500.", "—", "Overrides default.", "P2"),
    ("Set per-symbol spread for EURUSD.", "1.5 pips", "Spread applied in session.", "P2"),
    ("Timezone America/New_York + DST on.", "—", "Chart times in NY.", "P1"),
    ("Timezone UTC + DST off.", "—", "Times in UTC.", "P2"),
    ("Margin call 100% / Stop out 50%.", "defaults", "Liquidation rules apply in chart.", "P2"),
    ("Link data file from picker.", "CSV file", "File attached to session.", "P2"),
])

# Prop firm exhaustive
bulk(S2, 23, "Add Session — Prop Firm Mode", "Configuration", [
    ("Switch to Prop Firm mode.", "—", "Prop section expands; Standard fields remain.", "P0"),
    ("Prop on Futures asset class.", "Futures", "Prop tab may be disabled — verify.", "P1"),
    ("Firm preset FTMO.", "FTMO", "Default rules populate.", "P1"),
    ("Challenge type Evaluation.", "—", "Phase fields shown.", "P1"),
    ("Num phases = 1.", "1", "Only Phase 1 fields.", "P1"),
    ("Num phases = 2.", "2", "Phase 2 fields appear.", "P1"),
    ("Phase 1 daily loss 5%.", "5", "Saves.", "P1"),
    ("Phase 1 max DD 10%.", "10", "Saves.", "P1"),
    ("Phase 1 profit target 10%.", "10", "Saves.", "P1"),
    ("Phase 1 min days ON, value 4.", "4", "Min trading days rule.", "P1"),
    ("Phase 1 min days OFF.", "disabled", "Field grayed.", "P2"),
    ("Phase 1 amounts $ mode: daily loss $1000.", "$1000", "Dollar-based rules.", "P2"),
    ("Trailing drawdown ON.", "—", "Trailing DD active in chart.", "P1"),
    ("Consistency rule ON, 30%.", "30%", "Consistency enforced.", "P2"),
    ("Weekend holding ON.", "—", "Allows weekend positions.", "P2"),
    ("Max contracts ON, value 5.", "5", "Caps position size.", "P2"),
    ("Max lots ON with unit lots.", "lots", "Forex lot cap.", "P2"),
])

bulk(S2, 24, "Add Session — Save, Start, Limits", "Logic", [
    ("Fill required fields → Save (no launch).", "—", "Session in list as Not Started.", "P0"),
    ("Fill required → Start Session.", "—", "Redirects to chart; session Active.", "P0"),
    ("Missing name → buttons disabled.", "—", "Cannot submit.", "P0"),
    ("Missing ticker → disabled.", "—", "Cannot submit.", "P0"),
    ("Missing dates → disabled.", "—", "Cannot submit.", "P0"),
    ("Missing capital → disabled.", "—", "Cannot submit.", "P0"),
    ("At session plan cap → create another.", "—", "Session limit modal.", "P1"),
    ("Edit not-started session: change name, Save.", "—", "Updates in list.", "P0"),
    ("Edit not-started: change all fields, Save.", "—", "All persist on reopen.", "P1"),
])

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3 — TRADES
# ═══════════════════════════════════════════════════════════════════════════

bulk(S3, 25, "Trades Page — Shell & Header", "Structure", [
    ("Open Trades via left nav.", "—", "Trade table visible.", "P0"),
    ("Source switcher shows active source label.", "—", "Matches applied source.", "P0"),
    ("With no source applied, note empty state.", "—", "Prompt to select source or empty table.", "P1"),
])

bulk(S3, 26, "Trades — Filters, Import, Export, Compare", "Configuration", [
    ("Open Filters panel.", "—", "Date, scope, tags, outcome, privacy options.", "P0"),
    ("Filter date range: last 30 days.", "30d", "Table filters.", "P1"),
    ("Filter date range: custom 1 day.", "1 day", "Narrow results.", "P2"),
    ("Filter by outcome: winners only.", "winners", "Only positive P&L rows.", "P1"),
    ("Filter by outcome: losers only.", "losers", "Only negative P&L.", "P1"),
    ("Combine 3 filters simultaneously.", "multi", "Intersection correct.", "P2"),
    ("Clear all filters.", "—", "Full list returns.", "P1"),
    ("Export CSV with trades.", "—", "File downloads with correct columns.", "P0"),
    ("Export with zero trades.", "—", "Disabled or empty file message.", "P2"),
    ("Import valid CSV.", "valid template", "Trades added; count updates.", "P1"),
    ("Import malformed CSV.", "bad file", "Error message; no partial corrupt data.", "P1"),
    ("Import while trades exist.", "—", "Merge or duplicate behavior documented.", "P2"),
    ("Enable Compare mode.", "—", "Compare UI active.", "P1"),
    ("Clear compare.", "—", "Returns to single source.", "P1"),
])

bulk(S3, 27, "Trades Table — Columns, Sort, Expand", "Structure", [
    ("Verify default columns present.", "—", "ID, Source, Date, Symbol, Side, Qty, Status, P&L, R.", "P0"),
    ("Column picker: enable each optional column.", "all optional", "Each appears in table.", "P1"),
    ("Column picker: max 12 on overview.", "13th column", "Blocked or warning at 12.", "P2"),
    ("Save custom column view name.", "My View", "View restorable.", "P2"),
    ("Sort by Date asc/desc.", "—", "Chronological order.", "P1"),
    ("Sort by Net P&L.", "—", "Highest/lowest P&L.", "P1"),
    ("Sort by R.", "—", "R-multiple order.", "P1"),
    ("Expand row.", "—", "Detail: prices, excursion chart, tags, notes.", "P0"),
    ("Expand row with screenshots.", "—", "Images display.", "P1"),
    ("Edit from expanded row.", "—", "Editor opens.", "P0"),
])

bulk(S3, 28, "Add Trade — Source Picker", "Logic", [
    ("Add Trade → pick backtest session.", "session", "Editor scoped to session instruments.", "P0"),
    ("Add Trade → pick journal account.", "journal", "Discipline tab visible.", "P0"),
    ("Single eligible source auto-skip picker.", "1 source", "Editor opens directly.", "P2"),
    ("First-time warning modal → dismiss.", "—", "Editor opens.", "P2"),
    ("Warning 'Don't show again' → reload → Add Trade.", "—", "Modal skipped.", "P2"),
])

# Trade scenarios — long/short normal
bulk(S3, 29, "Add Trade — Trade & Risk (Normal Scenarios)", "Normal", [
    ("LONG win: entry 100, SL 95, TP 110, exit 110, qty 1.", "see left", "Positive P&L; R ≈ 2.", "P0"),
    ("LONG loss: exit at SL 95.", "—", "Negative P&L; R ≈ -1.", "P0"),
    ("SHORT win: entry 100, SL 105, TP 90, exit 90.", "—", "Positive P&L on short.", "P0"),
    ("SHORT loss: exit 105.", "—", "Negative P&L.", "P0"),
    ("OPEN trade: no exit rows, exit timing OFF.", "—", "Status Open; no close P&L.", "P0"),
    ("2 entry rows scale-in: 50% + 50%.", "2 entries", "Avg entry weighted correctly.", "P1"),
    ("5 entry rows.", "5 entries", "Avg entry still calculates.", "P2"),
    ("2 targets partial.", "2 TPs", "Planned R uses blended target.", "P1"),
    ("2 exit rows partial close.", "50% + 50%", "Exited size matches.", "P1"),
    ("Commission 5.00 on round trip.", "$5", "Net P&L reduced by commission.", "P1"),
    ("Spread 2 pips.", "2", "Affects fill prices.", "P2"),
])

# Field input extremes
TRADE_FIELD_TESTS = [
    ("Entry price 0.00001 (micro).", "micro price", "Accepts if >0.", "P2"),
    ("Entry price 999999.", "huge price", "Accepts; preview scales.", "P2"),
    ("Qty 0.01 lots (min).", "0.01", "Accepts.", "P1"),
    ("Qty 100 lots.", "100", "Accepts; risk preview updates.", "P2"),
    ("Qty 0.", "0", "Blocked.", "P0"),
    ("Qty negative.", "-1", "Blocked.", "P0"),
    ("Stop loss = 0.", "0", "Blocked.", "P0"),
    ("Stop loss negative.", "-5", "Blocked.", "P0"),
]
for t in TRADE_FIELD_TESTS:
    add(S3, 30, "Add Trade — Field Input Extremes", "Extreme", t[0], t[1], t[2], t[3])

# Validation matrix from validateDashAddTradeDraft
VALIDATION_TESTS = [
    ("Symbol not in source instrument list.", "GBPUSD on NQ-only session", "Error: instrument must belong to source.", "P0"),
    ("Journal backtest trade without strategy/setup.", "empty setup", "Error: choose strategy.", "P0"),
    ("Missing entry date/time.", "empty", "Error: entry date and time required.", "P0"),
    ("Entry date before session start.", "before range", "Error: within session range.", "P0"),
    ("Entry date after session end.", "after range", "Error: within session range.", "P0"),
    ("Live journal entry date tomorrow.", "future date", "Error: cannot be in future.", "P0"),
    ("Backtest entry datetime in future.", "future datetime", "Error: cannot be in future.", "P0"),
    ("Entry row price=0.", "0", "Error: price and size > 0.", "P0"),
    ("Entry row qty=0.", "0", "Error: price and size > 0.", "P0"),
    ("Missing stop loss.", "empty SL", "Error: SL required.", "P0"),
    ("LONG SL above entry.", "SL>entry", "Error: SL below entries.", "P0"),
    ("SHORT SL below entry.", "SL<entry", "Error: SL above entries.", "P0"),
    ("LONG target below entry.", "TP<entry", "Error: targets above entries.", "P0"),
    ("SHORT target above entry.", "TP>entry", "Error: targets below entries.", "P0"),
    ("Exit size > entry size.", "exit>entry", "Error: exited size exceeds entered.", "P0"),
    ("Exit rows without exit timing ON.", "exits only", "Error: enable exit time or clear exits.", "P1"),
    ("Exit timing ON but no exit rows.", "timing only", "Error: add exit or disable timing.", "P1"),
    ("Exit before entry time.", "exit<entry", "Error: exit after entry.", "P0"),
    ("Exit date outside session range.", "outside", "Error: within session range.", "P1"),
    ("Target price 0.", "TP=0", "Error: targets > 0.", "P1"),
    ("Negative commission.", "-1", "Error: commission >= 0.", "P1"),
    ("Negative spread.", "-1", "Error: spread >= 0.", "P1"),
]
for t in VALIDATION_TESTS:
    add(S3, 31, "Add Trade — Validation Errors", "Logic", t[0], t[1], t[2], t[3])

# Excursion tests
EXCURSION_TESTS = [
    ("Excursion mode None.", "none", "No high/low required; saves.", "P1"),
    ("Standard: high=105 low=95 entry=100.", "in range", "Saves; MFE/MAE calculate.", "P1"),
    ("Standard: high <= low.", "high=95 low=100", "Error: high above low.", "P0"),
    ("Standard: entry outside high-low.", "entry=110", "Error: entry inside range.", "P0"),
    ("Standard: exit outside range.", "exit outside", "Error: exit inside range.", "P1"),
    ("Standard: range >500R from entry.", "extreme range", "Error: too far to analyze.", "P1"),
    ("Extended: closed trade + post high/low + window 10.", "valid", "Saves.", "P2"),
    ("Extended on open trade.", "open", "Error: needs closed trade.", "P1"),
    ("Extended: post window 0.", "0", "Error: whole number > 0.", "P1"),
    ("Extended: post window 1.5 (non-integer).", "1.5", "Error: whole number.", "P2"),
]
for t in EXCURSION_TESTS:
    add(S3, 32, "Add Trade — Excursion MFE/MAE", "Logic", t[0], t[1], t[2], t[3])

# Warnings
bulk(S3, 32, "Add Trade — Warnings (non-blocking)", "Edge", [
    ("Very wide stop → R > 50.", "tiny stop distance", "Warning: unusually large R; can still save.", "P2"),
    ("No target set.", "empty TP", "Warning: planned R blank.", "P2"),
    ("No commission entered.", "empty", "Warning: net equals gross.", "P2"),
])

bulk(S3, 33, "Add Trade — Tags, Discipline, Notes", "Configuration", [
    ("Tags tab: select 3 pre-trade tags from strategy.", "3 tags", "Chips active.", "P1"),
    ("Tags tab: select 5 post-trade tags.", "5 tags", "All save.", "P1"),
    ("Strategy/setup required — select from dropdown.", "strategy", "Saves.", "P0"),
    ("Discipline tab (journal): plan review field.", "text", "Saves.", "P2"),
    ("Discipline: behavior tags multi-select.", "tags", "Saves.", "P2"),
    ("Notes: pre-trade 500 chars.", "500 chars", "Saves.", "P2"),
    ("Notes: post-trade empty.", "—", "Allowed.", "P2"),
    ("Screenshot pre: upload PNG.", "image", "Thumbnail shows.", "P1"),
    ("Screenshot post: upload 2 images if slots allow.", "2 images", "Both display.", "P2"),
    ("Save & Add Another.", "—", "Trade saved; form resets.", "P1"),
])

bulk(S3, 34, "Add Trade — Persistence & Edit", "Logic", [
    ("Save → verify row in table.", "—", "New row with correct P&L.", "P0"),
    ("Edit trade → change SL → Save.", "—", "Row updates; R recalculates.", "P0"),
    ("Edit trade → change tags → Save.", "—", "Expanded row shows new tags.", "P1"),
    ("Delete trade if available.", "—", "Removed from table.", "P2"),
])

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 4 — SOURCES
# ═══════════════════════════════════════════════════════════════════════════

bulk(S4, 35, "Sources Library — Open & Shell", "Structure", [
    ("Dashboard → click source chip.", "—", "Library overlay opens.", "P0"),
    ("Trades → source switcher.", "—", "Same library UI.", "P0"),
    ("Verify overlay: tree left, list center, footer.", "—", "Three-zone layout.", "P1"),
    ("Click outside overlay.", "—", "Stays open or closes — note behavior.", "P2"),
    ("Resize window with library open.", "—", "Layout adapts.", "P2"),
])

TREE_NODES = [
    ("Expand Backtest Sessions.", "—", "Standard and Prop Firm children.", "P0"),
    ("Expand Standard under Backtests.", "—", "Standard sessions listed.", "P0"),
    ("Expand Prop Firm under Backtests.", "—", "Prop sessions listed.", "P0"),
    ("Expand Live Journals.", "—", "Personal and Prop children.", "P0"),
    ("Expand Strategies tree node.", "—", "Strategy groups with children.", "P1"),
    ("Collapse all then re-expand.", "—", "State consistent.", "P2"),
]
for t in TREE_NODES:
    add(S4, 36, "Sources — Tree Navigation", "Structure", t[0], t[1], t[2], t[3])

# Filter matrix
for mode in ["Standard", "Prop"]:
    for status in ["All", "Not Started", "Active", "Completed"]:
        add(S4, 37, "Sources — Filter Matrix (Backtests)", "Configuration",
             f"Mode **{mode}** + Status **{status}**.", f"{mode}+{status}",
             "List shows only matching sessions; count in footer updates.", "P2")

bulk(S4, 37, "Sources — Filter Matrix (Backtests)", "Configuration", [
    ("Journal filter: Personal only.", "—", "Personal journal accounts.", "P1"),
    ("Journal filter: Prop only.", "—", "Prop journal accounts.", "P1"),
    ("Strategy view filter if present.", "—", "Filters strategy groups.", "P2"),
])

bulk(S4, 38, "Sources — Selection & Footer", "Logic", [
    ("Single-click session row.", "—", "Row highlighted; footer updates.", "P0"),
    ("Footer: name, type badge, trade count.", "—", "Matches selection.", "P0"),
    ("Select strategy parent.", "—", "All children selected or prompt.", "P1"),
    ("Uncheck 1 of 3 strategy children.", "partial", "Partial selection; N/M label.", "P1"),
    ("Select journal account.", "—", "Green journal badge.", "P0"),
    ("Select single journal entry (trade scope).", "—", "Trade count = 1.", "P2"),
    ("Cancel after changing selection.", "—", "Previous applied source unchanged.", "P0"),
    ("Go with new selection.", "—", "Dashboard reloads with new source.", "P0"),
])

bulk(S4, 39, "Sources — Go Behavior & Banner", "Logic", [
    ("Go: backtest session → dashboard metrics.", "session", "Metrics match session trades.", "P0"),
    ("Go: strategy all children.", "strategy full", "Aggregated P&L.", "P1"),
    ("Go: strategy 2 of 5 children.", "partial", "Metrics scope to 2 only.", "P1"),
    ("Go: journal account.", "journal", "All journal trades in scope.", "P0"),
    ("Banner after Go: correct type badge color.", "—", "Backtest=blue, Journal=green.", "P1"),
    ("Banner trade count matches table.", "—", "Counts equal.", "P0"),
    ("Empty journal: Create Journal CTA.", "—", "Button works.", "P2"),
    ("Empty journal: Add Trade CTA.", "—", "Opens add trade.", "P2"),
])

bulk(S4, 40, "Sources — Header Actions & Compare", "Configuration", [
    ("Library header Create Session.", "—", "New session flow.", "P1"),
    ("Library header Create Journal.", "—", "Journal modal.", "P1"),
    ("Library header Add Trade.", "—", "Add trade picker.", "P1"),
    ("Compare mode: select 2 backtest sessions.", "2 sessions", "Compare chip shows both.", "P1"),
    ("Compare: 3+ sources if multi-select.", "3+", "All active.", "P2"),
    ("Compare filters separate from main.", "—", "Compare scope independent.", "P2"),
    ("Clear compare trash icon.", "—", "Exits compare; single source.", "P1"),
    ("Dashboard metrics under compare.", "—", "Side-by-side or combined view correct.", "P1"),
])

# Cross-section integration
bulk(S4, 41, "Cross-Section Integration", "Logic", [
    ("Strategy ⋯ New Session → complete modal → Start.", "—", "Chart opens with strategy linked.", "P0"),
    ("Session Resume → place trade in chart → see in Trades.", "—", "Trade appears in table.", "P0"),
    ("Sources Go → Trades tab: same trade count.", "—", "Consistent scope.", "P0"),
    ("Add Trade tags from strategy builder → visible on trade row.", "—", "Tag presets match.", "P1"),
    ("Full loop: Build strategy → New session → Resume → Add manual trade → Sources verify.", "end-to-end", "All sections consistent.", "P0"),
])

# ─── Additional exhaustive matrices (appended) ─────────────────────────────

# Session card visuals per mode
for mode_label, mode_hint in [
    ("Standard backtest", "blue/std styling, date range bar, rollback/costs chips"),
    ("Prop Firm backtest", "gold styling, prop badge, firm name if shown"),
    ("Journal session", "green styling, 'Live · since' or manual label"),
]:
    add(S2, 42, "Sessions — Visual Structure per Mode", "Structure",
         f"Open **{mode_label}** session card in Cards view.", mode_label,
         f"Card shows correct {mode_hint}; progress bar 0–100%.", "P1")

bulk(S2, 42, "Sessions — Visual Structure per Mode", "Structure", [
    ("Rows view: verify column headers align with data.", "—", "No column shift on horizontal scroll.", "P1"),
    ("Rows view: horizontal scroll with 20+ sessions.", "—", "Sticky filter row stays visible.", "P2"),
    ("Progress bar 0% (not started).", "—", "Empty bar.", "P1"),
    ("Progress bar 50% (active).", "—", "Half fill.", "P1"),
    ("Progress bar 100% (completed).", "—", "Full fill.", "P1"),
    ("Negative P&L session: red styling.", "losing session", "P&L shown in red.", "P1"),
    ("Positive P&L session: green styling.", "winning session", "P&L shown in green.", "P1"),
])

# Session modal cross-product: trading mode × asset class
for tm in ["standard", "prop"]:
    for ac in ["Forex", "Futures", "Stocks", "Crypto"]:
        add(S2, 43, "Add Session — Mode × Asset Class Matrix", "Configuration",
             f"Create session: trading mode **{tm}**, asset class **{ac}**.", f"{tm}+{ac}",
             "Modal saves; session card shows correct mode+asset; chart loads symbol from class.", "P2")

# Strategy builder condition status matrix
for status in ["mandatory", "optional", "invalidate"]:
    add(S1, 44, "Builder — Condition Status Visuals", "Structure",
         f"Add condition with status **{status}**.", status,
         f"Node/badge color matches {status} (green/optional/red).", "P1")

bulk(S1, 44, "Builder — Condition Status Visuals", "Structure", [
    ("Canvas with 0 groups — Review step.", "empty flow", "Readiness shows flow incomplete.", "P1"),
    ("Canvas with 5 groups × 3 conditions each.", "15 conditions", "Performance OK; scroll works; save succeeds.", "P2"),
    ("Condition with very long description (2000 chars).", "long text", "Inspector scrolls; save succeeds or truncates safely.", "P2"),
    ("Group name with special characters.", "G&1 <test>", "Displays safely on canvas.", "P2"),
])

# Trades table visual edge cases
bulk(S3, 45, "Trades Table — Visuals & Edge Data", "Structure", [
    ("Table with 0 trades.", "empty source", "Empty state message; Export disabled.", "P1"),
    ("Table with 1 trade.", "1 row", "Row renders; expand works.", "P1"),
    ("Table with 100+ trades.", "large set", "Scroll performance OK; sort works.", "P2"),
    ("Very long symbol name in row.", "long symbol", "Truncates with ellipsis.", "P2"),
    ("Trade with edited badge.", "edited trade", "Badge visible.", "P2"),
    ("Discipline: According to Plan label.", "—", "Green/neutral styling.", "P2"),
    ("Discipline: Out of Plan label.", "—", "Warning styling.", "P2"),
    ("Excursion chart in expanded row.", "trade with excursion", "Mini chart renders.", "P1"),
    ("Row with 4 pre screenshots.", "4 images", "All thumbnails visible.", "P2"),
])

# Add trade — datetime edge matrix
for label, tin, exp, pri in [
    ("Entry midnight 00:00.", "00:00", "Accepts; saves correctly.", "P2"),
    ("Entry 23:59.", "23:59", "Accepts.", "P2"),
    ("Entry date = session start date (boundary).", "start date", "Accepts.", "P1"),
    ("Entry date = session end date (boundary).", "end date", "Accepts.", "P1"),
    ("Exit 1 minute after entry.", "+1 min", "Accepts.", "P1"),
    ("Exit same day, hours later.", "+8 hours", "Accepts.", "P1"),
]:
    add(S3, 46, "Add Trade — Date/Time Boundaries", "Edge", label, tin, exp, pri)

# Multi-entry price scenarios
for desc, entries, exp, pri in [
    ("Scale-in lower avg: entries 100+90, qty equal.", "100,90", "Avg entry 95.", "P1"),
    ("Scale-in higher avg: entries 90+100.", "90,100", "Avg entry 95.", "P1"),
    ("Unequal qty weighting: 10@100 + 1@110.", "weighted", "Avg closer to 100.", "P2"),
    ("3 targets at 1R 2R 3R.", "3 TPs", "Planned R reflects nearest or blended — note.", "P2"),
]:
    add(S3, 47, "Add Trade — Multi-Entry/Target Math", "Logic", desc, entries, exp, pri)

# Sources visual per source type
for kind, badge, action in [
    ("backtest session", "Backtest blue", "Dashboard shows session trades"),
    ("prop session", "Prop gold", "Prop metrics if any"),
    ("strategy aggregate", "Strategy", "Multi-session rollup"),
    ("journal account", "Journal green", "Live journal trades"),
    ("journal single entry", "Journal entry", "Single trade scope"),
]:
    add(S4, 48, "Sources — Type Badge & Footer Visuals", "Structure",
         f"Select **{kind}**; inspect footer badge.", kind,
         f"Badge shows {badge}; after Go: {action}.", "P1")

bulk(S4, 48, "Sources — Type Badge & Footer Visuals", "Structure", [
    ("List row: trade count label format.", "—", "e.g. '42 trades' readable.", "P2"),
    ("List row: date range format.", "—", "ISO or locale dates consistent.", "P2"),
    ("List row: market label (Forex, Futures).", "—", "Correct asset class.", "P2"),
    ("Strategy group with 0 children.", "empty strategy", "Empty state or disabled Go.", "P2"),
    ("Select then rapidly click Go 3 times.", "—", "No duplicate navigation or crash.", "P2"),
])

# Unrealistic / abuse inputs across sections
ABUSE = [
    (S1, 49, "Strategy name SQL injection", "'; DROP TABLE--", "Displays as text; no DB error.", "P0"),
    (S1, 49, "Strategy name HTML injection", "<img src=x onerror=alert(1)>", "Escaped; no script run.", "P0"),
    (S2, 49, "Session name script tag", "<script>alert(1)</script>", "Safe display in list.", "P0"),
    (S2, 49, "Capital field paste letters", "abc", "Rejected or ignored.", "P1"),
    (S2, 49, "Capital field paste huge number", "999999999999", "Capped or accepted with warning.", "P2"),
    (S3, 49, "Trade notes HTML paste", "<b>bold</b>", "Plain text or sanitized.", "P1"),
    (S3, 49, "Paste 10000 char note", "10k chars", "Truncated or scroll; no crash.", "P2"),
]
for sec, ph, instr, tin, exp, pri in ABUSE:
    add(sec, ph, "Abuse & Security Inputs", "Extreme", instr, tin, exp, pri)

# Responsive breakpoints
for w in [2560, 1920, 1440, 1280, 1024]:
    add(S1, 50, "Responsive Layout — Strategies", "Structure",
         f"Strategy Bank at viewport **{w}px** width.", f"{w}px",
         "Header, cards/rows, Build Strategy visible without overlap.", "P2")
    add(S2, 50, "Responsive Layout — Sessions", "Structure",
         f"Sessions page at **{w}px**.", f"{w}px", "Stats dashboard + list usable.", "P2")
    add(S3, 50, "Responsive Layout — Trades", "Structure",
         f"Trades table at **{w}px**.", f"{w}px", "Horizontal scroll or column wrap OK.", "P2")
    add(S4, 50, "Responsive Layout — Sources", "Structure",
         f"Sources library at **{w}px**.", f"{w}px", "Tree + list + footer visible.", "P2")

# Session modal toggles disabled-state visual
TOGGLE_TESTS = [
    ("P1 min days OFF → input grayed.", "sessP1MinDaysEnabled=false", "Cannot edit min days.", "P2"),
    ("P2 min days OFF.", "phase 2", "Same.", "P2"),
    ("Consistency rule OFF → % grayed.", "—", "Field disabled.", "P2"),
    ("Max contracts OFF.", "—", "Contracts input disabled.", "P2"),
    ("Max lots OFF.", "—", "Lots input disabled.", "P2"),
    ("MFE/MAE OFF → hours field disabled.", "—", "Grayed out.", "P2"),
    ("Supporting tickers OFF → picker hidden.", "—", "Section collapsed.", "P2"),
    ("Trading costs OFF → per-class overrides hidden.", "—", "Section collapsed.", "P2"),
]
for t in TOGGLE_TESTS:
    add(S2, 51, "Add Session — Toggle Disabled States", "Structure", t[0], t[1], t[2], t[3])

# Strategy bank community tab (disabled) extra checks
bulk(S1, 52, "Strategy Bank — Community Tab (disabled)", "Edge", [
    ("Community tab visible but disabled.", "—", "Grayed; cannot activate.", "P2"),
    ("If Community enabled in future build: sort by Win Rate, Name, P&L.", "—", "Skip if disabled; note N/A.", "P2"),
])

# Trades import/export file matrix
bulk(S3, 53, "Trades — Import/Export File Matrix", "Edge", [
    ("Export → open CSV in Excel.", "—", "Columns readable; dates parse.", "P1"),
    ("Export → re-import same file.", "—", "Duplicates handled (skip/merge) — note behavior.", "P2"),
    ("Import CSV wrong delimiter.", "semicolon file", "Error or partial — note.", "P2"),
    ("Import CSV missing required columns.", "bad headers", "Clear error message.", "P1"),
    ("Import empty CSV.", "0 rows", "No trades added; message shown.", "P2"),
    ("Import 1000-row CSV.", "stress", "Completes without browser freeze.", "P2"),
])

# Sources compare matrix
for n in [2, 3, 4]:
    add(S4, 54, "Sources — Compare Multi-Select", "Configuration",
         f"Compare mode: select **{n}** backtest sessions.", f"{n} sessions",
         f"Compare chip lists {n} sources; metrics update.", "P2")

bulk(S4, 55, "Sources — Cancel vs Go State Machine", "Logic", [
    ("Open library → select A → Cancel.", "—", "Still on previous source.", "P0"),
    ("Open → select A → Go → reopen → select B → Cancel.", "—", "Still on A.", "P0"),
    ("Open → select A → Go → reopen → select B → Go.", "—", "Now on B.", "P0"),
    ("Partial strategy children → Cancel.", "—", "Previous child selection restored.", "P1"),
    ("Partial strategy children → Go.", "—", "Dashboard scoped to partial set.", "P1"),
])

def build():
    wb = Workbook()
    ws_i = wb.active
    ws_i.title = "Instructions"
    ws_i["A1"] = "TALARIA DASHBOARD — EXHAUSTIVE UAT TRACKER"
    ws_i["A1"].font = Font(bold=True, size=14, color="1F4E79")

    lines = [
        ("", ""),
        ("Scope", "Every field, configuration, validation path, visual/layout check, normal + edge + extreme inputs."),
        ("Testers", "2 people — Tester A (Sections 1–2), Tester B (Sections 3–4), then swap."),
        ("Categories", "Normal | Edge | Extreme | Configuration | Input matrix | Logic | Structure"),
        ("Columns", "Test Input = exact value to enter; Expected Result = pass criteria."),
        ("Result", "☑ Pass | ☐ Fail | ⚠ Partial | — Not tested"),
        ("P0", "Launch blocker — stop and report immediately."),
        ("", f"Total steps: {len(STEPS)} across 4 sections"),
        ("1 — Strategies", "Phases 1–13, 44, 49, 50, 52: Bank + Builder + abuse/responsive"),
        ("2 — Sessions", "Phases 14–24, 42–43, 49–51: List + Modal + mode matrix + toggles"),
        ("3 — Trades", "Phases 25–34, 45–47, 49–50, 53: Table + validations + import/export"),
        ("4 — Sources", "Phases 35–41, 48, 49–50, 54–55: Library + compare + state machine"),
    ]
    ws_i.column_dimensions["A"].width = 18
    ws_i.column_dimensions["B"].width = 95
    for r, (a, b) in enumerate(lines, 2):
        ws_i.cell(r, 1, a).font = Font(bold=bool(a))
        ws_i.cell(r, 2, b).alignment = Alignment(wrap_text=True)

    headers = [
        "Step #", "Section", "Phase", "Phase Name", "Category", "Priority",
        "What to Do", "Test Input", "Expected Result",
        "Tester A", "Tester B", "Issue?", "Comments",
    ]
    ws = wb.create_sheet("QA Tracker")
    fill = PatternFill("solid", fgColor="1F4E79")
    hf = Font(bold=True, color="FFFFFF", size=9)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = fill
        cell.font = hf
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    colors = {"1 — Strategies": "E3F2FD", "2 — Sessions": "E8F5E9", "3 — Trades": "FFF8E1", "4 — Sources": "F3E5F5"}
    cur_sec = None
    cur_fill = None

    for i, s in enumerate(STEPS, 1):
        r = i + 1
        ws.cell(r, 1, i)
        ws.cell(r, 2, s["section"])
        ws.cell(r, 3, s["phase"])
        ws.cell(r, 4, s["phase_name"])
        ws.cell(r, 5, s["category"])
        ws.cell(r, 6, s["priority"])
        ws.cell(r, 7, s["instruction"])
        ws.cell(r, 8, s["test_input"])
        ws.cell(r, 9, s["expected"])
        ws.cell(r, 10, "— Not tested")
        ws.cell(r, 11, "— Not tested")
        ws.cell(r, 12, "— Not tested")
        for c in range(1, 13):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        if s["section"] != cur_sec:
            cur_sec = s["section"]
            cur_fill = PatternFill("solid", fgColor=colors.get(cur_sec, "FFFFFF"))
        for c in range(1, 5):
            ws.cell(r, c).fill = cur_fill

    last = len(STEPS) + 1
    dv = DataValidation(type="list", formula1=RESULT, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"J2:K{last}")
    dv2 = DataValidation(type="list", formula1=ISSUE, allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add(f"L2:L{last}")

    widths = [7, 14, 6, 26, 14, 7, 48, 22, 38, 12, 12, 9, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"

    # Summary
    ws_s = wb.create_sheet("Summary")
    for c, h in enumerate(["Phase", "Phase Name", "Section", "Steps", "P0", "A Pass", "B Pass", "Issues"], 1):
        ws_s.cell(1, c, h).fill = fill
        ws_s.cell(1, c, h).font = hf
    phases = []
    seen = set()
    for s in STEPS:
        if s["phase"] not in seen:
            seen.add(s["phase"])
            phases.append((s["phase"], s["phase_name"], s["section"]))
    tr = "'QA Tracker'"
    for i, (p, name, sec) in enumerate(phases, 2):
        ws_s.cell(i, 1, p)
        ws_s.cell(i, 2, name)
        ws_s.cell(i, 3, sec)
        ws_s.cell(i, 4, f'=COUNTIFS({tr}!C:C,{p})')
        ws_s.cell(i, 5, f'=COUNTIFS({tr}!C:C,{p},{tr}!F:F,"P0")')
        ws_s.cell(i, 6, f'=COUNTIFS({tr}!C:C,{p},{tr}!J:J,"☑ Pass")')
        ws_s.cell(i, 7, f'=COUNTIFS({tr}!C:C,{p},{tr}!I:I,"☑ Pass")')
        ws_s.cell(i, 8, f'=COUNTIFS({tr}!C:C,{p},{tr}!L:L,"☑ Yes")')

    # Matrix reference sheet
    ws_m = wb.create_sheet("Test Matrix Reference")
    ws_m["A1"] = "Built-in limits & validation reference (from codebase)"
    ws_m["A1"].font = Font(bold=True, size=12)
    ref = [
        ("", ""),
        ("STRATEGY BUILDER", ""),
        ("Strategy name", "Required, max 80 chars"),
        ("Description", "Optional, max 500 chars"),
        ("Tags", "Max 10"),
        ("Markets", "Required, at least 1 (Forex/Futures/Crypto/Stocks)"),
        ("Instruments", "Max 10"),
        ("Timeframes", "Required, max 6; custom TF supported"),
        ("Cover images", "Max 6"),
        ("Condition title", "~70 chars max"),
        ("", ""),
        ("SESSION MODAL", ""),
        ("Session name", "Required"),
        ("Primary tickers", "Required, max = plan (default 5)"),
        ("Supporting tickers", "Max = plan supporting cap"),
        ("Start/end dates", "Required for range mode"),
        ("Capital", "Required, > 0"),
        ("Prop firm", "Disabled for Futures asset class tab"),
        ("", ""),
        ("ADD TRADE VALIDATION", ""),
        ("Stop loss", "Required, > 0; below entries (long) / above (short)"),
        ("Targets", "Above entries (long) / below (short)"),
        ("Exit size", "Cannot exceed entry size"),
        ("Excursion", "High > low; entry inside range; max ~500R"),
        ("Extended excursion", "Requires closed trade + post range + integer window > 0"),
        ("Journal trade", "Strategy/setup required (non-live journal)"),
        ("", ""),
        ("SOURCES", ""),
        ("Filter combos", "Mode (Std/Prop) × Status (All/Not Started/Active/Completed)"),
        ("Strategy selection", "Partial child checkboxes supported"),
    ]
    for r, (a, b) in enumerate(ref, 2):
        ws_m.cell(r, 1, a).font = Font(bold=bool(a and not b))
        ws_m.cell(r, 2, b)
    ws_m.column_dimensions["A"].width = 28
    ws_m.column_dimensions["B"].width = 55

    # Per-section sheets
    for sec, title in [("1 — Strategies", "S1 Strategies"), ("2 — Sessions", "S2 Sessions"),
                       ("3 — Trades", "S3 Trades"), ("4 — Sources", "S4 Sources")]:
        wsp = wb.create_sheet(title)
        for c, h in enumerate(headers, 1):
            wsp.cell(1, c, h).fill = fill
            wsp.cell(1, c, h).font = hf
        row = 2
        for i, s in enumerate(STEPS, 1):
            if s["section"] != sec:
                continue
            wsp.cell(row, 1, i)
            for c, key in enumerate(["section", "phase", "phase_name", "category", "priority",
                                     "instruction", "test_input", "expected"], 2):
                wsp.cell(row, c, s[key])
            wsp.cell(row, 10, "— Not tested")
            wsp.cell(row, 11, "— Not tested")
            wsp.cell(row, 12, "— Not tested")
            row += 1
        if row > 2:
            dva = DataValidation(type="list", formula1=RESULT, allow_blank=True)
            wsp.add_data_validation(dva)
            dva.add(f"J2:K{row-1}")
            dvb = DataValidation(type="list", formula1=ISSUE, allow_blank=True)
            wsp.add_data_validation(dvb)
            dvb.add(f"L2:L{row-1}")
        wsp.freeze_panes = "A2"
        for i, w in enumerate(widths, 1):
            wsp.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")
    print(f"Steps: {len(STEPS)} | Phases: {len(phases)}")


if __name__ == "__main__":
    build()
